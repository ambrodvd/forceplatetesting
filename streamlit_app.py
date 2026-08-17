# -*- coding: utf-8 -*-
"""
Force Plate Test Report — app Streamlit (versione a file singolo)

Carica gli export XLSX di ForceMate/ForceDecks (IMTP, SJ, CMJ, CMJ RE),
calcola automaticamente medie, T-score rispetto alla popolazione di
riferimento e produce un profilo di forza con grafici Plotly e report HTML interattivo.

Struttura del file:
  PARTE 1 — Costanti e dati di popolazione
  PARTE 2 — Caricamento file (sidebar upload)
  PARTE 3 — Lettura/parsing dei file XLSX
  PARTE 4 — Analisi dati e confronto con la popolazione
  PARTE 5 — Report live (UI a schede)
  PARTE 6 — Report scaricabile (HTML interattivo)
"""

from __future__ import annotations

import io
import math
import re
import textwrap
import html as _html
import datetime as dt
from dataclasses import dataclass, field

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl
from fpdf import FPDF
from fpdf.fonts import FontFace


# ============================================================================
# PARTE 1 — COSTANTI E DATI DI POPOLAZIONE
# ============================================================================
# Equivalente ai fogli "DATI POP" (norme di popolazione) e "DATI TEST"
# (definizione delle metriche) del Google Sheet originale.

CATEGORIES = ["ISOMETRIC PULL TEST", "SQUAT JUMP TEST", "COUNTERMOVEMENT JUMP TEST", "COUNTERMOVEMENT JUMP REBOUND TEST"]

# Qualità fisica misurata da ciascun test: il nome del protocollo dice COME si
# è misurato, l'etichetta dice COSA significa. Serve all'atleta, che non ha
# motivo di sapere cosa sia un "countermovement jump rebound".
CATEGORY_QUALITY = {
    "ISOMETRIC PULL TEST": "Forza massima",
    "SQUAT JUMP TEST": "Potenza",
    "COUNTERMOVEMENT JUMP TEST": "Esplosività",
    "COUNTERMOVEMENT JUMP REBOUND TEST": "Reattività",
}

# Bande di valutazione basate sul T-score (media 50, deviazione standard 10)
BANDS = [
    (float("-inf"), 30, "MOLTO SOTTO LA MEDIA", "#B00020"),
    (30, 40, "SOTTO LA MEDIA", "#E4572E"),
    (40, 60, "NELLA MEDIA", "#F2A007"),
    (60, 70, "SOPRA LA MEDIA", "#8BC34A"),
    (70, 80, "BUONO", "#4CAF50"),
    (80, float("inf"), "OTTIMO", "#2E7D32"),
]


def banda_da_tscore(t):
    if t is None:
        return None, None
    for lo, hi, label, color in BANDS:
        if lo <= t < hi:
            return label, color
    return "OTTIMO", "#2E7D32"


# Soglie di lettura degli indici-rapporto (DSI, EUR): delimitano le tre zone
# di profilo mostrate nei grafici. Valori di DEFAULT, modificabili nella
# scheda "⚙️ Costanti" (a runtime si usa sempre st.session_state["idx_thr"]).
#   DSI  0.60 / 0.80  — Sheppard & Chapman (2011), via Science for Sport
#   EUR  1.00 / 1.10  — McGuigan et al. (2006), via VALD / Clubb
# NB: le soglie EUR di letteratura cadono praticamente sulla media di
# popolazione qui usata (1.108 U / 1.091 D): con i default un atleta "medio"
# finisce sul confine superiore. Da ricalibrare sui propri dati.
DEFAULT_INDEX_THRESHOLDS = {"dsi": (0.60, 0.80), "eur": (1.00, 1.10)}

# Etichette descrittive (profilo dell'atleta, NON indicazione di allenamento)
# e colori delle tre zone, in ordine basso / medio / alto.
INDEX_ZONE_LABELS = {
    "dsi": ("Atleta più forte che esplosivo", "Atleta bilanciato", "Atleta più esplosivo che forte"),
    "eur": ("Poco contributo elastico", "Contributo elastico medio", "Buon contributo elastico"),
}
INDEX_ZONE_COLORS = ("#FFB74D", "#81C784", "#4FC3F7")


def zona_da_indice(key, value, thr_low, thr_high):
    """Zona di profilo (etichetta, colore) per un indice-rapporto."""
    labels = INDEX_ZONE_LABELS.get(key)
    if value is None or labels is None:
        return None, None
    i = 0 if value < thr_low else (2 if value > thr_high else 1)
    return labels[i], INDEX_ZONE_COLORS[i]


# sd sempre positiva: la direzione "minore è meglio" è gestita dal flag
# lower_is_better nella metrica, non dal segno della deviazione standard.
#
# Questi sono i valori DI DEFAULT: l'utente può visionarli, modificarli,
# scaricarli e ricaricarli da un file Excel nella scheda "⚙️ Costanti"
# (vedi PARTE 4bis). A runtime l'app usa sempre st.session_state["pop"],
# inizializzato da questo dizionario alla prima esecuzione.
#
# NB: EUR è definito come rapporto puro CMJ height / SJ height (non come
# differenza percentuale), coerente con la tabella costanti fornita
# dall'utente (~1.11 uomini, ~1.09 donne).
DEFAULT_POP = {
    "imtp_peak_force":        dict(mean_m=2606.8,      sd_m=646.1163194,  mean_f=1575.0,      sd_f=386.145544),
    "imtp_rel_peak_force":    dict(mean_m=34.56,        sd_m=5.27,         mean_f=34.56,        sd_f=5.27),
    "sj_mean_power":          dict(mean_m=1180.0,       sd_m=414.1638849,  mean_f=823.0,        sd_f=240.2447942),
    "sj_height":              dict(mean_m=26.9,         sd_m=6.57,         mean_f=19.35,        sd_f=5.51),
    "sj_contraction_time":    dict(mean_m=0.445,        sd_m=0.127835797,  mean_f=0.46,         sd_f=0.137032268),
    "cmj_height":             dict(mean_m=30.01,        sd_m=6.5,          mean_f=20.91,        sd_f=5.84),
    "mrsi_cmj":               dict(mean_m=0.419,        sd_m=0.098531539,  mean_f=0.308,        sd_f=0.093831019),
    "dsi":                    dict(mean_m=0.7,          sd_m=0.074074074,  mean_f=0.7,          sd_f=0.074074074),
    "eur":                    dict(mean_m=1.107708605,  sd_m=0.038034873,  mean_f=1.090563116,  sd_f=0.02656191),
    "cmj_re_rebound_height":  dict(mean_m=38.5,         sd_m=5.5996817,    mean_f=38.5,         sd_f=5.5996817),
    "cmj_re_contact_time":    dict(mean_m=0.25,         sd_m=0.148005087,  mean_f=0.25,         sd_f=0.148005087),
    "cmj_re_rebound_impulse": dict(mean_m=541.5,        sd_m=53.82161458,  mean_f=541.5,        sd_f=53.82161458),
    "mrsi_cmj_re":            dict(mean_m=1.37,         sd_m=0.331705729,  mean_f=1.37,         sd_f=0.331705729),
}

# Etichetta e unità di misura di ciascuna costante, usate sia per la
# tabella a schermo sia per il file Excel scaricabile/ricaricabile.
# Le etichette ricalcano quelle del file "DATO" fornito dall'utente, in
# modo che un file con quella struttura venga riconosciuto automaticamente.
CONST_LABELS = {
    "imtp_peak_force":        ("IMTP abs", "N", "ISOMETRIC PULL TEST"),
    "imtp_rel_peak_force":    ("IMTP rel", "N/kg", "ISOMETRIC PULL TEST"),
    "sj_mean_power":          ("SJ mean power", "W", "SQUAT JUMP TEST"),
    "sj_height":               ("SJ height", "cm", "SQUAT JUMP TEST"),
    "sj_contraction_time":     ("SJ contraction time", "s", "SQUAT JUMP TEST"),
    "cmj_height":              ("CMJ height", "cm", "COUNTERMOVEMENT JUMP TEST"),
    "mrsi_cmj":                ("Mrsi-CMJ", "m/s", "COUNTERMOVEMENT JUMP TEST"),
    "cmj_re_rebound_height":   ("CMJ RE Rebound Jump height", "cm", "COUNTERMOVEMENT JUMP REBOUND TEST"),
    "cmj_re_contact_time":     ("CMJ RE Contact Time", "s", "COUNTERMOVEMENT JUMP REBOUND TEST"),
    "cmj_re_rebound_impulse":  ("CMJ RE propulsive impulse", "N\u00b7s", "COUNTERMOVEMENT JUMP REBOUND TEST"),
    "mrsi_cmj_re":             ("mRSI-CMJ RE", "m/s", "COUNTERMOVEMENT JUMP REBOUND TEST"),
    "dsi":                     ("DSI", "", "INDICI"),
    "eur":                     ("EUR", "", "INDICI"),
}


def costanti_dataframe(pop_dict):
    """Costruisce il DataFrame mostrato/editato nella scheda Costanti."""
    rows = []
    for key, (label, unit, categoria) in CONST_LABELS.items():
        c = pop_dict[key]
        rows.append({
            "Categoria": categoria, "Costante": label, "Unità": unit,
            "Media Uomini": c["mean_m"], "Dev.Std Uomini": c["sd_m"],
            "Media Donne": c["mean_f"], "Dev.Std Donne": c["sd_f"],
        })
    return pd.DataFrame(rows)


def genera_costanti_xlsx(pop_dict):
    """Esporta le costanti nello stesso formato del file caricabile:
    DATO | UdM | MEDIA | DEV ST | MEDIA2 | DEV ST3, con riga SESSO."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TABELLA metrics data"
    ws.append(["DATO", "UdM", "MEDIA", "DEV ST", "MEDIA2", "DEV ST3"])
    ws.append(["SESSO", None, "UOMO", "UOMO", "DONNA", "DONNA"])
    for key, (label, unit, _categoria) in CONST_LABELS.items():
        c = pop_dict[key]
        ws.append([label, unit, c["mean_m"], c["sd_m"], c["mean_f"], c["sd_f"]])
    for col in ("A", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["A"].width = 30
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_constants_workbook(file_like):
    """Legge un file Excel nel formato DATO/UdM/MEDIA/DEV ST/MEDIA2/DEV ST3
    e restituisce (updates, unmatched): un dict {key: {mean_m, sd_m, mean_f,
    sd_f}} per le costanti riconosciute, e la lista di righe non riconosciute."""
    wb = openpyxl.load_workbook(file_like, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    label_to_key = {label.strip().lower(): key for key, (label, _u, _c) in CONST_LABELS.items()}

    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] and str(row[0]).strip().upper() == "DATO":
            header_idx = i
            break
    if header_idx is None:
        return {}, []

    def to_num(v):
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            try:
                return float(v.replace(",", "."))
            except ValueError:
                return None
        return None

    updates, unmatched = {}, []
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if not name or name.upper() == "SESSO":
            continue
        if "DATI DA LETTERATURA" in name.upper():
            break
        key = label_to_key.get(name.lower())
        mean_m = to_num(row[2]) if len(row) > 2 else None
        sd_m = to_num(row[3]) if len(row) > 3 else None
        mean_f = to_num(row[4]) if len(row) > 4 else None
        sd_f = to_num(row[5]) if len(row) > 5 else None
        if key is None:
            unmatched.append(name)
            continue
        if None not in (mean_m, sd_m, mean_f, sd_f):
            updates[key] = dict(mean_m=mean_m, sd_m=abs(sd_m), mean_f=mean_f, sd_f=abs(sd_f))
    return updates, unmatched


def _safe_div(a, b):
    try:
        if a is None or b in (None, 0):
            return None
        return float(a) / float(b)
    except (TypeError, ValueError):
        return None


# Definizione metriche: key, label, category, jump_type (o None se derivata),
# raw_var (nome variabile grezza ForceMate) o derive (funzione per-rep),
# unit, pop_key (chiave in POP o None), lower_is_better, kind
#   kind: "score" (T-score da distribuzione pooled), "score_single"
#         (indice aggregato cross-test, es. DSI/EUR — letto per zona, non
#         per T-score), "info" (valore mostrato senza confronto)
METRICS = [
    dict(key="imtp_peak_force", label="IMTP Peak Force", category="ISOMETRIC PULL TEST",
         jump_type="imtp", raw_var="peak force", unit="N",
         pop_key="imtp_peak_force", lower_is_better=False, kind="score"),
    dict(key="imtp_rel_peak_force", label="IMTP Rel Peak Force", category="ISOMETRIC PULL TEST",
         jump_type="imtp", raw_var=None,
         derive=lambda rep: _safe_div(rep.get("peak force"), rep.get("body mass")),
         unit="N/kg", pop_key="imtp_rel_peak_force", lower_is_better=False, kind="score"),

    dict(key="sj_mean_power", label="SJ Mean Power", category="SQUAT JUMP TEST",
         jump_type="sj", raw_var="avg. propulsive power", unit="W",
         pop_key="sj_mean_power", lower_is_better=False, kind="score"),
    dict(key="sj_height", label="SJ Height", category="SQUAT JUMP TEST",
         jump_type="sj", raw_var="jump height ft", unit="cm",
         pop_key="sj_height", lower_is_better=False, kind="score"),
    dict(key="sj_contraction_time", label="SJ Contraction Time", category="SQUAT JUMP TEST",
         jump_type="sj", raw_var="time to takeoff", unit="s",
         pop_key="sj_contraction_time", lower_is_better=True, kind="score"),
    dict(key="sj_net_impulse", label="SJ Net Impulse", category="SQUAT JUMP TEST",
         jump_type="sj", raw_var="net impulse", unit="N\u00b7s",
         pop_key=None, lower_is_better=False, kind="info"),
    dict(key="sj_net_rel_impulse", label="SJ Net Rel Impulse", category="SQUAT JUMP TEST",
         jump_type="sj", raw_var=None,
         derive=lambda rep: _safe_div(rep.get("net impulse"), rep.get("body mass")),
         unit="N\u00b7s/kg", pop_key=None, lower_is_better=False, kind="info"),

    dict(key="cmj_net_impulse", label="CMJ Net Impulse", category="COUNTERMOVEMENT JUMP TEST",
         jump_type="cmj", raw_var="net impulse", unit="N\u00b7s",
         pop_key=None, lower_is_better=False, kind="info"),
    dict(key="cmj_net_rel_impulse", label="CMJ Net Rel Impulse", category="COUNTERMOVEMENT JUMP TEST",
         jump_type="cmj", raw_var=None,
         derive=lambda rep: _safe_div(rep.get("net impulse"), rep.get("body mass")),
         unit="N\u00b7s/kg", pop_key=None, lower_is_better=False, kind="info"),
    dict(key="cmj_contraction_time", label="CMJ Contraction Time", category="COUNTERMOVEMENT JUMP TEST",
         jump_type="cmj", raw_var="time to takeoff", unit="s",
         pop_key=None, lower_is_better=True, kind="info"),
    dict(key="cmj_height", label="CMJ Height", category="COUNTERMOVEMENT JUMP TEST",
         jump_type="cmj", raw_var="jump height ft", unit="cm",
         pop_key="cmj_height", lower_is_better=False, kind="score"),
    dict(key="mrsi_cmj", label="mRSI-CMJ", category="COUNTERMOVEMENT JUMP TEST",
         jump_type="cmj", raw_var="rsi modified", unit="m/s",
         pop_key="mrsi_cmj", lower_is_better=False, kind="score"),
    dict(key="cmj_peak_force", label="CMJ Peak Force", category="COUNTERMOVEMENT JUMP TEST",
         jump_type="cmj", raw_var="peak propulsive force", unit="N",
         pop_key=None, lower_is_better=False, kind="info"),

    dict(key="cmj_re_initial_height", label="CMJ RE Jump Height (iniziale)", category="COUNTERMOVEMENT JUMP REBOUND TEST",
         jump_type="cmrj", raw_var="jump height ft", unit="cm",
         pop_key=None, lower_is_better=False, kind="info"),
    dict(key="cmj_re_rebound_height", label="CMJ RE Rebound Jump Height", category="COUNTERMOVEMENT JUMP REBOUND TEST",
         jump_type="cmrj", raw_var="rebound jump height ft", unit="cm",
         pop_key="cmj_re_rebound_height", lower_is_better=False, kind="score"),
    dict(key="cmj_re_contact_time", label="CMJ RE Contact Time", category="COUNTERMOVEMENT JUMP REBOUND TEST",
         jump_type="cmrj", raw_var="rebound contact time", unit="s",
         pop_key="cmj_re_contact_time", lower_is_better=True, kind="score"),
    dict(key="cmj_re_rebound_impulse", label="CMJ RE Rebound Propulsive Impulse", category="COUNTERMOVEMENT JUMP REBOUND TEST",
         jump_type="cmrj", raw_var="rebound propulsive impulse", unit="N\u00b7s",
         pop_key="cmj_re_rebound_impulse", lower_is_better=False, kind="score"),
    dict(key="mrsi_cmj_re", label="mRSI-CMJ RE", category="COUNTERMOVEMENT JUMP REBOUND TEST",
         jump_type="cmrj", raw_var="rebound rsi modified", unit="m/s",
         pop_key="mrsi_cmj_re", lower_is_better=False, kind="score"),
    dict(key="unbalanced_landing_raw", label="Braking Impulse Sym. Index (CMJ RE)", category="COUNTERMOVEMENT JUMP REBOUND TEST",
         jump_type="cmrj", raw_var="braking impulse sym. index", unit="%",
         pop_key=None, lower_is_better=False, kind="info"),

    # Indici derivati da medie aggregate cross-test (non per-rep)
    dict(key="dsi", label="DSI (Dynamic Strength Index - Peak force CMJ/Peak force IMTP)", category="INDICI",
         jump_type=None, raw_var=None, unit="",
         pop_key="dsi", lower_is_better=False, kind="score_single"),
    dict(key="eur", label="EUR (Eccentric Utilisation Ratio - CMJ Height / SJ Height)", category="INDICI",
         jump_type=None, raw_var=None, unit="",
         pop_key="eur", lower_is_better=False, kind="score_single"),
]

# Controlli a soglia (equivalenti a "Jump to rebound Ratio", "CMJ to CMJ RE
# check", "Unbalanced landing check" del foglio originale, più il controllo
# sul tempo di contatto). Riguardano esclusivamente il CMJ Rebound, quindi
# vengono mostrati insieme alla categoria COUNTERMOVEMENT JUMP REBOUND TEST
# nel Dettaglio Test.
# scale/decimals/suffix controllano solo la formattazione a schermo: il
# "value" resta sempre nell'unità nativa del calcolo (rapporto puro per i
# primi tre, secondi per il tempo di contatto).
CHECKS = [
    dict(key="jump_to_rebound_ratio", label="Jump to Rebound Ratio",
         desc="Rapporto tra altezza del rimbalzo e altezza del salto iniziale nel CMJ RE.",
         threshold=0.60, direction="min", scale=100, decimals=1, suffix="%"),
    dict(key="cmj_to_cmjre_check", label="CMJ to CMJ RE Check",
         desc="Rapporto tra l'altezza del salto iniziale nel CMJ RE e l'altezza del CMJ standard.",
         threshold=0.85, direction="min", scale=100, decimals=1, suffix="%"),
    dict(key="unbalanced_landing_check", label="Unbalanced Landing Check",
         desc="Indice di simmetria dell'impulso frenante in atterraggio (CMJ RE). Valori assoluti alti indicano un atterraggio sbilanciato.",
         threshold=0.50, direction="max", scale=100, decimals=1, suffix="%"),
    dict(key="cmj_re_contact_time_check", label="CMJ RE Contact Time Check",
         desc="Tempo di contatto nel rimbalzo del CMJ RE: un tempo di contatto troppo lungo indica un rimbalzo poco reattivo.",
         threshold=0.250, direction="max", scale=1000, decimals=0, suffix=" ms"),
]

JUMP_TYPE_LABELS = {"imtp": "IMTP", "sj": "Squat Jump", "cmj": "CMJ", "cmrj": "CMJ Rebound"}


# ============================================================================
# PARTE 1bis — CATALOGO METRICHE EXTRA (ricerca libera, senza dato di
# popolazione/T-score)
# ============================================================================
# L'app mostra di default solo un set curato di metriche (vedi METRICS
# sopra), ciascuna con confronto di popolazione. Questi elenchi coprono
# TUTTE le variabili che ForceMate/ForceDecks può esportare per ciascun
# tipo di test, così l'utente può cercare e aggiungere qualsiasi altra
# metrica dalla scheda Dettaglio Test — senza T-score (nessun dato di
# popolazione disponibile per queste), solo media/dev.std/CV% per
# ripetizione, con le stesse esclusioni già impostate sui salti.
# Nomi già lowercase per combaciare con le chiavi salvate in rep["vars"]
# (il parsing normalizza sempre con .strip().lower()).
_SJ_METRICS_CATALOG = [
    'avg landing rfd', 'avg landing rfd sym. index', 'avg propulsive force sym. index',
    'avg propulsive power sym. index', 'avg. propulsive force', 'avg. propulsive power',
    'avg. propulsive velocity', 'avg. rfd', 'avg. rfd sym. index', 'body mass', 'body weight',
    'body weight sd', 'contact time', 'flight threshold', 'flight time', 'force peak power',
    'initiation threshold', 'jump height ft', 'jump height ni', 'jump momentum', 'jump start time',
    'jump threshold time', 'landing peak force time', 'landing rfd 0-20ms',
    'landing rfd 0-20ms sym. index', 'landing rfd 0-40ms', 'landing rfd 0-40ms sym. index',
    'landing rfd 0-60ms', 'landing rfd 0-60ms sym. index', 'landing rfd 0-80ms',
    'landing rfd 0-80ms sym. index', 'landing time', 'left avg landing rfd',
    'left avg propulsive force', 'left avg propulsive power', 'left avg rfd',
    'left landing rfd 0-20ms', 'left landing rfd 0-40ms', 'left landing rfd 0-60ms',
    'left landing rfd 0-80ms', 'left net impulse', 'left p1 impulse', 'left p2 impulse',
    'left peak force', 'left peak landing force', 'left peak propulsive force',
    'left peak propulsive power', 'left propulsive impulse', 'left propulsive rfd',
    'left propulsive work', 'left time to peak force', 'left time to peak landing force',
    'left time to peak power', 'left time to peak propulsive force',
    'left time to peak propulsive power', 'net impulse', 'net impulse sym. index', 'p1 avg force',
    'p1 avg power', 'p1 avg velocity', 'p1 duration', 'p1 impulse', 'p1 impulse sym. index',
    'p1 p2 duration ratio', 'p1 p2 force ratio', 'p1 p2 power ratio', 'p1 p2 velocity ratio',
    'p1 peak force', 'p1 peak power', 'p1 peak velocity', 'p2 avg force', 'p2 avg power',
    'p2 avg velocity', 'p2 duration', 'p2 impulse', 'p2 impulse sym. index', 'p2 peak force',
    'p2 peak power', 'p2 peak velocity', 'peak force', 'peak force sym. index', 'peak force time',
    'peak landing force', 'peak landing force sym. index', 'peak power', 'peak propulsive force',
    'peak propulsive force sym. index', 'peak propulsive power',
    'peak propulsive power sym. index', 'peak propulsive velocity', 'peak velocity',
    'propulsive duration', 'propulsive impulse', 'propulsive impulse sym. index', 'propulsive rfd',
    'propulsive rfd sym. index', 'propulsive start time', 'propulsive work',
    'propulsive work sym. index', 'rel. avg. propulsive force', 'rel. propulsive impulse',
    'relative peak force', 'relative peak landing force', 'relative peak power',
    'right avg landing rfd', 'right avg propulsive force', 'right avg propulsive power',
    'right avg rfd', 'right landing rfd 0-20ms', 'right landing rfd 0-40ms',
    'right landing rfd 0-60ms', 'right landing rfd 0-80ms', 'right net impulse',
    'right p1 impulse', 'right p2 impulse', 'right peak force', 'right peak landing force',
    'right peak propulsive force', 'right peak propulsive power', 'right propulsive impulse',
    'right propulsive rfd', 'right propulsive work', 'right time to peak force',
    'right time to peak landing force', 'right time to peak power',
    'right time to peak propulsive force', 'right time to peak propulsive power', 'rsi',
    'rsi modified', 'takeoff time', 'takeoff velocity', 'time to peak force',
    'time to peak landing force', 'time to peak landing force sym. index', 'time to peak power',
    'time to peak power sym. index', 'time to peak propulsive force',
    'time to peak propulsive force sym. index', 'time to peak propulsive power',
    'time to peak propulsive power sym. index', 'time to peak si', 'time to takeoff',
    'velocity peak power',
]

_CMJ_EXTRA_METRICS_CATALOG = [
    'avg braking force sym. index', 'avg braking power sym. index', 'avg eccentric force',
    'avg eccentric force sym. index', 'avg eccentric power', 'avg eccentric power sym. index',
    'avg eccentric velocity', 'avg. braking force', 'avg. braking power', 'avg. braking velocity',
    'braking duration', 'braking end time', 'braking impulse', 'braking impulse sym. index',
    'braking rfd', 'braking rfd sym. index', 'braking work', 'braking work sym. index',
    'decel rfd', 'decel rfd sym. index', 'displacement depth', 'eccentric impulse',
    'eccentric impulse sym. index', 'eccentric rfd', 'eccentric rfd sym. index', 'eccentric work',
    'eccentric work sym. index', 'force at min displacement', 'left avg braking force',
    'left avg braking power', 'left avg eccentric force', 'left avg eccentric power',
    'left braking impulse', 'left braking rfd', 'left braking work', 'left decel rfd',
    'left eccentric impulse', 'left eccentric rfd', 'left eccentric work',
    'left peak braking force', 'left peak braking power', 'left peak eccentric force',
    'left peak eccentric power', 'left time to peak braking force',
    'left time to peak braking power', 'left time to peak eccentric force',
    'left time to peak eccentric power', 'min braking velocity', 'min eccentric velocity',
    'min unweight force', 'peak braking force', 'peak braking force sym. index',
    'peak braking power', 'peak braking power sym. index', 'peak eccentric force',
    'peak eccentric force sym. index', 'peak eccentric power', 'peak eccentric power sym. index',
    'rel. min unweight force', 'relative force at min displacement', 'right avg braking force',
    'right avg braking power', 'right avg eccentric force', 'right avg eccentric power',
    'right braking impulse', 'right braking rfd', 'right braking work', 'right decel rfd',
    'right eccentric impulse', 'right eccentric rfd', 'right eccentric work',
    'right peak braking force', 'right peak braking power', 'right peak eccentric force',
    'right peak eccentric power', 'right time to peak braking force',
    'right time to peak braking power', 'right time to peak eccentric force',
    'right time to peak eccentric power', 'rsi exponential', 'time to peak braking force',
    'time to peak braking force sym. index', 'time to peak braking power',
    'time to peak braking power sym. index', 'time to peak eccentric force',
    'time to peak eccentric force sym. index', 'time to peak eccentric power',
    'time to peak eccentric power sym. index', 'unweighted duration', 'with armswing',
]

_CMRJ_EXTRA_METRICS_CATALOG = [
    'left rebound avg braking force', 'left rebound avg braking power',
    'left rebound avg propulsive force', 'left rebound avg propulsive power',
    'left rebound braking impulse', 'left rebound braking rfd', 'left rebound braking work',
    'left rebound p1 impulse', 'left rebound p2 impulse', 'left rebound peak braking force',
    'left rebound peak braking power', 'left rebound peak propulsive force',
    'left rebound peak propulsive power', 'left rebound propulsive impulse',
    'left rebound propulsive rfd', 'left rebound propulsive work',
    'left rebound time to peak braking force', 'left rebound time to peak braking power',
    'left rebound time to peak propulsive force', 'left rebound time to peak propulsive power',
    'rebound avg braking force', 'rebound avg braking force symmetry index',
    'rebound avg braking power', 'rebound avg braking power symmetry index',
    'rebound avg braking velocity', 'rebound avg propulsive force',
    'rebound avg propulsive force symmetry index', 'rebound avg propulsive power',
    'rebound avg propulsive power symmetry index', 'rebound avg propulsive velocity',
    'rebound braking impulse', 'rebound braking impulse symmetry index', 'rebound braking rfd',
    'rebound braking rfd symmetry index', 'rebound braking work',
    'rebound braking work symmetry index', 'rebound contact time', 'rebound flight time',
    'rebound force at min displacement', 'rebound jump height ft', 'rebound jump momentum',
    'rebound min braking velocity', 'rebound p1 avg force', 'rebound p1 avg power',
    'rebound p1 avg velocity', 'rebound p1 duration', 'rebound p1 impulse',
    'rebound p1 impulse symmetry index', 'rebound p1 p2 duration ratio',
    'rebound p1 p2 force ratio', 'rebound p1 p2 power ratio', 'rebound p1 p2 velocity ratio',
    'rebound p1 peak force', 'rebound p1 peak power', 'rebound p1 peak velocity',
    'rebound p2 avg force', 'rebound p2 avg power', 'rebound p2 avg velocity',
    'rebound p2 duration', 'rebound p2 impulse', 'rebound p2 impulse symmetry index',
    'rebound p2 peak force', 'rebound p2 peak power', 'rebound p2 peak velocity',
    'rebound peak braking force', 'rebound peak braking force symmetry index',
    'rebound peak braking power', 'rebound peak braking power symmetry index',
    'rebound peak propulsive force', 'rebound peak propulsive force symmetry index',
    'rebound peak propulsive power', 'rebound peak propulsive power symmetry index',
    'rebound peak propulsive velocity', 'rebound propulsive impulse',
    'rebound propulsive impulse symmetry index', 'rebound propulsive rfd',
    'rebound propulsive rfd symmetry index', 'rebound propulsive work',
    'rebound propulsive work symmetry index', 'rebound relative force at min displacement',
    'rebound rsi', 'rebound rsi modified', 'rebound time to peak braking force',
    'rebound time to peak braking force symmetry index', 'rebound time to peak braking power',
    'rebound time to peak braking power symmetry index', 'rebound time to peak propulsive force',
    'rebound time to peak propulsive force symmetry index',
    'rebound time to peak propulsive power',
    'rebound time to peak propulsive power symmetry index', 'rebound time to takeoff',
    'right rebound avg braking force', 'right rebound avg braking power',
    'right rebound avg propulsive force', 'right rebound avg propulsive power',
    'right rebound braking impulse', 'right rebound braking rfd', 'right rebound braking work',
    'right rebound p1 impulse', 'right rebound p2 impulse', 'right rebound peak braking force',
    'right rebound peak braking power', 'right rebound peak propulsive force',
    'right rebound peak propulsive power', 'right rebound propulsive impulse',
    'right rebound propulsive rfd', 'right rebound propulsive work',
    'right rebound time to peak braking force', 'right rebound time to peak braking power',
    'right rebound time to peak propulsive force', 'right rebound time to peak propulsive power',
]

_IMTP_METRICS_CATALOG = [
    'end of rise force', 'end of rise force net', 'end of rise force relative',
    'end of rise force relative net', 'end of rise torque', 'force 100ms', 'force 100ms left',
    'force 100ms left net', 'force 100ms left relative', 'force 100ms left relative net',
    'force 100ms net', 'force 100ms relative', 'force 100ms relative net', 'force 100ms right',
    'force 100ms right net', 'force 100ms right relative', 'force 100ms right relative net',
    'force 100ms si', 'force 150ms', 'force 150ms left', 'force 150ms left net',
    'force 150ms left relative', 'force 150ms left relative net', 'force 150ms net',
    'force 150ms relative', 'force 150ms relative net', 'force 150ms right',
    'force 150ms right net', 'force 150ms right relative', 'force 150ms right relative net',
    'force 150ms si', 'force 200ms', 'force 200ms left', 'force 200ms left net',
    'force 200ms left relative', 'force 200ms left relative net', 'force 200ms net',
    'force 200ms relative', 'force 200ms relative net', 'force 200ms right',
    'force 200ms right net', 'force 200ms right relative', 'force 200ms right relative net',
    'force 200ms si', 'force 250ms', 'force 250ms left', 'force 250ms left net',
    'force 250ms left relative', 'force 250ms left relative net', 'force 250ms net',
    'force 250ms relative', 'force 250ms relative net', 'force 250ms right',
    'force 250ms right net', 'force 250ms right relative', 'force 250ms right relative net',
    'force 250ms si', 'force 50ms', 'force 50ms left', 'force 50ms left net',
    'force 50ms left relative', 'force 50ms left relative net', 'force 50ms net',
    'force 50ms relative', 'force 50ms relative net', 'force 50ms right', 'force 50ms right net',
    'force 50ms right relative', 'force 50ms right relative net', 'force 50ms si',
    'force at max rfd', 'force at max rfd left', 'force at max rfd left net',
    'force at max rfd left relative', 'force at max rfd left relative net', 'force at max rfd net',
    'force at max rfd relative', 'force at max rfd relative net', 'force at max rfd right',
    'force at max rfd right net', 'force at max rfd right relative',
    'force at max rfd right relative net', 'force at max rfd si', 'impulse 100ms', 'impulse 150ms',
    'impulse 200ms', 'impulse 250ms', 'impulse 50ms', 'impulse at max rfd', 'max rfd',
    'onset force', 'onset force net', 'onset force relative', 'onset force relative net',
    'onset time', 'onset torque', 'peak force', 'peak force left', 'peak force left net',
    'peak force left relative', 'peak force left relative net', 'peak force net',
    'peak force relative', 'peak force relative net', 'peak force right', 'peak force right net',
    'peak force right relative', 'peak force right relative net', 'peak force si', 'peak torque',
    'peak torque left', 'peak torque right', 'rfd 100ms', 'rfd 150ms', 'rfd 200ms', 'rfd 250ms',
    'rfd 50ms', 'rfd impulse', 'steadiness rsme force', 'steadiness rsme rfd', 'system weight',
    'time to force end of rise', 'time to peak force', 'time to peak rfd', 'torque 100ms',
    'torque 100ms left', 'torque 100ms right', 'torque 150ms', 'torque 150ms left',
    'torque 150ms right', 'torque 200ms', 'torque 200ms left', 'torque 200ms right',
    'torque 250ms', 'torque 250ms left', 'torque 250ms right', 'torque 50ms', 'torque 50ms left',
    'torque 50ms right', 'torque at max rfd', 'torque at max rfd left', 'torque at max rfd right',
]

# Elenco per jump_type usato dal multiselect di ricerca. SJ non include le
# metriche della fase eccentrica/di contromovimento (non presenti in un
# export SJ reale); CMJ le aggiunge; CMJ RE aggiunge anche le variabili
# "rebound ...".
EXTRA_METRICS_CATALOG = {
    "sj": sorted(_SJ_METRICS_CATALOG),
    "cmj": sorted(set(_SJ_METRICS_CATALOG) | set(_CMJ_EXTRA_METRICS_CATALOG)),
    "cmrj": sorted(set(_SJ_METRICS_CATALOG) | set(_CMJ_EXTRA_METRICS_CATALOG) | set(_CMRJ_EXTRA_METRICS_CATALOG)),
    "imtp": sorted(_IMTP_METRICS_CATALOG),
}


def extra_metric_label(raw_var):
    """Etichetta leggibile per una metrica extra scelta dall'utente (solo
    la prima lettera maiuscola, il resto è già in minuscolo)."""
    return raw_var[:1].upper() + raw_var[1:]


# ============================================================================
# PARTE 1ter — DESCRIZIONI DELLE METRICHE (help "?")
# ============================================================================
# Testi tratti da ccathletics.dk/what-we-measure/ (glossario ufficiale
# ForceMate/ForceDecks), usati per l'help "?" mostrato accanto a ogni
# metrica in Dettaglio Test e nel report. Se una metrica non ha una
# descrizione nota, semplicemente non mostra alcun help (nessuna icona).

# Metriche generiche di salto (SJ/CMJ/CMJ RE): stessa fase, stesso nome,
# stessa definizione a prescindere dal tipo di test.
_CORE_METRIC_DESCRIPTIONS = {
    "jump height ft": "Altezza del salto calcolata dal tempo di volo (Flight Time).",
    "jump height ni": "Altezza del salto calcolata dall'impulso netto (Net Impulse).",
    "takeoff velocity": "Velocità al distacco da terra.",
    "rsi": "Tempo di volo / tempo di contatto (DJ). (Time-to-Takeoff per CMJ/SJ).",
    "rsi modified": "Altezza del salto / tempo di contatto (DJ). (Time-to-Takeoff per CMJ/SJ). Usa l'impulso netto quando disponibile.",
    "rsi exponential": "Formula proprietaria CC Athletics per l'RSI nei Drop Jump: Tempo di volo\u00b2 / Tempo di contatto.",
    "jump momentum": "Quantità di moto generata durante il salto, calcolata come prodotto tra massa corporea e velocità al distacco.",
    "displacement depth": "Profondità massima raggiunta durante la fase di contromovimento, cioè quanto scende il centro di massa prima dell'inizio della fase propulsiva.",
    "net impulse": "Impulso totale (al netto del peso corporeo) applicato durante le fasi di frenata+propulsione.",
    "eccentric impulse": "Impulso netto sull'intera fase eccentrica (scarico + frenata). Poiché l'atleta parte e finisce questa fase da fermo, le due sotto-fasi tendono ad annullarsi a vicenda: valori vicini a zero sono normali.",
    "propulsive impulse": "Impulso netto (forza meno peso corporeo) generato durante la fase propulsiva (dal punto più basso al distacco). Rappresenta l'integrale forza-tempo che accelera il corpo verso l'alto. Equivalente al 'Concentric Impulse' di altri sistemi.",
    "rel. propulsive impulse": "Impulso durante la fase propulsiva relativo alla massa corporea.",
    "braking impulse": "Impulso netto (forza meno peso corporeo) generato solo durante la fase di frenata/decelerazione (dal picco di velocità eccentrica al punto più basso). Quantifica quanto efficacemente l'atleta decelera dopo aver raggiunto la velocità massima verso il basso.",
    "peak force": "Picco di forza istantanea durante l'intero salto.",
    "min unweight force": "Forza minima misurata durante la fase di scarico (unweighting).",
    "rel. min unweight force": "Forza minima durante lo scarico, espressa come multiplo del peso corporeo.",
    "peak braking force": "Picco di forza istantanea durante la sola fase di frenata/decelerazione (dal picco di velocità eccentrica al punto più basso). Per l'intera fase eccentrica, vedi 'Peak Eccentric Force'.",
    "avg. braking force": "Forza media durante la sola fase di frenata/decelerazione (dal picco di velocità eccentrica al punto più basso). Per l'intera fase eccentrica, vedi 'Avg Eccentric Force'.",
    "avg braking force": "Forza media durante la sola fase di frenata/decelerazione (dal picco di velocità eccentrica al punto più basso). Per l'intera fase eccentrica, vedi 'Avg Eccentric Force'.",
    "avg eccentric force": "Forza media durante l'intera fase eccentrica (scarico + frenata). Rappresenta la forza media dall'inizio del movimento al punto più basso del contromovimento.",
    "peak eccentric force": "Picco di forza istantanea durante l'intera fase eccentrica (scarico + frenata). Rappresenta la forza massima dall'inizio del movimento al punto più basso del contromovimento.",
    "avg. propulsive force": "Forza media durante la fase propulsiva.",
    "avg propulsive force": "Forza media durante la fase propulsiva.",
    "rel. avg. propulsive force": "Forza media durante la fase propulsiva, relativa alla massa corporea.",
    "peak propulsive force": "Picco di forza istantanea durante la fase propulsiva.",
    "relative peak force": "Picco di forza espresso come multiplo del peso corporeo.",
    "rel. peak force": "Picco di forza espresso come multiplo del peso corporeo.",
    "force peak power": "Forza istantanea nel momento di picco potenza (sull'intero dataset).",
    "force at min displacement": "Forza totale (di reazione al suolo) misurata nel punto di massima profondità di spostamento durante il contromovimento.",
    "relative force at min displacement": "Forza al punto di minimo spostamento, espressa come multiplo del peso corporeo.",
    "peak braking power": "Picco di potenza istantanea durante la sola fase di frenata/decelerazione (valore negativo). Per l'intera fase eccentrica, vedi 'Peak Eccentric Power'.",
    "avg. braking power": "Potenza meccanica media durante la sola fase di frenata/decelerazione (valore negativo). Per l'intera fase eccentrica, vedi 'Avg Eccentric Power'.",
    "avg braking power": "Potenza meccanica media durante la sola fase di frenata/decelerazione (valore negativo). Per l'intera fase eccentrica, vedi 'Avg Eccentric Power'.",
    "braking work": "Lavoro totale svolto durante la sola fase di frenata/decelerazione (valore negativo). Per l'intera fase eccentrica, vedi 'Eccentric Work'.",
    "peak eccentric power": "Picco di potenza istantanea durante l'intera fase eccentrica (scarico + frenata). Riportato come valore assoluto. Valori alti indicano una buona capacità di potenza eccentrica.",
    "avg eccentric power": "Potenza meccanica media durante l'intera fase eccentrica (scarico + frenata). Riportata come valore assoluto. Rappresenta il tasso medio di assorbimento di energia nella fase discendente.",
    "eccentric work": "Lavoro totale svolto durante l'intera fase eccentrica (scarico + frenata). Riportato come valore assoluto. Quantifica l'energia totale assorbita dall'inizio del movimento al punto più basso.",
    "peak propulsive power": "Potenza massima durante la fase propulsiva.",
    "avg. propulsive power": "Potenza media durante la fase propulsiva.",
    "avg propulsive power": "Potenza media durante la fase propulsiva.",
    "propulsive work": "Lavoro totale svolto durante la fase propulsiva.",
    "peak power": "Picco di potenza istantanea durante l'intero salto.",
    "relative peak power": "Picco di potenza espresso in watt per chilogrammo di peso corporeo.",
    "min braking velocity": "Velocità minima durante la fase di frenata.",
    "avg. braking velocity": "Velocità media durante la fase di frenata.",
    "avg braking velocity": "Velocità media durante la fase di frenata.",
    "min eccentric velocity": "Velocità minima (più negativa) durante l'intera fase eccentrica \u2014 il picco di velocità verso il basso (Eccentric Peak Velocity, EPV), che segna il passaggio dallo scarico alla frenata.",
    "avg eccentric velocity": "Velocità media durante l'intera fase eccentrica (scarico + frenata).",
    "peak propulsive velocity": "Velocità massima raggiunta durante la fase propulsiva.",
    "avg. propulsive velocity": "Velocità media durante la fase propulsiva.",
    "avg propulsive velocity": "Velocità media durante la fase propulsiva.",
    "peak velocity": "Picco di velocità istantanea durante la fase di propulsione.",
    "velocity peak power": "Velocità istantanea nel momento di picco potenza (sull'intero dataset).",
    "avg. rfd": "Tasso medio di sviluppo della forza (Rate of Force Development, RFD).",
    "avg rfd": "Tasso medio di sviluppo della forza (Rate of Force Development, RFD).",
    "braking rfd": "Pendenza media della forza durante la fase di frenata.",
    "decel rfd": "Tasso di sviluppo della forza nella porzione di decelerazione della fase di frenata, misurato dalla velocità minima alla fine della frenata. Simile a 'Braking RFD' o 'Load' di altri sistemi, ma focalizzato sulla fase tardiva della frenata.",
    "eccentric rfd": "Tasso di sviluppo della forza durante la fase eccentrica, calcolato dal valore iniziale al picco di forza.",
    "propulsive rfd": "Tasso di sviluppo della forza durante la fase propulsiva, calcolato dal valore iniziale al picco di forza.",
    "p1 impulse": "Impulso generato durante la prima metà della fase propulsiva (posizione iniziale a triplice flessione). Rappresenta il primo 50% dell'impulso propulsivo.",
    "p2 impulse": "Impulso generato durante la seconda metà della fase propulsiva (fino alla triplice estensione). Rappresenta l'ultimo 50% dell'impulso propulsivo.",
    "p1 duration": "Durata della prima metà della fase propulsiva.",
    "p2 duration": "Durata della seconda metà della fase propulsiva.",
    "p1 avg force": "Forza media durante la prima metà della fase propulsiva (posizione di squat profondo).",
    "p2 avg force": "Forza media durante la seconda metà della fase propulsiva (estensione fino al distacco).",
    "p1 p2 force ratio": "Rapporto tra la forza media in P1 e quella in P2. Sopra 1.0 = produzione di forza maggiore nella posizione più profonda; sotto 1.0 = maggiore durante l'estensione.",
    "p1 p2 duration ratio": "Rapporto tra la durata di P1 e quella di P2. Sopra 1.0 = più tempo nella prima metà dell'impulso; sotto 1.0 = fase iniziale più rapida.",
    "p1 peak force": "Picco di forza durante la fase propulsiva iniziale (P1).",
    "p1 peak velocity": "Picco di velocità durante la fase propulsiva iniziale (P1).",
    "p1 avg velocity": "Velocità media durante la fase propulsiva iniziale (P1).",
    "p1 peak power": "Picco di potenza durante la fase propulsiva iniziale (P1).",
    "p1 avg power": "Potenza media durante la fase propulsiva iniziale (P1).",
    "p2 peak force": "Picco di forza durante la fase propulsiva finale (P2).",
    "p2 peak velocity": "Picco di velocità durante la fase propulsiva finale (P2).",
    "p2 avg velocity": "Velocità media durante la fase propulsiva finale (P2).",
    "p2 peak power": "Picco di potenza durante la fase propulsiva finale (P2).",
    "p2 avg power": "Potenza media durante la fase propulsiva finale (P2).",
    "p1 p2 velocity ratio": "Rapporto tra la velocità media di P1 e P2, indica le caratteristiche di velocità tra fase iniziale e finale.",
    "p1 p2 power ratio": "Rapporto tra la potenza media di P1 e P2, indica le caratteristiche di potenza tra fase iniziale e finale.",
    "peak landing force": "Forza massima registrata durante la fase di atterraggio del salto. Misura la capacità di assorbire le forze in atterraggio.",
    "relative peak landing force": "Picco di forza in atterraggio, espresso come multiplo del peso corporeo.",
    "time to peak landing force": "Tempo dal contatto in atterraggio al picco di forza in atterraggio. Indica quanto rapidamente si raggiungono i picchi di forza durante l'atterraggio.",
    "landing peak force time": "Tempo dal contatto in atterraggio al picco di forza in atterraggio. Indica quanto rapidamente si raggiungono i picchi di forza durante l'atterraggio.",
    "avg landing rfd": "Tasso medio di sviluppo della forza in atterraggio, dal peso corporeo al picco di forza in atterraggio. Indica la rigidità in atterraggio.",
    "landing rfd 0-20ms": "Tasso di sviluppo della forza nei primi 20ms dopo l'atterraggio.",
    "landing rfd 0-40ms": "Tasso di sviluppo della forza nei primi 40ms dopo l'atterraggio.",
    "landing rfd 0-60ms": "Tasso di sviluppo della forza nei primi 60ms dopo l'atterraggio.",
    "landing rfd 0-80ms": "Tasso di sviluppo della forza nei primi 80ms dopo l'atterraggio.",
    "time to peak force": "Tempo impiegato per raggiungere il picco di forza istantanea (dall'inizio del salto).",
    "peak force time": "Tempo impiegato per raggiungere il picco di forza istantanea (dall'inizio del salto).",
    "unweighted duration": "Tempo impiegato per completare la fase di scarico (unweighting).",
    "braking duration": "Durata della sola fase di frenata/decelerazione, misurata dal picco di velocità eccentrica (EPV) al punto più basso del contromovimento.",
    "propulsive duration": "Durata dalla fine della fase di frenata (quando la velocità diventa positiva) fino al distacco. A volte chiamato 'Concentric Time'.",
    "flight time": "Tempo trascorso in fase di volo.",
    "time to peak power": "Tempo dall'inizio del salto al momento di picco potenza.",
    "contact time": "Tempo di contatto con il suolo tra atterraggio e distacco.",
    "contraction time": "Tempo totale di contrazione muscolare dall'inizio del movimento al distacco. Per il CMJ: fasi eccentrica + concentrica. Per lo SJ: solo la fase propulsiva (concentrica).",
    "time to takeoff": "Tempo dall'inizio del movimento al distacco. Indica il tempo totale di contatto col suolo prima del decollo.",
    "takeoff time": "Tempo dall'inizio del movimento al distacco. Indica il tempo totale di contatto col suolo prima del decollo.",
    "time to peak braking force": "Tempo per raggiungere il picco di forza durante la fase di frenata, misurato dall'inizio della frenata.",
    "time to peak braking power": "Tempo per raggiungere il picco di potenza durante la fase di frenata, misurato dall'inizio della frenata.",
    "time to peak eccentric force": "Tempo per raggiungere il picco di forza durante la fase eccentrica, misurato dall'inizio del salto.",
    "time to peak eccentric power": "Tempo per raggiungere il picco di potenza durante la fase eccentrica, misurato dall'inizio del salto.",
    "time to peak propulsive force": "Tempo per raggiungere il picco di forza durante la fase propulsiva, misurato dall'inizio della propulsione.",
    "time to peak propulsive power": "Tempo per raggiungere il picco di potenza durante la fase propulsiva, misurato dall'inizio della propulsione.",
    "system weight": "Peso misurato dell'atleta durante la fase di pesata, appena prima del salto.",
    "body mass": "Massa corporea dell'atleta, misurata prima del test.",
    "body weight": "Peso corporeo dell'atleta registrato per questo test.",
    "body weight sd": "Deviazione standard del peso corporeo durante la fase di pesata (indica quanto l'atleta è rimasto fermo).",
    "propulsive start time": "Momento in cui inizia la fase propulsiva (quando la forza passa da frenata a propulsiva).",
}

# Descrizioni specifiche IMTP (Isometric Mid-Thigh Pull): nomi e definizioni
# non coincidono con quelli dei test di salto, quindi hanno priorità quando
# jump_type == "imtp".
_IMTP_METRIC_DESCRIPTIONS = {
    "peak force": "Picco di forza (valore di forza più alto durante la contrazione volontaria massimale).",
    "peak force net": "Picco di forza netta (valore di forza più alto durante la contrazione volontaria massimale).",
    "peak force relative": "Picco di forza espresso come percentuale del peso corporeo.",
    "peak force relative net": "Picco di forza netta espresso come percentuale del peso corporeo.",
    "onset force": "Forza al momento dell'inizio (onset) della contrazione.",
    "onset force net": "Forza netta al momento dell'inizio (onset) della contrazione.",
    "onset force relative": "Forza all'inizio della contrazione, espressa come percentuale del peso corporeo.",
    "onset force relative net": "Forza netta all'inizio della contrazione, espressa come percentuale del peso corporeo.",
    "onset torque": "Coppia (torque) al momento dell'inizio della contrazione.",
    "onset time": "Istante di inizio della contrazione (dall'inizio del segnale).",
    "end of rise force": "Forza raggiunta nel punto in cui la derivata forza-tempo si stabilizza per la prima volta (fine della fase di salita).",
    "end of rise force net": "Forza netta raggiunta nel punto in cui la derivata forza-tempo si stabilizza per la prima volta.",
    "end of rise force relative": "Forza a fine salita, espressa come percentuale del peso corporeo.",
    "end of rise force relative net": "Forza netta a fine salita, espressa come percentuale del peso corporeo.",
    "end of rise torque": "Coppia raggiunta nel punto in cui la derivata coppia-tempo si stabilizza per la prima volta.",
    "force at max rfd": "Forza nel punto in cui si raggiunge l'RFD massimo.",
    "force at max rfd net": "Forza netta nel punto in cui si raggiunge l'RFD massimo.",
    "force at max rfd relative": "Forza nel punto di RFD massimo, espressa come percentuale del peso corporeo.",
    "force at max rfd relative net": "Forza netta nel punto di RFD massimo, espressa come percentuale del peso corporeo.",
    "max rfd": "Picco del tasso di sviluppo della forza (Rate of Force Development) durante la contrazione.",
    "impulse at max rfd": "Impulso dall'inizio della contrazione al punto di RFD massimo.",
    "rfd impulse": "Impulso accumulato fino al punto di massimo tasso di sviluppo della forza.",
    "steadiness rsme force": "Errore quadratico medio (RMSE) della forza durante la fase di stazionarietà (steady state).",
    "steadiness rsme rfd": "Errore quadratico medio (RMSE) dell'RFD durante la fase di stazionarietà (steady state).",
    "time to force end of rise": "Tempo dall'inizio della contrazione alla fine della fase di salita (End Of Rise).",
    "time to peak force": "Tempo dall'inizio della contrazione al picco di forza durante la contrazione volontaria massimale.",
    "time to peak rfd": "Tempo dall'inizio della contrazione al picco di RFD.",
    "peak torque": "Picco di coppia (valore di coppia più alto durante la contrazione volontaria massimale).",
    "torque at max rfd": "Coppia nel punto in cui si raggiunge l'RFD massimo.",
    "system weight": "Peso corporeo misurato sulla pedana, appena prima della trazione.",
}
for _ms in (50, 100, 150, 200, 250):
    _IMTP_METRIC_DESCRIPTIONS[f"force {_ms}ms"] = f"Forza a {_ms}ms dall'inizio della contrazione."
    _IMTP_METRIC_DESCRIPTIONS[f"force {_ms}ms net"] = f"Forza netta a {_ms}ms dall'inizio della contrazione."
    _IMTP_METRIC_DESCRIPTIONS[f"force {_ms}ms relative"] = f"Forza a {_ms}ms dall'inizio della contrazione, come percentuale del peso corporeo."
    _IMTP_METRIC_DESCRIPTIONS[f"force {_ms}ms relative net"] = f"Forza netta a {_ms}ms dall'inizio della contrazione, come percentuale del peso corporeo."
    _IMTP_METRIC_DESCRIPTIONS[f"rfd {_ms}ms"] = f"Tasso di sviluppo della forza (RFD) a {_ms}ms dall'inizio della contrazione."
    _IMTP_METRIC_DESCRIPTIONS[f"impulse {_ms}ms"] = f"Impulso dall'inizio della contrazione a {_ms}ms."
    _IMTP_METRIC_DESCRIPTIONS[f"torque {_ms}ms"] = f"Coppia a {_ms}ms dall'inizio della contrazione."

# Metriche "rebound ..." (CMJ RE) documentate esplicitamente sul sito con
# testo proprio (non solo "stessa metrica in fase di rimbalzo").
_REBOUND_METRIC_OVERRIDES = {
    "rebound jump height ft": "Altezza del salto di rimbalzo, calcolata dal tempo di volo.",
    "rebound contact time": "Durata del contatto a terra tra l'atterraggio dal CMJ e il distacco del salto di rimbalzo.",
    "rebound rsi": "Reactive Strength Index del salto di rimbalzo (altezza del salto \u00f7 tempo di contatto).",
    "rebound rsi modified": "Reactive Strength Index modificato del salto di rimbalzo (altezza del salto \u00f7 tempo di contatto, formulazione modificata).",
    "rebound jump momentum": "Quantità di moto verticale al distacco del salto di rimbalzo (massa corporea \u00d7 velocità al distacco).",
    "rebound braking impulse": "Impulso totale generato durante la fase di frenata del rimbalzo.",
    "rebound propulsive impulse": "Impulso totale generato durante la fase propulsiva del rimbalzo.",
    "rebound force at min displacement": "Forza verticale nel punto più basso di spostamento del centro di massa durante il contatto di rimbalzo.",
    "rebound relative force at min displacement": "Forza al punto di minimo spostamento durante il rimbalzo, normalizzata sul peso corporeo.",
    "rebound avg braking force": "Forza verticale media durante la fase di frenata del rimbalzo.",
    "rebound peak braking force": "Picco di forza verticale durante la fase di frenata del rimbalzo.",
    "rebound avg propulsive force": "Forza verticale media durante la fase propulsiva del rimbalzo.",
    "rebound peak propulsive force": "Picco di forza verticale durante la fase propulsiva del rimbalzo.",
    "rebound avg braking power": "Potenza media durante la fase di frenata del rimbalzo.",
    "rebound peak braking power": "Picco di potenza durante la fase di frenata del rimbalzo.",
    "rebound avg propulsive power": "Potenza media durante la fase propulsiva del rimbalzo.",
    "rebound peak propulsive power": "Picco di potenza durante la fase propulsiva del rimbalzo.",
    "rebound braking work": "Lavoro totale svolto durante la fase di frenata del rimbalzo.",
    "rebound propulsive work": "Lavoro totale svolto durante la fase propulsiva del rimbalzo.",
    "rebound min braking velocity": "Velocità minima durante la fase di frenata del rimbalzo.",
    "rebound avg braking velocity": "Velocità media durante la fase di frenata del rimbalzo.",
    "rebound peak propulsive velocity": "Velocità massima raggiunta durante la fase propulsiva del rimbalzo.",
    "rebound avg propulsive velocity": "Velocità media durante la fase propulsiva del rimbalzo.",
    "rebound avg braking rfd": "Pendenza media della forza durante la fase di frenata del rimbalzo.",
    "rebound propulsive rfd": "Tasso di sviluppo della forza durante la fase propulsiva del rimbalzo, calcolato dal valore iniziale al picco di forza.",
    "rebound p1 impulse": "Impulso durante la prima metà della fase propulsiva del rimbalzo.",
    "rebound p2 impulse": "Impulso durante la seconda metà della fase propulsiva del rimbalzo.",
    "rebound p1 duration": "Durata della prima metà della fase propulsiva del rimbalzo.",
    "rebound p1 peak force": "Picco di forza durante la fase propulsiva iniziale del rimbalzo (P1).",
    "rebound p1 avg force": "Forza media durante la prima metà della fase propulsiva del rimbalzo.",
    "rebound p1 peak velocity": "Picco di velocità durante la fase propulsiva iniziale del rimbalzo (P1).",
    "rebound p1 avg velocity": "Velocità media durante la fase propulsiva iniziale del rimbalzo (P1).",
    "rebound p1 peak power": "Picco di potenza durante la fase propulsiva iniziale del rimbalzo (P1).",
    "rebound p1 avg power": "Potenza media durante la fase propulsiva iniziale del rimbalzo (P1).",
    "rebound p2 duration": "Durata della seconda metà della fase propulsiva del rimbalzo.",
    "rebound p2 peak force": "Picco di forza durante la fase propulsiva finale del rimbalzo (P2).",
    "rebound p2 avg force": "Forza media durante la seconda metà della fase propulsiva del rimbalzo.",
    "rebound p2 peak velocity": "Picco di velocità durante la fase propulsiva finale del rimbalzo (P2).",
    "rebound p2 avg velocity": "Velocità media durante la fase propulsiva finale del rimbalzo (P2).",
    "rebound p2 peak power": "Picco di potenza durante la fase propulsiva finale del rimbalzo (P2).",
    "rebound p2 avg power": "Potenza media durante la fase propulsiva finale del rimbalzo (P2).",
    "rebound p1 p2 force ratio": "Rapporto tra la forza media in P1 e in P2 del rimbalzo.",
    "rebound p1 p2 velocity ratio": "Rapporto tra la velocità media di P1 e P2 del rimbalzo.",
    "rebound p1 p2 power ratio": "Rapporto tra la potenza media di P1 e P2 del rimbalzo.",
    "rebound p1 p2 duration ratio": "Rapporto tra la durata di P1 e la durata di P2 del rimbalzo.",
    "rebound flight time": "Durata della fase di volo del salto di rimbalzo.",
    "rebound time to peak braking force": "Tempo dall'inizio del contatto di rimbalzo al picco di forza in frenata.",
    "rebound time to takeoff": "Tempo dall'inizio del contatto di rimbalzo al distacco del rimbalzo.",
    "rebound time to peak braking power": "Tempo per raggiungere il picco di potenza durante la fase di frenata del rimbalzo, misurato dall'inizio della frenata.",
    "rebound time to peak propulsive force": "Tempo per raggiungere il picco di forza durante la fase propulsiva del rimbalzo, misurato dall'inizio della propulsione.",
    "rebound time to peak propulsive power": "Tempo per raggiungere il picco di potenza durante la fase propulsiva del rimbalzo, misurato dall'inizio della propulsione.",
}

# Descrizioni manuali per le metriche derivate dall'app stessa (nessun
# raw_var diretto, quindi non ricavabili dal sito).
_DERIVED_METRIC_DESCRIPTIONS = {
    "imtp_rel_peak_force": "IMTP Peak Force diviso per la massa corporea (N/kg): permette di confrontare atleti di taglia diversa.",
    "sj_net_rel_impulse": "SJ Net Impulse diviso per la massa corporea (N\u00b7s/kg).",
    "cmj_net_rel_impulse": "CMJ Net Impulse diviso per la massa corporea (N\u00b7s/kg).",
    "dsi": "Dynamic Strength Index: rapporto tra CMJ Peak Propulsive Force e IMTP Peak Force. Indica quanta della forza massima isometrica viene espressa durante un gesto balistico come il CMJ. Non è un indice 'più alto = meglio': entrambi gli estremi descrivono un profilo diverso, non uno migliore.",
    "eur": "Eccentric Utilisation Ratio: rapporto tra CMJ Height e SJ Height. Indica quanto la fase eccentrica (contromovimento) contribuisce alla prestazione rispetto a un salto puramente concentrico. Da leggere sempre insieme ai valori assoluti: due atleti con lo stesso EUR possono avere prestazioni molto diverse.",
}


def get_metric_description(jump_type, raw_var):
    """Cerca la descrizione di una metrica (per l'help "?"), con fallback
    su prefissi lato sinistro/destro, fase di rimbalzo (rebound) e indice
    di simmetria. Ritorna None se non trovata (nessun help, per design)."""
    if not raw_var:
        return None
    key = raw_var.strip().lower()

    if jump_type == "imtp" and key in _IMTP_METRIC_DESCRIPTIONS:
        return _IMTP_METRIC_DESCRIPTIONS[key]
    if key in _REBOUND_METRIC_OVERRIDES:
        return _REBOUND_METRIC_OVERRIDES[key]
    if key in _CORE_METRIC_DESCRIPTIONS:
        return _CORE_METRIC_DESCRIPTIONS[key]

    # Nell'IMTP "left"/"right" compare come token in mezzo al nome (es.
    # "force 100ms left net"), non come prefisso: lo rimuoviamo e cerchiamo
    # la base risultante.
    if jump_type == "imtp":
        for side_label, token in (("lato sinistro", "left"), ("lato destro", "right")):
            if re.search(rf"\b{token}\b", key):
                candidate = re.sub(rf"\b{token}\b", "", key)
                candidate = re.sub(r"\s+", " ", candidate).strip()
                if candidate in _IMTP_METRIC_DESCRIPTIONS:
                    return f"{_IMTP_METRIC_DESCRIPTIONS[candidate]} ({side_label})"

    for side_label, prefix in (("lato sinistro", "left "), ("lato destro", "right ")):
        if key.startswith(prefix):
            base_desc = get_metric_description(jump_type, key[len(prefix):])
            if base_desc:
                return f"{base_desc} ({side_label})"

    if key.startswith("rebound "):
        base_desc = get_metric_description(jump_type, key[len("rebound "):])
        if base_desc:
            return f"{base_desc} (fase di rimbalzo / rebound)"

    for suffix in (" sym. index", " symmetry index", " si"):
        if key.endswith(suffix):
            base_desc = get_metric_description(jump_type, key[: -len(suffix)].strip())
            if base_desc:
                return f"Indice di simmetria sinistra-destra per: {base_desc}"

    return None


# Attacca la descrizione a ciascuna metrica curata (usata per l'help "?"
# nella scheda Dettaglio Test e nel report). Fatto qui, dopo la definizione
# di get_metric_description, per evitare di duplicare il testo nel
# dizionario METRICS stesso.
for _m in METRICS:
    if _m.get("raw_var"):
        _m["desc"] = get_metric_description(_m.get("jump_type"), _m["raw_var"])
    else:
        _m["desc"] = _DERIVED_METRIC_DESCRIPTIONS.get(_m["key"])


# ============================================================================
# PARTE 2 — CARICAMENTO FILE (sidebar upload)
# ============================================================================

st.set_page_config(page_title="Force Plate Test Report", page_icon="🏋️", layout="wide")

# Palette brand DU Coaching (Destination Unknown), allineata al tema
# Streamlit fornito dall'utente:
#   primaryColor           #e94a26  -> ACCENT (riferimenti di popolazione/benchmark)
#   secondaryBackgroundColor #0bb6ff -> PRIMARY (dati dell'atleta: linee/punti nei grafici)
#   textColor               #065678 -> TEXT_COLOR (testi, titoli, chrome UI)
#   backgroundColor          #ffffff -> BG_COLOR (sfondo grafici/report)
PRIMARY = "#0bb6ff"     # azzurro — dati dell'atleta (linee/punti nei grafici)
ACCENT = "#e94a26"      # arancio — riferimenti di popolazione/benchmark
TEXT_COLOR = "#065678"  # blu scuro — testi, titoli, chrome dei widget
BG_COLOR = "#ffffff"    # sfondo

# Applica la palette brand ai widget nativi di Streamlit (tab attive,
# pulsanti, download button, checkbox) iniettando CSS direttamente nel file,
# così il branding funziona subito senza dover distribuire/ricordare anche
# un file .streamlit/config.toml separato. Nota: usa selettori CSS interni
# di Streamlit che potrebbero cambiare in futuri aggiornamenti della libreria.
st.markdown(f"""
<style>
    .stTabs [data-baseweb="tab-highlight"] {{ background-color: {TEXT_COLOR} !important; }}
    .stTabs [aria-selected="true"] {{ color: {TEXT_COLOR} !important; }}
    .stButton > button, .stDownloadButton > button {{
        border-color: {TEXT_COLOR} !important;
        color: {TEXT_COLOR} !important;
    }}
    .stButton > button:hover, .stDownloadButton > button:hover,
    .stButton > button:active, .stDownloadButton > button:active {{
        border-color: {ACCENT} !important;
        color: {ACCENT} !important;
    }}
    input[type="checkbox"] {{ accent-color: {TEXT_COLOR} !important; }}
</style>
""", unsafe_allow_html=True)

st.sidebar.title("📂 Import dati")

uploaded = st.sidebar.file_uploader(
    "Carica i file XLSX esportati da ForceMate", type=["xlsx"], accept_multiple_files=True
)

st.sidebar.markdown("---")

csv_reload = st.sidebar.file_uploader(
    "🔁 Ricarica export completo (CSV)", type=["csv"],
    help="Carica un CSV precedentemente esportato dalla scheda 'Dettaglio Test' "
         "(forceplate_fulldata_...) per rivedere grafici e tabelle senza i file XLSX "
         "originali. Se presente, ha precedenza sui file XLSX caricati sotto.",
)


# Le costanti di popolazione (scheda "⚙️ Costanti") non dipendono dai file
# caricati: inizializziamo subito lo stato in modo che quella scheda sia
# consultabile/modificabile anche prima di caricare qualsiasi file.
if "pop" not in st.session_state:
    st.session_state["pop"] = {k: dict(v) for k, v in DEFAULT_POP.items()}

# Soglie degli indici a rapporto (DSI, EUR): modificabili nella scheda
# "⚙️ Costanti", insieme alle costanti di popolazione.
if "idx_thr" not in st.session_state:
    st.session_state["idx_thr"] = {k: list(v) for k, v in DEFAULT_INDEX_THRESHOLDS.items()}


# ============================================================================
# PARTE 3 — LETTURA / PARSING DEI FILE XLSX
# ============================================================================
# Ogni file ForceMate ha: colonne A-F = metadati (generali, attributi
# atleta, campi custom); colonna G = unità di misura; blocchi da 3 colonne
# ("Jump 1", "Jump 2", ...) con [variabile, valore, spacer] per ripetizione.
# Il "jump type" è taggato per singola ripetizione, non per file (es. un
# file SJ può contenere per errore una rep taggata "cmj").

@dataclass
class ParsedFile:
    filename: str
    metadata: dict
    reps: list = field(default_factory=list)  # [{"jump_type","vars","units"}, ...]


def _parse_date(raw):
    if raw is None:
        return None
    if isinstance(raw, dt.datetime):
        return raw
    s = str(raw).strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_forcemate_workbook(file_like, filename: str) -> ParsedFile:
    wb = openpyxl.load_workbook(file_like, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ParsedFile(filename=filename, metadata={}, reps=[])

    if _is_imtp_trial_layout(rows):
        return _parse_imtp_trial_rows(rows, filename)

    max_col = max(len(r) for r in rows)

    def cell(r_idx, c_idx):
        row = rows[r_idx - 1] if 0 < r_idx <= len(rows) else ()
        return row[c_idx - 1] if 0 < c_idx <= len(row) else None

    def kv_lookup(col_key, col_val, max_r=15):
        out = {}
        for r in range(1, max_r + 1):
            k = cell(r, col_key)
            if k:
                out[str(k).strip().lower()] = cell(r, col_val)
        return out

    generali = kv_lookup(1, 2)
    attributi = kv_lookup(3, 4)
    custom = kv_lookup(5, 6)

    metadata = {
        "nome": attributi.get("name") or generali.get("name") or "Atleta",
        "sesso": (attributi.get("gender") or "").strip().upper() if attributi.get("gender") else None,
        "altezza_cm": attributi.get("height"),
        "peso_kg_input": attributi.get("weight"),
        "data_test": _parse_date(generali.get("date")),
        "device": generali.get("device"),
        "team": generali.get("team"),
        "test_period": custom.get("test period"),
        "test_type": custom.get("test type"),
    }

    block_starts = [c for c in range(8, max_col + 1)
                    if cell(1, c) and str(cell(1, c)).strip().lower().startswith("jump")]

    reps = []
    for bs in block_starts:
        var_col, val_col, unit_col = bs, bs + 1, 7
        variables, units, jump_type = {}, {}, None
        for r in range(2, len(rows) + 1):
            var_name = cell(r, var_col)
            if not var_name:
                continue
            var_name = str(var_name).strip().lower()
            val = cell(r, val_col)
            unit = cell(r, unit_col)
            if var_name == "jump type":
                jump_type = str(val).strip().lower() if val else None
                continue
            if var_name == "tags":
                continue
            num = None
            if isinstance(val, (int, float)):
                num = float(val)
            elif isinstance(val, str):
                try:
                    num = float(val.replace(",", "."))
                except ValueError:
                    num = None
            variables[var_name] = num if num is not None else val
            if unit:
                units[var_name] = str(unit).strip()
        reps.append({"jump_type": jump_type, "vars": variables, "units": units})

    return ParsedFile(filename=filename, metadata=metadata, reps=reps)


def apply_type_override(parsed: ParsedFile, override_type):
    if not override_type:
        return parsed
    for rep in parsed.reps:
        if not rep["jump_type"]:
            rep["jump_type"] = override_type
    return parsed


def guess_type_from_filename(filename: str):
    fn = filename.upper()
    for t in ("CMRJ", "IMTP", "CMJ", "SJ", "DJ"):
        if t in fn:
            return t.lower()
    return None


def _is_imtp_trial_layout(rows) -> bool:
    """Rileva il layout Unit/Trial N (IMTP senza intestazione ForceMate
    standard): riga 1, colonna A = 'Unit', e almeno una colonna successiva
    che inizia con 'Trial'."""
    if not rows or not rows[0]:
        return False
    header = rows[0]
    if not header or str(header[6]).strip().lower() != "unit":
        return False
    return any(v and str(v).strip().lower().startswith("trial") for v in header[6:])


def _to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip().replace(",", "."))
        except ValueError:
            return None
    return None


def _parse_imtp_trial_rows(rows, filename: str) -> ParsedFile:
    max_col = max(len(r) for r in rows)

    def cell(r_idx, c_idx):
        row = rows[r_idx - 1] if 0 < r_idx <= len(rows) else ()
        return row[c_idx - 1] if 0 < c_idx <= len(row) else None

    # Ogni blocco "Trial N" occupa 3 colonne: variabile, valore, spacer.
    # La colonna A contiene sempre l'unità di misura della riga.
    block_starts = [c for c in range(2, max_col + 1)
                    if cell(1, c) and str(cell(1, c)).strip().lower().startswith("trial")]

    reps = []
    for bs in block_starts:
        var_col, val_col = bs, bs + 1
        variables, units = {}, {}
        for r in range(2, len(rows) + 1):
            var_name = cell(r, var_col)
            if not var_name:
                continue
            var_name = str(var_name).strip().lower()
            val = cell(r, val_col)
            unit = cell(r, 1)
            num = _to_num(val)
            variables[var_name] = num if num is not None else val
            if unit:
                units[var_name] = str(unit).strip()

        # Non c'è un campo "body mass" in kg: c'è "system weight" (N),
        # cioè il peso corporeo come forza. Lo convertiamo per poter
        # calcolare IMTP Rel Peak Force. ASSUNZIONE DA VERIFICARE: g=9.80665.
        sw = variables.get("system weight")
        if "body mass" not in variables and isinstance(sw, (int, float)):
            variables["body mass"] = sw / 9.80665

        # Nessun tag "jump type" nel file: resta None e viene assegnato
        # dal selettore in sidebar (apply_type_override), come richiesto.
        reps.append({"jump_type": None, "vars": variables, "units": units})

    # Questo layout non contiene nome/sesso/peso/data: verranno ereditati
    # dagli altri file caricati insieme (vedi sesso_da_file()).
    metadata = {
        "nome": None, "sesso": None, "altezza_cm": None, "peso_kg_input": None,
        "data_test": None, "device": None, "team": None,
        "test_period": None, "test_type": None,
    }
    return ParsedFile(filename=filename, metadata=metadata, reps=reps)


FULL_META_COLS = {"Nome", "Sesso", "Data test", "File", "Jump Type", "Indice Ripetizione"}


def parse_full_export_to_parsed(file_like) -> ParsedFile:
    """Ricostruisce un ParsedFile dall'export completo (una riga per
    ripetizione, una colonna per variabile grezza), per rivedere grafici e
    tabelle senza i file XLSX originali. I dati sono già per ripetizione e
    completi: nulla va ricostruito a ritroso, nemmeno il peso corporeo
    delle metriche 'Rel'."""
    df = pd.read_csv(file_like, encoding="utf-8-sig")
    if "Jump Type" not in df.columns:
        raise ValueError("Il file non sembra un export completo: manca la colonna 'Jump Type'.")

    var_cols = [c for c in df.columns if c not in FULL_META_COLS]
    if not var_cols:
        raise ValueError("Nessuna colonna di variabili trovata nel file.")

    reps = []
    for _, row in df.iterrows():
        jump_type = str(row["Jump Type"]).strip().lower()
        if jump_type in ("", "—", "-", "nan", "none"):
            jump_type = None
        variables = {}
        for c in var_cols:
            v = pd.to_numeric(row[c], errors="coerce")
            if pd.notna(v):
                variables[c] = float(v)
        if variables:
            reps.append({"jump_type": jump_type, "vars": variables, "units": {}})

    if not reps:
        raise ValueError("Nessuna ripetizione valida nel file.")

    def _primo(colonna):
        if colonna in df.columns and len(df):
            v = str(df[colonna].iloc[0]).strip()
            if v and v.lower() not in ("nan", "none", "-"):
                return v
        return None

    metadata = {
        "nome": _primo("Nome"), "sesso": _primo("Sesso"), "altezza_cm": None,
        "peso_kg_input": None, "data_test": None, "device": None, "team": None,
        "test_period": None, "test_type": None, "periodo_override": _primo("Data test"),
    }
    return ParsedFile(filename="(da CSV)", metadata=metadata, reps=reps)


def _slug_nome(text):
    """Nome atleta utilizzabile in un nome file."""
    s = re.sub(r"[^A-Za-z0-9]+", "_", str(text or "").strip())
    return re.sub(r"_+", "_", s).strip("_") or "atleta"


def _data_file_token(files, periodo_str):
    """Data del test in formato ISO per il nome file: si preferisce la data
    reale dei file XLSX, poi quella scritta nel CSV ricaricato, infine oggi."""
    date_tests = [pf.metadata.get("data_test") for pf in files if pf.metadata.get("data_test")]
    if date_tests:
        return min(date_tests).strftime("%Y-%m-%d")
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", str(periodo_str or ""))
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return dt.date.today().isoformat()


def build_full_raw_export(files, nome, sesso, periodo):
    """Dump grezzo di TUTTE le variabili presenti nei file XLSX caricati,
    non solo quelle riconosciute dal catalogo dell'app (curate o extra):
    pensato per studiare nuove metriche e costruire in futuro nuovi dati
    di popolazione. Una riga per ripetizione; le colonne sono l'unione di
    tutte le variabili incontrate in tutti i file/ripetizioni, quindi
    normale che compaiano molte celle vuote (una ripetizione non ha tutte
    le variabili di tutti gli altri tipi di test).

    Le ripetizioni deselezionate nella scheda Dettaglio Test vengono
    escluse, coerentemente con l'altro export e con medie/T-score."""
    all_vars = set()
    rep_entries = []
    tipo_counter = {}
    for pf in files:
        for i, rep in enumerate(pf.reps):
            jt = rep.get("jump_type")
            # Indice progressivo PER TIPO DI TEST attraverso tutti i file,
            # nello stesso ordine di collect_reps_all(): è la chiave con cui
            # sono salvate le checkbox "Prova N" (incl_{jump_type}_{indice}).
            # Va incrementato anche per le ripetizioni scartate, altrimenti
            # le successive si disallineano.
            idx_tipo = tipo_counter.get(jt, 0)
            tipo_counter[jt] = idx_tipo + 1
            if jt and not is_rep_included(jt, idx_tipo):
                continue
            rep_entries.append((pf.filename, i, rep))
            all_vars.update(rep["vars"].keys())
    all_vars_sorted = sorted(all_vars)

    rows = []
    for filename, idx, rep in rep_entries:
        row = {
            "Nome": nome, "Sesso": sesso, "Data test": periodo,
            "File": filename, "Jump Type": rep.get("jump_type") or "—",
            "Indice Ripetizione": idx + 1,
        }
        for var in all_vars_sorted:
            row[var] = rep["vars"].get(var)
        rows.append(row)
    return pd.DataFrame(rows)


parsed_files = []
if csv_reload is not None:
    try:
        parsed_files = [parse_full_export_to_parsed(io.BytesIO(csv_reload.getvalue()))]
        st.sidebar.success(f"Dati ricaricati da '{csv_reload.name}'.")
        if uploaded:
            st.sidebar.caption("File XLSX caricati sotto ignorati: è attivo il CSV ricaricato sopra.")
    except Exception as e:
        st.sidebar.error(f"Impossibile leggere il CSV: {e}")
elif uploaded:
    st.sidebar.markdown("---")
    st.sidebar.subheader("Tipo di esercizio per file")
    st.sidebar.caption("Verificare soprattutto l'IMTP: spesso non viene taggato automaticamente dal software.")

    for f in uploaded:
        pf = parse_forcemate_workbook(io.BytesIO(f.getvalue()), f.name)
        guess = guess_type_from_filename(f.name)
        detected_types = {r["jump_type"] for r in pf.reps if r["jump_type"]}
        default = guess or (list(detected_types)[0] if len(detected_types) == 1 else None)
        options = ["imtp", "sj", "cmj", "cmrj"]
        idx = options.index(default) if default in options else 0
        choice = st.sidebar.selectbox(
            f.name, options, index=idx, format_func=lambda x: JUMP_TYPE_LABELS.get(x, x),
            key=f"type_{f.name}",
            help="Applicato solo alle ripetizioni prive di tag automatico nel file."
        )
        parsed_files.append(apply_type_override(pf, choice))


# ============================================================================
# PARTE 4 — ANALISI DATI E CONFRONTO CON LA POPOLAZIONE
# ============================================================================
# Equivalente alle funzioni MEDIA_DINAMICA_JUMP / DEV_STD_DINAMICA_JUMP /
# CONTA_DINAMICA_JUMP dello script Apps Script originale: i valori vengono
# aggregati per tipo di test attraverso TUTTI i file caricati, poi
# confrontati con le norme di popolazione tramite Z-score / T-score.

def pooled_stats(values):
    values = [v for v in values if isinstance(v, (int, float)) and not math.isnan(v)]
    n = len(values)
    if n == 0:
        return dict(n=0, mean=None, sd=None)
    mean = sum(values) / n
    if n < 2:
        return dict(n=n, mean=mean, sd=None)
    var = sum((v - mean) ** 2 for v in values) / n  # ddof=0, come nello script originale
    return dict(n=n, mean=mean, sd=math.sqrt(var))


def z_t_score(value, pop_mean, pop_sd, lower_is_better):
    if value is None or pop_mean is None or not pop_sd:
        return None, None
    z = (value - pop_mean) / abs(pop_sd)
    if lower_is_better:
        z = -z
    return z, 50 + 10 * z


def collect_reps_all(files, jump_type):
    """Tutte le ripetizioni di un tipo, nell'ordine di caricamento — senza
    applicare le esclusioni scelte dall'utente. Usata per popolare le
    colonne 'Salto N' nel Dettaglio Test."""
    return [rep for pf in files for rep in pf.reps if rep["jump_type"] == jump_type]


def is_rep_included(jump_type, index, default=True):
    return st.session_state.get(f"incl_{jump_type}_{index}", default)


def collect_reps(files, jump_type):
    """Ripetizioni di un tipo, filtrate escludendo i salti deselezionati
    dall'utente nella scheda Dettaglio Test (checkbox 'Salto N')."""
    reps_all = collect_reps_all(files, jump_type)
    return [rep for i, rep in enumerate(reps_all) if is_rep_included(jump_type, i)]


def compute_metric_values(files, metric):
    reps = collect_reps(files, metric["jump_type"])
    values = []
    for rep in reps:
        if metric.get("raw_var"):
            v = rep["vars"].get(metric["raw_var"])
        elif metric.get("derive"):
            v = metric["derive"](rep["vars"])
        else:
            v = None
        if isinstance(v, (int, float)):
            values.append(v)
    return values


def per_rep_metric_values(files, metric):
    """Valore della metrica per OGNI ripetizione (inclusi i salti esclusi
    dal calcolo), nello stesso ordine delle colonne 'Salto N'."""
    reps_all = collect_reps_all(files, metric["jump_type"])
    values = []
    for rep in reps_all:
        if metric.get("raw_var"):
            v = rep["vars"].get(metric["raw_var"])
        elif metric.get("derive"):
            v = metric["derive"](rep["vars"])
        else:
            v = None
        values.append(v if isinstance(v, (int, float)) else None)
    return values


def best_worst_indices(values, incl_mask, lower_is_better):
    """Indice del salto migliore e peggiore tra quelli inclusi e numerici."""
    candidates = [(i, v) for i, (v, inc) in enumerate(zip(values, incl_mask))
                  if inc and isinstance(v, (int, float))]
    if len(candidates) < 2:
        return None, None
    key = (lambda t: t[1])
    best = (min if lower_is_better else max)(candidates, key=key)[0]
    worst = (max if lower_is_better else min)(candidates, key=key)[0]
    if best == worst:
        return None, None
    return best, worst


def sesso_da_file(files):
    for pf in files:
        if pf.metadata.get("sesso"):
            return pf.metadata["sesso"]
    return "UOMO"


def build_results(files, pop_dict, thresholds):
    sesso_raw = sesso_da_file(files)
    is_female = sesso_raw and sesso_raw.upper().startswith("F")

    support = {}
    results = []

    for metric in METRICS:
        if metric["kind"] in ("score", "info") and metric.get("jump_type"):
            values = compute_metric_values(files, metric)
            stats = pooled_stats(values)
            support[metric["key"]] = stats

            z = t = banda = colore = None
            pop_mean = pop_sd = None
            if metric.get("pop_key"):
                pop = pop_dict[metric["pop_key"]]
                pop_mean = pop["mean_f"] if is_female else pop["mean_m"]
                pop_sd = pop["sd_f"] if is_female else pop["sd_m"]
                if metric["kind"] == "score":
                    z, t = z_t_score(stats["mean"], pop_mean, pop_sd, metric["lower_is_better"])
                    banda, colore = banda_da_tscore(t)

            results.append(dict(
                key=metric["key"], label=metric["label"], category=metric["category"],
                unit=metric["unit"], kind=metric["kind"], n=stats["n"], mean=stats["mean"], sd=stats["sd"],
                z=z, t=t, banda=banda, colore=colore, pop_mean=pop_mean, pop_sd=pop_sd,
                desc=metric.get("desc"),
            ))

    # --- Indici derivati da medie aggregate (DSI, EUR) ---
    cmj_peak = support.get("cmj_peak_force", {}).get("mean")
    imtp_peak = support.get("imtp_peak_force", {}).get("mean")
    dsi_val = (cmj_peak / imtp_peak) if (cmj_peak and imtp_peak) else None

    # EUR = CMJ height / SJ height (rapporto puro), coerente con la tabella
    # costanti fornita dall'utente (~1,11 uomini, ~1,09 donne).
    cmj_h = support.get("cmj_height", {}).get("mean")
    sj_h = support.get("sj_height", {}).get("mean")
    eur_val = (cmj_h / sj_h) if (cmj_h and sj_h) else None

    for key, val in (("dsi", dsi_val), ("eur", eur_val)):
        pop = pop_dict[key]
        pop_mean = pop["mean_f"] if is_female else pop["mean_m"]
        metric_def = next(m for m in METRICS if m["key"] == key)
        thr_low, thr_high = thresholds[key]
        zona, zona_colore = zona_da_indice(key, val, thr_low, thr_high)
        # Niente T-score per DSI/EUR: non sono metriche "più alto = meglio"
        # ma indici diagnostici, dove entrambi gli estremi descrivono un
        # profilo diverso, non uno migliore. Resta la media di popolazione,
        # mostrata come riferimento nell'help e nel report.
        results.append(dict(
            key=key, label=metric_def["label"], category="INDICI", unit="",
            kind="score_single", n=(1 if val is not None else 0), mean=val, sd=None,
            z=None, t=None, banda=None, colore=None, pop_mean=pop_mean, pop_sd=None,
            zona=zona, zona_colore=zona_colore, thr_low=thr_low, thr_high=thr_high,
            desc=metric_def.get("desc"),
        ))

    # --- Checks a soglia ---
    def make_check(defn, ratio):
        passed = None
        if ratio is not None:
            passed = (ratio >= defn["threshold"]) if defn["direction"] == "min" else (abs(ratio) <= defn["threshold"])
        return dict(**defn, value=ratio, passed=passed)

    initial_h = support.get("cmj_re_initial_height", {}).get("mean")
    rebound_h = support.get("cmj_re_rebound_height", {}).get("mean")
    landing_sym = support.get("unbalanced_landing_raw", {}).get("mean")
    cmj_h_standalone = support.get("cmj_height", {}).get("mean")
    cmj_re_contact_time = support.get("cmj_re_contact_time", {}).get("mean")

    checks_out = [
        make_check(CHECKS[0], (rebound_h / initial_h) if (initial_h and rebound_h) else None),
        make_check(CHECKS[1], (initial_h / cmj_h_standalone) if (initial_h and cmj_h_standalone) else None),
        make_check(CHECKS[2], (landing_sym / 100.0) if landing_sym is not None else None),
        make_check(CHECKS[3], cmj_re_contact_time if cmj_re_contact_time is not None else None),
    ]

    return dict(results=results, support=support, checks=checks_out, sesso=sesso_raw)


def profilo_forza(results):
    out = {}
    for cat in CATEGORIES:
        ts = [r["t"] for r in results if r["category"] == cat and r["t"] is not None]
        out[cat] = (sum(ts) / len(ts)) if ts else None
    return out


# ============================================================================
# PARTE 4bis — GRAFICI A QUADRANTI E A BANDE
# ============================================================================
# Due famiglie di grafici bidimensionali:
#
#  - quadrant_chart (RSQ): incrocia due metriche e divide il piano nei
#    quattro quadranti definiti dal crosshair centrato sulla media di
#    popolazione. Legge il livello assoluto nelle due metriche.
#
#  - ratio_wedge_chart (DSI, EUR): per un indice che è un RAPPORTO tra le
#    due metriche. Qui i quadranti non servono: tutti i punti con lo stesso
#    rapporto stanno su una retta per l'origine, quindi il piano va diviso
#    da rette diagonali (y = x / soglia), non da una croce. Il crosshair di
#    popolazione resta come contesto sul livello assoluto.

def _wrap_label(text, width=16):
    """Spezza un'etichetta lunga su più righe (per gli assi angolari del
    radar e per le etichette delle zone), così non deborda oltre il bordo
    del grafico invece di essere tagliata."""
    return "<br>".join(textwrap.wrap(text, width=width, break_long_words=False))


def quadrant_chart(x_mean, y_mean, x_sd, y_sd, x_label, y_label, quadrant_defs,
                   pop_x=None, pop_y=None, point_color=PRIMARY, height=430):
    """quadrant_defs: dict con chiavi 'tl','tr','bl','br' -> (etichetta, colore).
    Il crosshair (linee tratteggiate e confini dei quadranti) è centrato sulla
    media di popolazione (pop_x, pop_y); se non disponibile per una metrica si
    usa la media del test come fallback. Il punto mostrato è la media del test
    con barre d'errore pari alla deviazione standard delle ripetizioni incluse.
    Ritorna una go.Figure oppure None se manca la media del test su uno dei due assi."""
    if x_mean is None or y_mean is None:
        return None
    x_sd = x_sd or 0
    y_sd = y_sd or 0
    cx = pop_x if pop_x is not None else x_mean
    cy = pop_y if pop_y is not None else y_mean

    # Raggio simmetrico intorno al crosshair (cx, cy): copre la distanza
    # massima tra il crosshair e gli estremi del punto (media ± dev.std),
    # più un margine di respiro. Essendo simmetrico, il crosshair cade
    # sempre esattamente al centro del grafico e i quattro quadranti
    # risultano sempre della stessa dimensione.
    x_extent = abs(x_mean - cx) + x_sd
    y_extent = abs(y_mean - cy) + y_sd
    x_r = x_extent * 1.3 if x_extent > 0 else (abs(cx) * 0.15 or 1)
    y_r = y_extent * 1.3 if y_extent > 0 else (abs(cy) * 0.15 or 1)
    x0, x1 = cx - x_r, cx + x_r
    y0, y1 = cy - y_r, cy + y_r

    tl_label, tl_color = quadrant_defs["tl"]
    tr_label, tr_color = quadrant_defs["tr"]
    bl_label, bl_color = quadrant_defs["bl"]
    br_label, br_color = quadrant_defs["br"]

    fig = go.Figure()
    for x_a, x_b, y_a, y_b, color in (
        (x0, cx, cy, y1, tl_color), (cx, x1, cy, y1, tr_color),
        (x0, cx, y0, cy, bl_color), (cx, x1, y0, cy, br_color),
    ):
        fig.add_shape(type="rect", x0=x_a, x1=x_b, y0=y_a, y1=y_b,
                      fillcolor=color, opacity=0.15, line_width=0)

    for x_pos, y_pos, text, anchor in (
        ((x0 + cx) / 2, y1, tl_label, "top"), ((cx + x1) / 2, y1, tr_label, "top"),
        ((x0 + cx) / 2, y0, bl_label, "bottom"), ((cx + x1) / 2, y0, br_label, "bottom"),
    ):
        fig.add_annotation(x=x_pos, y=y_pos, text=text, showarrow=False,
                           font=dict(size=10, color="#666"), yanchor=anchor)

    fig.add_vline(x=cx, line_dash="dash", line_color=ACCENT)
    fig.add_hline(y=cy, line_dash="dash", line_color=ACCENT)

    fig.add_trace(go.Scatter(
        x=[x_mean], y=[y_mean], mode="markers",
        marker=dict(size=12, color=point_color, line=dict(width=1.5, color="white")),
        error_x=dict(type="data", array=[x_sd], visible=True, color=point_color, thickness=1.5, width=6),
        error_y=dict(type="data", array=[y_sd], visible=True, color=point_color, thickness=1.5, width=6),
        name="Media test \u00b1 Dev.Std",
        hovertemplate=(f"Media test<br>{x_label}: %{{x:.2f}} \u00b1 {x_sd:.2f}"
                       f"<br>{y_label}: %{{y:.2f}} \u00b1 {y_sd:.2f}<extra></extra>"),
    ))

    fig.update_layout(
        xaxis_title=x_label, yaxis_title=y_label,
        xaxis_range=[x0, x1], yaxis_range=[y0, y1],
        xaxis_showgrid=False, yaxis_showgrid=False,
        xaxis=dict(automargin=True, title=dict(standoff=12)),
        yaxis=dict(automargin=True, title=dict(standoff=12)),
        height=height, margin=dict(t=30, b=50, l=70, r=20), showlegend=False,
        plot_bgcolor=BG_COLOR, paper_bgcolor=BG_COLOR, font=dict(color=TEXT_COLOR),
    )
    return fig


def _clip_half_plane(poly, f, keep_positive):
    """Sutherland-Hodgman: ritaglia un poligono contro il semipiano f>=0
    (keep_positive) oppure f<=0. f(x, y) è lineare, quindi l'intersezione con
    ogni lato si trova per interpolazione lineare dei valori di f."""
    out = []
    for i in range(len(poly)):
        cur, prv = poly[i], poly[i - 1]
        fc, fp = f(*cur), f(*prv)
        inc = fc >= 0 if keep_positive else fc <= 0
        inp = fp >= 0 if keep_positive else fp <= 0
        if inc != inp and (fp - fc) != 0:
            t = fp / (fp - fc)
            out.append((prv[0] + t * (cur[0] - prv[0]), prv[1] + t * (cur[1] - prv[1])))
        if inc:
            out.append(cur)
    return out


def _poly_path(poly):
    if len(poly) < 3:
        return None
    return "M " + " L ".join(f"{x},{y}" for x, y in poly) + " Z"


def _poly_centroid(poly):
    """Baricentro del poligono (per posizionare l'etichetta della zona)."""
    n, a, cx, cy = len(poly), 0.0, 0.0, 0.0
    for i in range(n):
        x0, y0 = poly[i - 1]
        x1, y1 = poly[i]
        cr = x0 * y1 - x1 * y0
        a += cr
        cx += (x0 + x1) * cr
        cy += (y0 + y1) * cr
    if a == 0:
        return sum(p[0] for p in poly) / n, sum(p[1] for p in poly) / n
    return cx / (3 * a), cy / (3 * a)


def ratio_wedge_chart(x_mean, y_mean, x_sd, y_sd, x_label, y_label, key,
                      thr_low, thr_high, pop_x=None, pop_y=None,
                      point_color=PRIMARY, height=470):
    """Grafico a bande per un indice-rapporto (indice = x / y).

    A differenza di quadrant_chart le zone NON sono i quattro quadranti del
    crosshair, ma tre spicchi delimitati dalle rette y = x/soglia uscenti
    dall'origine: essendo l'indice un rapporto, tutti i punti con lo stesso
    indice stanno su una retta per l'origine, quindi è la posizione rispetto
    alle diagonali a rappresentare l'indice. Le linee tratteggiate della media
    di popolazione restano come contesto sul livello assoluto e vengono
    disegnate SOLO se quel dato di popolazione esiste davvero.

    L'origine può restare fuori dal range: gli spicchi sono ritagliati sul
    riquadro visibile. Il range verticale viene però allargato quanto basta
    perché entrambe le diagonali attraversino il riquadro, altrimenti un
    atleta lontano dalle soglie vedrebbe un grafico tinto di un solo colore,
    senza alcun confine visibile su cui orientarsi."""
    if x_mean is None or y_mean is None or not thr_low or not thr_high:
        return None
    x_sd, y_sd = x_sd or 0, y_sd or 0
    cx = pop_x if pop_x is not None else x_mean
    cy = pop_y if pop_y is not None else y_mean

    x_ext, y_ext = abs(x_mean - cx) + x_sd, abs(y_mean - cy) + y_sd
    x_r = x_ext * 1.3 if x_ext > 0 else (abs(cx) * 0.15 or 1)
    y_r = y_ext * 1.3 if y_ext > 0 else (abs(cy) * 0.15 or 1)
    for thr in (thr_low, thr_high):
        y_r = max(y_r, abs(cx / thr - cy) * 1.08)
    x0, x1 = cx - x_r, cx + x_r
    y0, y1 = cy - y_r, cy + y_r

    box = [(x0, y0), (x1, y0), (x1, y1), (x0, y1)]
    f_low = lambda x, y: y - x / thr_low
    f_high = lambda x, y: y - x / thr_high
    zones = [
        (_clip_half_plane(box, f_low, True), 0),
        (_clip_half_plane(_clip_half_plane(box, f_low, False), f_high, True), 1),
        (_clip_half_plane(box, f_high, False), 2),
    ]
    labels = INDEX_ZONE_LABELS[key]

    fig = go.Figure()
    for poly, i in zones:
        path = _poly_path(poly)
        if not path:
            continue
        fig.add_shape(type="path", path=path, fillcolor=INDEX_ZONE_COLORS[i],
                      opacity=0.18, line_width=0, layer="below")
        lx, ly = _poly_centroid(poly)
        fig.add_annotation(x=lx, y=ly, text=_wrap_label(labels[i], 20), showarrow=False,
                           font=dict(size=10, color="#666"), align="center")

    for thr in (thr_low, thr_high):
        fig.add_shape(type="line", x0=x0, y0=x0 / thr, x1=x1, y1=x1 / thr,
                      line=dict(color="#8d8d8d", dash="dot", width=3))

    # Le linee di popolazione compaiono solo dove il dato esiste: oggi
    # cmj_peak_force non ha una norma, quindi il DSI mostra la sola
    # orizzontale (IMTP). Aggiungendo la norma mancante a DEFAULT_POP /
    # CONST_LABELS e impostando pop_key sulla metrica, la verticale
    # comparirà da sola senza altre modifiche.
    # Peso visivo: le diagonali (che rappresentano l'indice) sono spesse,
    # il crosshair di popolazione è sottile — è l'indice il soggetto del
    # grafico, la popolazione è solo contesto.
    if pop_x is not None:
        fig.add_vline(x=pop_x, line_dash="dash", line_color=ACCENT, line_width=1)
    if pop_y is not None:
        fig.add_hline(y=pop_y, line_dash="dash", line_color=ACCENT, line_width=1)

    fig.add_trace(go.Scatter(
        x=[x_mean], y=[y_mean], mode="markers",
        marker=dict(size=12, color=point_color, line=dict(width=1.5, color="white")),
        error_x=dict(type="data", array=[x_sd], visible=True, color=point_color, thickness=1.5, width=6),
        error_y=dict(type="data", array=[y_sd], visible=True, color=point_color, thickness=1.5, width=6),
        name="Media test \u00b1 Dev.Std",
        hovertemplate=(f"Media test<br>{x_label}: %{{x:.2f}} \u00b1 {x_sd:.2f}"
                       f"<br>{y_label}: %{{y:.2f}} \u00b1 {y_sd:.2f}<extra></extra>"),
    ))

    fig.update_layout(
        xaxis_title=x_label, yaxis_title=y_label,
        xaxis_range=[x0, x1], yaxis_range=[y0, y1],
        xaxis=dict(showgrid=False, automargin=True, title=dict(standoff=12)),
        yaxis=dict(showgrid=False, automargin=True, title=dict(standoff=12)),
        height=height, margin=dict(t=30, b=50, l=70, r=20), showlegend=False,
        plot_bgcolor=BG_COLOR, paper_bgcolor=BG_COLOR, font=dict(color=TEXT_COLOR),
    )
    return fig


def ratio_band_strip(key, value, thr_low, thr_high, label, decimals=2, height=120):
    """Barra orizzontale a tre bande con il valore dell'indice: la zona si
    legge a colpo d'occhio, senza dover interpretare una pendenza. Range =
    soglie ± 0.2, allargato automaticamente se il valore cade fuori.

    Lo spessore della fascia NON si imposta direttamente: add_vrect riempie
    sempre tutta l'area di plot, quindi la fascia è alta esattamente
    (height - margin.t - margin.b) = 10px. Per spostare le etichette senza
    ingrassare la fascia si alzano margin_t E height della stessa quantità.
    Il marcatore (13px) e il suo valore sono più alti della fascia: senza
    cliponaxis=False verrebbero tagliati dai bordi dell'area di plot."""
    if value is None or not thr_low or not thr_high:
        return None
    lo = min(thr_low - 0.2, value - 0.05)
    hi = max(thr_high + 0.2, value + 0.05)
    labels = INDEX_ZONE_LABELS[key]

    fig = go.Figure()
    for a, b, i in ((lo, thr_low, 0), (thr_low, thr_high, 1), (thr_high, hi, 2)):
        fig.add_vrect(x0=a, x1=b, fillcolor=INDEX_ZONE_COLORS[i], opacity=0.35, line_width=0)
        # Le etichette sono annotazioni in coordinate "paper" sopra l'area di
        # plot: non partecipano all'automargin, quindi lo spazio per loro va
        # riservato a mano con margin_t. Wrap a 22 caratteri = due righe al
        # massimo per le etichette attuali.
        fig.add_annotation(x=(a + b) / 2, y=1.10, xref="x", yref="paper", yanchor="bottom",
                           text=_wrap_label(labels[i], 22), showarrow=False,
                           font=dict(size=11, color="#484343"))
    fig.add_trace(go.Scatter(
        x=[value], y=[0], mode="markers+text",
        marker=dict(size=13, color=PRIMARY, symbol="diamond", line=dict(width=1.5, color="white")),
        text=[f"{value:.{decimals}f}"], textposition="bottom center",
        textfont=dict(color=TEXT_COLOR, size=13),
        cliponaxis=False,
        hovertemplate=f"{label}: %{{x:.{decimals}f}}<extra></extra>",
    ))
    fig.update_layout(
        xaxis=dict(range=[lo, hi], showgrid=False, tickvals=[thr_low, thr_high],
                   ticktext=[f"{thr_low:.2f}", f"{thr_high:.2f}"], automargin=True,
                   title=dict(text=label, standoff=12)),
        yaxis=dict(range=[-1, 1], showgrid=False, showticklabels=False, zeroline=False),
        height=height, margin=dict(t=60, b=50, l=20, r=20), showlegend=False,
        plot_bgcolor=BG_COLOR, paper_bgcolor=BG_COLOR, font=dict(color=TEXT_COLOR),
    )
    return fig


RSQ_QUADRANTS = dict(
    tl=("Alta Reattività", "#4FC3F7"),
    tr=("Forza-Dominante", "#EF5350"),
    bl=("Elastico-Dominante", "#66BB6A"),
    br=("Bassa Reattività", "#FFEE58"),
)

# Metriche (altezza, tempo) usate per il grafico RSQ di ciascuna categoria
# che include un mRSI. Definito a livello di modulo per essere condiviso
# tra la scheda Dettaglio Test e il report HTML.
RSQ_METRIC_KEYS = {
    "COUNTERMOVEMENT JUMP TEST": ("cmj_height", "cmj_contraction_time"),
    "COUNTERMOVEMENT JUMP REBOUND TEST": ("cmj_re_rebound_height", "cmj_re_contact_time"),
}


def build_rsq_chart(cat, results):
    """Grafico RSQ per una categoria, se applicabile. Ritorna None altrimenti."""
    if cat not in RSQ_METRIC_KEYS:
        return None
    h_key, t_key = RSQ_METRIC_KEYS[cat]
    h_metric = next(m for m in METRICS if m["key"] == h_key)
    t_metric = next(m for m in METRICS if m["key"] == t_key)
    r_h = next(r for r in results if r["key"] == h_key)
    r_t = next(r for r in results if r["key"] == t_key)
    return quadrant_chart(
        x_mean=r_t["mean"], y_mean=r_h["mean"], x_sd=r_t["sd"], y_sd=r_h["sd"],
        x_label=f"{t_metric['label']} ({t_metric['unit']})",
        y_label=f"{h_metric['label']} ({h_metric['unit']})",
        quadrant_defs=RSQ_QUADRANTS, pop_x=r_t["pop_mean"], pop_y=r_h["pop_mean"],
        point_color=PRIMARY,
    )


def build_dsi_chart(results, thresholds):
    """DSI: CMJ Peak Propulsive Force (x) vs IMTP Peak Force (y). La linea
    verticale compare solo quando cmj_peak_force avrà un dato di popolazione
    (oggi pop_key=None -> pop_mean None -> linea nascosta)."""
    r_cmj_peak = next((r for r in results if r["key"] == "cmj_peak_force"), None)
    r_imtp_peak = next((r for r in results if r["key"] == "imtp_peak_force"), None)
    if r_cmj_peak is None or r_imtp_peak is None:
        return None
    thr_low, thr_high = thresholds["dsi"]
    return ratio_wedge_chart(
        x_mean=r_cmj_peak["mean"], y_mean=r_imtp_peak["mean"],
        x_sd=r_cmj_peak["sd"], y_sd=r_imtp_peak["sd"],
        x_label="CMJ Peak Propulsive Force (N)", y_label="IMTP Peak Force (N)",
        key="dsi", thr_low=thr_low, thr_high=thr_high,
        pop_x=r_cmj_peak["pop_mean"], pop_y=r_imtp_peak["pop_mean"],
    )


def build_eur_chart(results, thresholds):
    """EUR: CMJ Height (x) vs SJ Height (y). Entrambe hanno un dato di
    popolazione, quindi il crosshair è completo."""
    r_sj_h = next((r for r in results if r["key"] == "sj_height"), None)
    r_cmj_h = next((r for r in results if r["key"] == "cmj_height"), None)
    if r_sj_h is None or r_cmj_h is None:
        return None
    thr_low, thr_high = thresholds["eur"]
    return ratio_wedge_chart(
        x_mean=r_cmj_h["mean"], y_mean=r_sj_h["mean"], x_sd=r_cmj_h["sd"], y_sd=r_sj_h["sd"],
        x_label="CMJ Height (cm)", y_label="SJ Height (cm)",
        key="eur", thr_low=thr_low, thr_high=thr_high,
        pop_x=r_cmj_h["pop_mean"], pop_y=r_sj_h["pop_mean"],
    )


def build_tscore_bar_chart(cat_results):
    """Grafico a barre orizzontali del T-score per una categoria (usato sia
    nella scheda Dettaglio Test sia nel report HTML). Ritorna None se nessuna
    metrica della categoria ha un T-score calcolabile."""
    scored = [r for r in cat_results if r["t"] is not None]
    if not scored:
        return None

    scored_sorted = sorted(scored, key=lambda r: r["t"])
    labels = [r["label"] for r in scored_sorted]
    tvals = [r["t"] for r in scored_sorted]
    colors = [r["colore"] for r in scored_sorted]

    fig = go.Figure()

    for lo, hi, band_label, band_color in BANDS:
        lo_c, hi_c = max(lo, 0), min(hi, 100)
        if lo_c >= hi_c:
            continue
        # Il rettangolo di sfondo (banda) resta ancorato ai dati (asse x),
        # ma l'etichetta viene disegnata separatamente appena sopra l'area
        # di plot (yref="paper", y>1) invece che dentro di essa: così non
        # si sovrappone mai alla barra più in alto, indipendentemente dal
        # numero di metriche mostrate.
        fig.add_vrect(x0=lo_c, x1=hi_c, fillcolor=band_color, opacity=0.3, line_width=0)
        fig.add_annotation(
            x=(lo_c + hi_c) / 2, y=1.02, xref="x", yref="paper", yanchor="bottom",
            text=band_label, showarrow=False, font=dict(size=9, color="#484343"),
        )

    fig.add_trace(go.Bar(
        x=tvals, y=labels, base=0, orientation="h",
        marker_color=colors, customdata=tvals,
        text=[f"{t:.0f}" for t in tvals], textposition="outside",
        textfont=dict(color="#484343", size=12),
        hovertemplate="%{y}: T-score %{customdata:.0f}<extra></extra>",
    ))
    fig.add_vline(x=50, line_dash="dash", line_color=ACCENT)
    fig.update_layout(
        xaxis_title="T-score", xaxis_range=[0, 100],
        xaxis=dict(showgrid=False, ticks="", automargin=True, title=dict(standoff=18)),
        yaxis=dict(showgrid=False, automargin=True),
        height=max(360, 70 * len(scored_sorted) + 150),
        margin=dict(t=70, b=60, l=10, r=20), showlegend=False,
        plot_bgcolor=BG_COLOR, paper_bgcolor=BG_COLOR,
        font=dict(color=TEXT_COLOR),
    )
    return fig


def build_radar_chart(cats_valide, profilo, nome):
    """Radar (Scatterpolar) del profilo di forza (usato sia nella scheda
    Profilo di Forza sia nel report HTML). Ritorna None se non ci sono
    categorie valide."""
    if not cats_valide:
        return None
    vals = [profilo[c] for c in cats_valide]
    vals_closed = vals + [vals[0]]
    cats_closed = [_wrap_label(c) for c in cats_valide] + [_wrap_label(cats_valide[0])]

    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=[50] * (len(cats_valide) + 1), theta=cats_closed, mode="lines",
        line=dict(color="rgba(233,74,38,0.6)", dash="dash"), name="Media popolazione (T=50)"
    ))
    fig.add_trace(go.Scatterpolar(
        r=vals_closed, theta=cats_closed, fill="toself",
        line=dict(color=PRIMARY, width=3), fillcolor="rgba(11,182,255,0.25)", name=nome
    ))
    fig.update_layout(
        polar=dict(
            domain=dict(x=[0.12, 0.88], y=[0.08, 0.92]),
            radialaxis=dict(range=[0, 100], showticklabels=True, ticks="", tickfont=dict(color=TEXT_COLOR)),
        ),
        font=dict(color=TEXT_COLOR),
        paper_bgcolor=BG_COLOR,
        showlegend=True, height=560, margin=dict(t=60, b=60, l=60, r=60),
    )
    return fig


if parsed_files:
    meta0 = parsed_files[0].metadata
    nome = meta0.get("nome") or "Atleta"
    sesso = sesso_da_file(parsed_files) or "-"
    if meta0.get("periodo_override"):
        periodo = meta0["periodo_override"]
    else:
        date_tests = [pf.metadata.get("data_test") for pf in parsed_files if pf.metadata.get("data_test")]
        data_min = min(date_tests).strftime("%d/%m/%Y") if date_tests else "-"
        data_max = max(date_tests).strftime("%d/%m/%Y") if date_tests else "-"
        periodo = data_min if data_min == data_max else f"{data_min} → {data_max}"
else:
    nome, sesso, periodo = "Atleta", "-", "-"


# ============================================================================
# PARTE 5 — REPORT LIVE (interfaccia a schede)
# ============================================================================

if parsed_files:
    st.title(f"🏋️ Force Plate Test Report — {nome}")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Sesso", sesso)
    c2.metric("Data test", periodo)
    c3.metric("File caricati", len(parsed_files))
    c4.metric("Ripetizioni totali", sum(len(pf.reps) for pf in parsed_files))
else:
    st.title("🏋️ Force Plate Test Report")
    st.info(
        "Carica uno o più file XLSX esportati da ForceMate (IMTP, SJ, CMJ, CMJ RE) dalla "
        "barra laterale per vedere il dettaglio dei test, il profilo di forza e generare il "
        "report. Nel frattempo puoi consultare e modificare le costanti di popolazione nella "
        "scheda '⚙️ Costanti'."
    )

tab_costanti, tab_dettaglio, tab_profilo, tab_comparazione, tab_report = st.tabs(
    ["⚙️ Costanti", "🔍 Dettaglio Test", "📊 Profilo di Forza", "🔀 Comparazione", "📄 Report"]
)

with tab_costanti:
    st.markdown(
        "Valori di riferimento della popolazione (media e deviazione standard, per uomini e "
        "donne) usati per calcolare i T-score. Modificabili direttamente nella tabella; le "
        "modifiche si applicano subito alle altre schede."
    )
    st.caption(
        "Nota: per DSI ed EUR la media di popolazione non viene usata per il T-score ma è utile come riferimento e "
        "viene mostrata accanto al valore dell'atleta. Le soglie di lettura di DSI ed EUR si "
        "modificano qui sotto."
    )

    df_pop = costanti_dataframe(st.session_state["pop"])
    edited_pop = st.data_editor(
        df_pop, use_container_width=True, hide_index=True, key="pop_editor",
        column_config={
            "Categoria": st.column_config.TextColumn(disabled=True),
            "Costante": st.column_config.TextColumn(disabled=True),
            "Unità": st.column_config.TextColumn(disabled=True),
            "Media Uomini": st.column_config.NumberColumn(format="%.4f"),
            "Dev.Std Uomini": st.column_config.NumberColumn(format="%.4f", min_value=0.0),
            "Media Donne": st.column_config.NumberColumn(format="%.4f"),
            "Dev.Std Donne": st.column_config.NumberColumn(format="%.4f", min_value=0.0),
        },
    )
    for key, row in zip(CONST_LABELS.keys(), edited_pop.to_dict("records")):
        mm, sm, mf, sf = row["Media Uomini"], row["Dev.Std Uomini"], row["Media Donne"], row["Dev.Std Donne"]
        if None not in (mm, sm, mf, sf):
            st.session_state["pop"][key] = dict(mean_m=float(mm), sd_m=float(sm), mean_f=float(mf), sd_f=float(sf))

    # --- Soglie degli indici a rapporto (DSI, EUR) ---
    # Stanno qui, e non accanto ai grafici, perché sono costanti di lettura
    # esattamente come le norme di popolazione sopra. Il blocco viene
    # eseguito PRIMA di build_results(), quindi una modifica si applica
    # nello stesso rerun (nessun click aggiuntivo).
    st.markdown("---")
    st.markdown("**Soglie di lettura degli indici (DSI, EUR)**")
    st.caption(
        "Delimitano le tre zone di profilo nei grafici di DSI ed EUR. Default di letteratura: "
        "DSI 0.60 / 0.80, EUR 1.00 / 1.10. Nota: le soglie EUR di letteratura cadono quasi sulla "
        "media di popolazione qui usata (1.108 U / 1.091 D), quindi un atleta 'medio' finisce sul "
        "confine superiore — da valutare una ricalibrazione sui propri dati."
    )
    tc1, tc2, tc3, tc4 = st.columns(4)
    dsi_low = tc1.number_input("DSI — soglia bassa", key="thr_dsi_low", min_value=0.0, step=0.05,
                               format="%.2f", value=float(st.session_state["idx_thr"]["dsi"][0]))
    dsi_high = tc2.number_input("DSI — soglia alta", key="thr_dsi_high", min_value=0.0, step=0.05,
                                format="%.2f", value=float(st.session_state["idx_thr"]["dsi"][1]))
    eur_low = tc3.number_input("EUR — soglia bassa", key="thr_eur_low", min_value=0.0, step=0.05,
                               format="%.2f", value=float(st.session_state["idx_thr"]["eur"][0]))
    eur_high = tc4.number_input("EUR — soglia alta", key="thr_eur_high", min_value=0.0, step=0.05,
                                format="%.2f", value=float(st.session_state["idx_thr"]["eur"][1]))
    for _idx_key, _lo, _hi in (("dsi", dsi_low, dsi_high), ("eur", eur_low, eur_high)):
        if _hi <= _lo:
            st.warning(f"{_idx_key.upper()}: la soglia alta deve essere maggiore della soglia bassa "
                       "(soglie non aggiornate).")
        else:
            st.session_state["idx_thr"][_idx_key] = [_lo, _hi]

    st.markdown("---")
    col_dl, col_ul, col_reset = st.columns([1, 1, 1])
    with col_dl:
        st.markdown("**Scarica costanti**")
        xlsx_bytes = genera_costanti_xlsx(st.session_state["pop"])
        st.download_button(
            "⬇️ Scarica (.xlsx)", data=xlsx_bytes, file_name="costanti_force_plate.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col_ul:
        st.markdown("**Ricarica da file**")
        const_file = st.file_uploader(
            "Formato: DATO / UdM / MEDIA / DEV ST / MEDIA2 / DEV ST3", type=["xlsx"], key="const_uploader"
        )
        if const_file is not None:
            updates, unmatched = parse_constants_workbook(io.BytesIO(const_file.getvalue()))
            if updates:
                st.session_state["pop"].update(updates)
                if "pop_editor" in st.session_state:
                    del st.session_state["pop_editor"]
                st.success(f"Aggiornate {len(updates)} costanti da '{const_file.name}'.")
                if unmatched:
                    st.caption("Righe non riconosciute: " + ", ".join(unmatched))
                st.rerun()
            else:
                st.warning("Nessuna costante riconosciuta nel file caricato. Verifica che segua il formato indicato.")
    with col_reset:
        st.markdown("**Ripristina**")
        if st.button("↺ Valori di default"):
            st.session_state["pop"] = {k: dict(v) for k, v in DEFAULT_POP.items()}
            st.session_state["idx_thr"] = {k: list(v) for k, v in DEFAULT_INDEX_THRESHOLDS.items()}
            for _k in ("pop_editor", "thr_dsi_low", "thr_dsi_high", "thr_eur_low", "thr_eur_high"):
                if _k in st.session_state:
                    del st.session_state[_k]
            st.rerun()

if parsed_files:
    results_bundle = build_results(parsed_files, st.session_state["pop"], st.session_state["idx_thr"])
    results = results_bundle["results"]
    checks = results_bundle["checks"]
    profilo = profilo_forza(results)
else:
    results, checks, profilo = [], [], {}

with tab_dettaglio:
    if not parsed_files:
        st.info("Carica dei file dalla barra laterale per vedere il dettaglio dei test.")
    else:
        for cat in CATEGORIES:
            cat_results = [r for r in results if r["category"] == cat and r["mean"] is not None]
            if not cat_results:
                continue
            st.markdown(f"### {cat}")

            # Grafico a barre orizzontali che partono da sinistra (0): la
            # lunghezza della barra è il T-score assoluto. Le bande di
            # valutazione sono mostrate come contesto sullo sfondo (vedi
            # BANDS) e la media di popolazione (T-score = 50) resta segnata
            # da una linea tratteggiata di riferimento.
            tscore_fig = build_tscore_bar_chart(cat_results)
            if tscore_fig:
                st.plotly_chart(tscore_fig, use_container_width=True)

            # --- Salti individuali: inclusione e valori per colonna ---
            cat_metrics = [m for m in METRICS if m["category"] == cat and m.get("jump_type")]
            jump_type = cat_metrics[0]["jump_type"] if cat_metrics else None
            reps_all = collect_reps_all(parsed_files, jump_type) if jump_type else []
            n_reps = len(reps_all)

            if n_reps == 0:
                st.info("Nessuna ripetizione disponibile per questa categoria.")
                st.markdown("---")
                continue

            # --- Metriche extra: ricerca libera su tutte le altre variabili
            # presenti nel file ForceDecks per questo tipo di test, oltre a
            # quelle curate sopra. Non avendo un dato di popolazione non
            # hanno T-score, ma vengono comunque calcolate (media, dev.std,
            # CV%) sulle stesse ripetizioni incluse/escluse qui sotto.
            existing_raw_vars = {m["raw_var"] for m in cat_metrics if m.get("raw_var")}
            catalog_options = [v for v in EXTRA_METRICS_CATALOG.get(jump_type, []) if v not in existing_raw_vars]
            extra_selected = st.multiselect(
                "🔎 Aggiungi metriche extra (cerca per nome)", options=catalog_options,
                key=f"extra_metrics_{jump_type}",
                help="Tutte le altre variabili disponibili nell'export ForceDecks per questo tipo di test. "
                     "Non avendo un dato di popolazione, sono mostrate senza T-score.",
            )
            extra_metric_defs = [
                dict(key=f"extra::{jump_type}::{name}", label=extra_metric_label(name), category=cat,
                     jump_type=jump_type, raw_var=name, unit="", pop_key=None,
                     lower_is_better=False, kind="info", desc=get_metric_description(jump_type, name))
                for name in extra_selected
            ]

            st.markdown("**Ripetizioni incluse nel calcolo**")
            st.caption("Deseleziona un salto per escluderlo da media, dev.std e T-score di questa categoria.")
            chk_cols = st.columns(n_reps)
            incl_mask = []
            for i in range(n_reps):
                key = f"incl_{jump_type}_{i}"
                checked = chk_cols[i].checkbox(f"Prova {i + 1}", value=st.session_state.get(key, True), key=key)
                incl_mask.append(checked)

            extra_results = []
            for m in extra_metric_defs:
                stats = pooled_stats(compute_metric_values(parsed_files, m))
                extra_results.append(dict(
                    key=m["key"], label=m["label"], category=cat, unit=m["unit"], kind="info",
                    n=stats["n"], mean=stats["mean"], sd=stats["sd"], z=None, t=None, banda=None,
                    colore=None, pop_mean=None, pop_sd=None,
                ))
            empty_extra_labels = [r["label"] for r in extra_results if r["n"] == 0]
            if empty_extra_labels:
                st.warning(
                    "⚠️ Nessun dato trovato per: " + ", ".join(empty_extra_labels) + ". La variabile "
                    "potrebbe non essere presente in questo export ForceDecks (es. campo non calcolato "
                    "dal software per questo tipo di test)."
                )
            results_lookup = results + extra_results
            cat_metrics = cat_metrics + extra_metric_defs

            rows, best_per_row, worst_per_row = [], [], []
            cv_warnings = []
            metric_descriptions = []  # (label, desc) per l'expander sotto la tabella
            jump_cols = [f"Prova {i + 1}" for i in range(n_reps)]
            for m in cat_metrics:
                r = next((x for x in results_lookup if x["key"] == m["key"]), None)
                values = per_rep_metric_values(parsed_files, m)
                best_i, worst_i = best_worst_indices(values, incl_mask, m["lower_is_better"])
                best_per_row.append(best_i)
                worst_per_row.append(worst_i)
                desc = m.get("desc")
                if desc:
                    metric_descriptions.append((m["label"], desc))
                row = {"Metrica": m["label"], "Unità": m["unit"]}
                for i, v in enumerate(values):
                    row[jump_cols[i]] = round(v, 3) if isinstance(v, (int, float)) else None
                row["Media"] = round(r["mean"], 3) if r and r["mean"] is not None else None
                row["Media Pop."] = round(r["pop_mean"], 3) if r and r["pop_mean"] is not None else "—"
                row["Dev.Std"] = round(r["sd"], 3) if r and r["sd"] else None
                # CV% = coefficiente di variazione (dev.std / media * 100): misura
                # la variabilità relativa tra le ripetizioni incluse.
                cv = None
                if r and r["mean"] not in (None, 0) and r["sd"] is not None:
                    cv = abs(r["sd"] / r["mean"]) * 100
                row["CV%"] = round(cv, 1) if cv is not None else None
                if cv is not None and cv > 10:
                    cv_warnings.append(m["label"])
                row["T-score"] = round(r["t"], 1) if r and r["t"] is not None else "—"
                row["Valutazione"] = (r["banda"] if r and r["banda"] else "—")
                rows.append(row)

            df = pd.DataFrame(rows).reset_index(drop=True)

            def _highlight(row):
                styles = [""] * len(row)
                ridx = row.name
                best_i, worst_i = best_per_row[ridx], worst_per_row[ridx]
                for i, col in enumerate(jump_cols):
                    pos = df.columns.get_loc(col)
                    if not incl_mask[i]:
                        styles[pos] = "color:#adb5bd; text-decoration: line-through;"
                    elif i == best_i:
                        styles[pos] = "background-color:#3b8f3d; font-weight:600;"
                    elif i == worst_i:
                        styles[pos] = "background-color:#d13242; font-weight:600;"
                cv_val = row.get("CV%")
                if isinstance(cv_val, (int, float)) and cv_val > 10:
                    styles[df.columns.get_loc("CV%")] = "background-color:#E4572E; color:white; font-weight:600;"
                return styles

            styled = df.style.apply(_highlight, axis=1)
            st.dataframe(styled, use_container_width=True, hide_index=True)
            if cv_warnings:
                st.warning(
                    "⚠️ Coefficiente di variazione (CV%) superiore al 10% per: " + ", ".join(cv_warnings) +
                    ". Indica un'elevata variabilità tra le ripetizioni incluse: valutarne l'affidabilità."
                )
            if metric_descriptions:
                with st.expander("ℹ️ Cosa significano queste metriche?"):
                    for label, desc in metric_descriptions:
                        st.markdown(f"**{label}** — {desc}")

            # --- Grafico RSQ (Reactive Strength Quadrant) per le categorie
            # che includono un mRSI: mette in relazione la media di altezza
            # del salto (asse Y) e la media del tempo di contatto/contrazione
            # (asse X), con il crosshair centrato sulla media di popolazione.
            if cat in RSQ_METRIC_KEYS:
                rsq_fig = build_rsq_chart(cat, results)
                h_key, t_key = RSQ_METRIC_KEYS[cat]
                r_t = next(r for r in results if r["key"] == t_key)
                if rsq_fig:
                    st.markdown("**RSQ — Reactive Strength Quadrant**")
                    caption = (
                        "Il grafico dà contesto all'mRSI mostrando non solo quanto l'atleta è reattivo, "
                        "ma come esprime quella reattività: le linee tratteggiate sono centrate sulla media "
                        "di popolazione, il punto è la media del test con barre d'errore (± dev.std)."
                    )
                    if r_t["pop_mean"] is None:
                        caption += (
                            " Nota: per il tempo di contatto/contrazione non è disponibile un dato di "
                            "popolazione, quindi la linea verticale usa la media del test."
                        )
                    st.caption(caption)
                    st.plotly_chart(rsq_fig, use_container_width=True)

            # I controlli a soglia riguardano esclusivamente il CMJ Rebound:
            # vengono mostrati qui, subito dopo la tabella della COUNTERMOVEMENT JUMP REBOUND TEST.
            if cat == "COUNTERMOVEMENT JUMP REBOUND TEST":
                st.markdown("#### ✅ Controlli Tecnici (CMJ Rebound)")
                st.caption("Controlli a soglia sul CMJ Rebound, per individuare asimmetrie o esecuzioni tecnicamente scorrette.")
                for c in checks:
                    ccol1, ccol2 = st.columns([3, 1])
                    with ccol1:
                        st.markdown(f"**{c['label']}**")
                        st.caption(c["desc"])
                    with ccol2:
                        if c["value"] is None:
                            st.markdown("—")
                        else:
                            icon = "✅" if c["passed"] else "⚠️"
                            val_disp = c["value"] * c["scale"]
                            th_disp = c["threshold"] * c["scale"]
                            st.markdown(f"##### {icon} {val_disp:.{c['decimals']}f}{c['suffix']}")
                            soglia_lbl = "min" if c["direction"] == "min" else "max"
                            st.caption(f"soglia {soglia_lbl} {th_disp:.{c['decimals']}f}{c['suffix']}")

            st.markdown("---")

        # --- Export unico: TUTTE le variabili presenti nei file XLSX, una
        # riga per ripetizione, escluse quelle deselezionate. È lo stesso
        # file che si ricarica dalla sidebar e che alimenta la Comparazione.
        st.markdown("### 📤 Esporta dati del test")
        st.caption(
            "Esporta OGNI variabile presente nei file caricati, una riga per ripetizione, "
            "escluse le prove deselezionate sopra. È il file da ricaricare nella barra "
            "laterale per rivedere questo test e da usare nella scheda 🔀 Comparazione."
        )
        df_full = build_full_raw_export(parsed_files, nome, sesso, periodo)
        csv_full_bytes = df_full.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "⬇️ Scarica dati completi (.csv)", data=csv_full_bytes,
            file_name=f"forceplate_fulldata_{_slug_nome(nome)}_"
                      f"{_data_file_token(parsed_files, periodo)}.csv",
            mime="text/csv",
        )

with tab_profilo:
    if not parsed_files:
        st.info("Carica dei file dalla barra laterale per vedere il profilo di forza.")
    else:
        cats_valide = [c for c in CATEGORIES if profilo.get(c) is not None]
        if not cats_valide:
            st.warning("Nessuna metrica con confronto di popolazione disponibile: carica almeno un test tra IMTP, SJ, CMJ o CMJ RE.")
        else:
            radar_fig = build_radar_chart(cats_valide, profilo, nome)
            st.plotly_chart(radar_fig, use_container_width=True)

            cols = st.columns(len(cats_valide))
            for col, cat in zip(cols, cats_valide):
                t = profilo[cat]
                banda, colore = banda_da_tscore(t)
                with col.container(border=True):
                    st.markdown(
                        f"<div style='font-size:11px;letter-spacing:0.5px;opacity:0.65'>{cat}</div>"
                        f"<div style='font-size:17px;font-weight:600;color:{TEXT_COLOR};"
                        f"margin:2px 0 4px 0'>{CATEGORY_QUALITY.get(cat, '')}</div>"
                        f"<div style='font-size:30px;font-weight:700;color:{colore}'>{t:.0f}</div>"
                        f"<div style='font-size:13px;opacity:0.75'>{banda}</div>",
                        unsafe_allow_html=True,
                    )

        # --- Indici a rapporto (DSI, EUR): barra a bande + grafico a spicchi.
        # Le soglie si modificano nella scheda ⚙️ Costanti e si applicano
        # subito anche ai report scaricabili. Le etichette delle zone
        # descrivono il PROFILO dell'atleta, non l'indicazione di
        # allenamento: quella resta una decisione del preparatore, da
        # scrivere nel campo "Analisi" della scheda Report.
        for idx_key, titolo, decimali, caption in (
            ("dsi", "DSI (Dynamic Strength Index)", 3,
             "CMJ Peak Propulsive Force vs IMTP Peak Force. Le linee tratteggiate arancioni sono la "
             "media di popolazione e il punto è la media del test con barre d'errore"),
            ("eur", "EUR (Eccentric Utilisation Ratio)", 3,
             "CMJ Height vs SJ Height. Le linee tratteggiate arancioni sono la media di popolazione."),
        ):
            r_idx = next((r for r in results if r["key"] == idx_key and r["mean"] is not None), None)
            fig_wedge = (build_dsi_chart if idx_key == "dsi" else build_eur_chart)(
                results, st.session_state["idx_thr"])
            if not fig_wedge and not r_idx:
                continue

            st.markdown(f"#### {titolo}")
            if r_idx:
                help_parts = []
                if r_idx.get("zona"):
                    help_parts.append(r_idx["zona"])
                if r_idx["pop_mean"] is not None:
                    help_parts.append(f"Media popolazione: {r_idx['pop_mean']:.3f}")
                lo, hi = st.session_state["idx_thr"][idx_key]
                help_parts.append(f"Soglie {lo:.2f} / {hi:.2f} (modificabili in ⚙️ Costanti)")
                st.metric(idx_key.upper(), f"{r_idx['mean']:.{decimali}f}",
                          help=" — ".join(help_parts))
                strip = ratio_band_strip(idx_key, r_idx["mean"], lo, hi,
                                         label=idx_key.upper(), decimals=decimali)
                if strip:
                    st.plotly_chart(strip, use_container_width=True)
            if fig_wedge:
                st.plotly_chart(fig_wedge, use_container_width=True)
                st.caption(caption)

# ============================================================================
# PARTE 5bis — COMPARAZIONE TRA SESSIONI
# ============================================================================
# Confronta il test attualmente caricato con lo storico dell'atleta, caricato
# come CSV prodotti da "Dettaglio Test → Esporta TUTTE le metriche grezze".
#
# Due confronti con statuto diverso:
#
#  - CONTRO LA MEDIA delle sedute precedenti (esclusa l'attuale): è il
#    confronto statistico. La media di k sedute è più precisa della singola,
#    quindi l'incertezza si stringe: SE = TE * sqrt(1 + 1/k) invece di
#    TE * sqrt(2) del testa a testa.
#
#  - CONTRO IL MIGLIOR VALORE storico: puramente DESCRITTIVO, senza giudizio
#    di significatività. Il massimo di k misure rumorose è per costruzione
#    una sovrastima (il record è quasi sempre il giorno in cui l'atleta era
#    in forma E il rumore ha aiutato), e la distorsione cresce con k:
#    applicarci un test darebbe "peggiorato" in modo sistematico.
#
# L'app non dice MAI se un cambiamento è un bene: mostra direzione (freccia)
# e ampiezza (scala SWC di Hopkins). Il giudizio di merito resta al
# preparatore, che è l'unico a conoscere obiettivo e contesto.
#
#   TE  = errore tipico between-day (rumore della misura), in CV%
#   SE  = TE * sqrt(1 + 1/k)           incertezza della differenza dalla media
#   IC  = delta ± 1.645 * SE           intervallo di confidenza al 90%
#   SWC = 0.2 * SD                     soglia di rilevanza pratica

Z90 = 1.645
SWC_FACTOR = 0.2
MIN_SESSIONS_TE_ATLETA = 4  # servono almeno 3 differenze consecutive

# CV% between-day di letteratura, usati finché non ci sono abbastanza
# sessioni per stimare il TE sull'atleta stesso. Il primo pattern che
# matcha vince: ordine dal più specifico al più generico.
DEFAULT_TE_CV = [
    (("asimmetria", "asymmetry", "sym. index", "symmetry index"), 20.0),
    (("rfd",), 15.0),
    (("dsi", "eur"), 10.0),
    (("eccentric", "eccentrica", "braking", "frenata", "unweight"), 9.0),
    (("rsi",), 7.0),
    (("contraction time", "contact time", "time to", "duration", "durata", "time"), 7.0),
    (("power", "potenza"), 5.0),
    (("impulse", "impulso", "work"), 5.0),
    (("velocity", "velocità"), 4.0),
    (("height", "altezza"), 4.0),
    (("force", "forza", "torque"), 4.0),
    (("mass", "weight", "peso"), 1.5),
]
FALLBACK_TE_CV = 8.0

# Indici di simmetria: il segno indica il lato dominante, non la qualità,
# quindi vanno confrontati in valore assoluto (da -2% a +6% è un aumento
# dello sbilanciamento, non una diminuzione).
COMP_ABS_KEYWORDS = ("sym. index", "symmetry index", "asimmetria", "asymmetry")
COMP_ABS_SUFFIXES = (" si",)

# Scala di ampiezza di Hopkins, in multipli di SWC (= 0,2 SD, quindi
# 1x = 0.2 SD, 3x = 0.6, 6x = 1.2, 10x = 2.0 deviazioni standard).
# Il livello "Trascurabile" sotto 1x non è nella scala originale ma serve:
# senza, i cambiamenti minuscoli non avrebbero etichetta.
SWC_SCALE = [
    (1.0, "Trascurabile", "#e4e4dd"),
    (3.0, "Piccolo", "#cfe3f7"),
    (6.0, "Moderato", "#9ecbf0"),
    (10.0, "Grande", "#4FC3F7"),
    (float("inf"), "Molto grande", "#0bb6ff"),
]

# Semaforo sull'attendibilita' del cambiamento, non sulla sua desiderabilita':
# verde = c'e' stato un cambiamento reale; giallo = non si puo' dire; rosso =
# quasi certamente nessun cambiamento rilevante.
ESITO_REALE = "\U0001F7E2 Alta"
ESITO_INCERTO = "\U0001F7E1 Incerta"
ESITO_STABILE = "\U0001F534 Bassa"
ESITO_ND = "\u26A0\uFE0F Non valutabile"
ESITO_COLORI = {
    ESITO_REALE: "#2E7D32",
    ESITO_INCERTO: "#B8860B",
    ESITO_STABILE: "#B00020",
    ESITO_ND: "#8d8d8d",
}

# Colonne di intestazione dell'export completo: tutto il resto è variabile.
COMP_META_COLS = {"Nome", "Sesso", "Data test", "File", "Jump Type", "Indice Ripetizione"}

# Bridge tra variabili grezze e catalogo curato: dove esiste una metrica
# curata si usano la sua etichetta e la sua norma di popolazione.
COMP_CURATED_BY_RAW = {
    (m["jump_type"], m["raw_var"]): m
    for m in METRICS if m.get("jump_type") and m.get("raw_var")
}
COMP_DERIVED = [m for m in METRICS if m.get("jump_type") and m.get("derive")]

COMP_DEFAULT_PRIMARY_KEYS = (
    "imtp_peak_force", "sj_height", "cmj_height", "mrsi_cmj", "cmj_re_rebound_height",
)


def _metric_id(jump_type, var_name):
    """Identità di una metrica tra sessioni: tipo di test + nome variabile.
    Il tipo serve perché la stessa variabile grezza (es. 'peak force')
    esiste in test diversi con significati diversi."""
    return f"{jump_type}||{var_name}"


def _metric_label(jump_type, var_name):
    m = COMP_CURATED_BY_RAW.get((jump_type, var_name))
    if m:
        return m["label"]
    m = next((x for x in COMP_DERIVED
              if x["key"] == var_name and x["jump_type"] == jump_type), None)
    if m:
        return m["label"]
    return extra_metric_label(var_name)


def _metric_unit(jump_type, var_name):
    m = COMP_CURATED_BY_RAW.get((jump_type, var_name))
    if m:
        return m.get("unit") or ""
    m = next((x for x in COMP_DERIVED
              if x["key"] == var_name and x["jump_type"] == jump_type), None)
    return (m.get("unit") or "") if m else ""


def is_symmetry_metric(label):
    lab = str(label).strip().lower()
    return any(k in lab for k in COMP_ABS_KEYWORDS) or lab.endswith(COMP_ABS_SUFFIXES)


def comp_lower_is_better(jump_type, var_name):
    """Solo per scegliere QUALE valore storico è il migliore. Non viene
    usato per giudicare il cambiamento: quello lo fa il preparatore."""
    m = COMP_CURATED_BY_RAW.get((jump_type, var_name))
    if m is None:
        m = next((x for x in COMP_DERIVED
                  if x["key"] == var_name and x["jump_type"] == jump_type), None)
    if m is not None:
        return bool(m["lower_is_better"]), True
    if is_symmetry_metric(var_name):
        return True, True
    return False, False


def te_cv_default(label):
    """CV% between-day di letteratura per una metrica, dal suo nome."""
    lab = str(label).lower()
    for keys, cv in DEFAULT_TE_CV:
        if any(k in lab for k in keys):
            return cv
    return FALLBACK_TE_CV


def _sd_campionaria(values):
    n = len(values)
    if n < 2:
        return None
    m = sum(values) / n
    return math.sqrt(sum((v - m) ** 2 for v in values) / (n - 1))


def fmt_valore(v, segno=False):
    """Numero leggibile senza notazione scientifica: i decimali si adattano
    alla grandezza. "{:.3g}" trasformerebbe 1815 in "1.82e+03", illeggibile
    per un allenatore che si aspetta dei newton."""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    if not isinstance(v, (int, float)):
        return str(v)
    a = abs(v)
    dec = 0 if a >= 100 else (1 if a >= 10 else (2 if a >= 1 else 3))
    return f"{v:+.{dec}f}" if segno else f"{v:.{dec}f}"


def scala_ampiezza(rapporto):
    """(etichetta, colore) dal rapporto |delta| / SWC."""
    for soglia, etichetta, colore in SWC_SCALE:
        if rapporto < soglia:
            return etichetta, colore
    return SWC_SCALE[-1][1], SWC_SCALE[-1][2]


def comp_pop_sd_map(pop_dict, sesso_atleta):
    """{metric_id: SD di popolazione} per le metriche curate che hanno una
    norma. Le altre ricadranno sulla SD storica dell'atleta."""
    is_female = bool(sesso_atleta) and str(sesso_atleta).upper().startswith("F")
    out = {}
    for m in METRICS:
        if not m.get("pop_key"):
            continue
        sd = pop_dict[m["pop_key"]]["sd_f" if is_female else "sd_m"]
        if not sd:
            continue
        if m.get("jump_type") and m.get("raw_var"):
            out[_metric_id(m["jump_type"], m["raw_var"])] = float(sd)
        elif m.get("jump_type") and m.get("derive"):
            out[_metric_id(m["jump_type"], m["key"])] = float(sd)
        elif m["key"] in ("dsi", "eur"):
            out[_metric_id("indici", m["key"])] = float(sd)
    return out


# ----------------------------------------------------------------------------
# Da ripetizioni a formato lungo
# ----------------------------------------------------------------------------
COMP_RATIO_IDS = {_metric_id("indici", "dsi"), _metric_id("indici", "eur")}


def reps_to_long(reps_by_type, session_label, data_test=None, ordine=0):
    """reps_by_type: {jump_type: [dict_variabili_per_ripetizione, ...]}.

    Aggrega media/dev.std/n per ogni variabile presente, aggiunge le metriche
    derivate del catalogo (rapporti per kg) e infine gli indici cross-test
    DSI ed EUR calcolati dalle medie della sessione."""
    rows = []
    medie = {}

    def _add(jump_type, var, etichetta, stats):
        medie[(jump_type, var)] = stats["mean"]
        rows.append(dict(
            session=session_label, data=data_test, ordine=ordine, jump_type=jump_type,
            test=JUMP_TYPE_LABELS.get(jump_type, jump_type), var=var,
            metric_id=_metric_id(jump_type, var), metrica=etichetta,
            unit=_metric_unit(jump_type, var),
            mean=stats["mean"], sd=stats["sd"], n=stats["n"],
        ))

    for jump_type, reps in reps_by_type.items():
        if not jump_type or not reps:
            continue
        for var in sorted({k for rep in reps for k in rep.keys()}):
            stats = pooled_stats([rep.get(var) for rep in reps
                                  if isinstance(rep.get(var), (int, float))])
            if stats["n"]:
                _add(jump_type, var, _metric_label(jump_type, var), stats)
        for m in COMP_DERIVED:
            if m["jump_type"] != jump_type:
                continue
            stats = pooled_stats([v for v in (m["derive"](rep) for rep in reps)
                                  if isinstance(v, (int, float))])
            if stats["n"]:
                _add(jump_type, m["key"], m["label"], stats)

    # Indici cross-test: stesse formule della PARTE 4, calcolate qui sulle
    # medie di questa sessione. Essendo rapporti fra test diversi non
    # esistono "per ripetizione" e non hanno dev.std.
    cmj_peak = medie.get(("cmj", "peak propulsive force"))
    imtp_peak = medie.get(("imtp", "peak force"))
    cmj_h = medie.get(("cmj", "jump height ft"))
    sj_h = medie.get(("sj", "jump height ft"))
    for key, val in (("dsi", (cmj_peak / imtp_peak) if (cmj_peak and imtp_peak) else None),
                     ("eur", (cmj_h / sj_h) if (cmj_h and sj_h) else None)):
        if val is None:
            continue
        rows.append(dict(
            session=session_label, data=data_test, ordine=ordine, jump_type="indici",
            test="Indici", var=key, metric_id=_metric_id("indici", key),
            metrica=key.upper(), unit="", mean=val, sd=None, n=1,
        ))

    return pd.DataFrame(rows)


def current_reps_by_type(files):
    """Ripetizioni della sessione attuale, già filtrate dalle checkbox
    'Prova N' della scheda Dettaglio Test."""
    out = {}
    for jump_type in ("imtp", "sj", "cmj", "cmrj"):
        reps = collect_reps(files, jump_type)
        if reps:
            out[jump_type] = [rep["vars"] for rep in reps]
    return out


def _parse_data_test(valore):
    """La colonna 'Data test' contiene 'gg/mm/aaaa' oppure un intervallo
    'gg/mm/aaaa → gg/mm/aaaa': si prende la prima data utile."""
    if valore is None:
        return None
    testo = str(valore).strip()
    match = re.search(r"(\d{2})/(\d{2})/(\d{4})", testo)
    if not match:
        return None
    try:
        return dt.date(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    except ValueError:
        return None


def parse_full_export_csv(file_like, fallback_label, ordine=0):
    """Legge un CSV prodotto da 'Scarica export completo (.csv)':
    una riga per ripetizione, una colonna per variabile grezza."""
    df = pd.read_csv(file_like, encoding="utf-8-sig")
    if "Jump Type" not in df.columns:
        raise ValueError(
            "manca la colonna 'Jump Type'. Serve il file prodotto da "
            "\"Scarica export completo (.csv)\", non l'altro export."
        )

    etichetta, data_test = fallback_label, None
    if "Data test" in df.columns and len(df):
        v = str(df["Data test"].iloc[0]).strip()
        if v and v.lower() not in ("-", "nan", "none"):
            etichetta = v
            data_test = _parse_data_test(v)

    var_cols = [c for c in df.columns if c not in COMP_META_COLS]
    if not var_cols:
        raise ValueError("nessuna colonna di variabili trovata.")

    reps_by_type = {}
    for _, row in df.iterrows():
        jump_type = str(row["Jump Type"]).strip().lower()
        if not jump_type or jump_type in ("—", "-", "nan", "none"):
            continue
        rep = {}
        for c in var_cols:
            v = pd.to_numeric(row[c], errors="coerce")
            if pd.notna(v):
                rep[c] = float(v)
        if rep:
            reps_by_type.setdefault(jump_type, []).append(rep)

    if not reps_by_type:
        raise ValueError("nessuna ripetizione con un tipo di test valido.")
    return reps_to_long(reps_by_type, etichetta, data_test, ordine)


def ordina_sessioni(long_frames):
    """Riordina le sessioni per 'Data test'. Serve al TE dell'atleta, che si
    calcola sulle differenze tra sedute CONSECUTIVE: con i file caricati in
    ordine sparso il TE risulterebbe sbagliato. Le sessioni senza data
    leggibile restano in coda, nell'ordine di caricamento."""
    con_data, senza_data = [], []
    for frame in long_frames:
        if frame.empty:
            continue
        data = frame["data"].iloc[0]
        (con_data if data is not None else senza_data).append((data, frame))
    con_data.sort(key=lambda t: t[0])
    ordinati = [f for _, f in con_data] + [f for f in senza_data]
    for i, frame in enumerate(ordinati):
        frame["ordine"] = i
    return ordinati


# ----------------------------------------------------------------------------
# Typical Error dall'atleta
# ----------------------------------------------------------------------------
def athlete_te_table(all_long):
    """TE individuale per metrica: SD delle differenze tra sessioni
    consecutive / sqrt(2). Molto più preciso dei valori di letteratura,
    perché ogni atleta ha la propria stabilità."""
    te = {}
    if all_long is None or all_long.empty:
        return te
    for mid, g in all_long.groupby("metric_id"):
        vals = [v for v in g.sort_values("ordine")["mean"].tolist()
                if isinstance(v, (int, float)) and not math.isnan(v)]
        if len(vals) < MIN_SESSIONS_TE_ATLETA:
            continue
        if mid in COMP_RATIO_IDS and all(v > 0 for v in vals):
            vals = [math.log(v) for v in vals]
            base = 1.0  # in scala log il TE è già una frazione -> CV%
        else:
            base = sum(abs(v) for v in vals) / len(vals)
            if base == 0:
                continue
        sd_diff = _sd_campionaria([b - a for a, b in zip(vals, vals[1:])])
        if not sd_diff:
            continue
        cv = (sd_diff / math.sqrt(2)) / base * 100.0
        if cv > 0 and not math.isinf(cv):
            te[mid] = (cv, len(vals))
    return te


# ----------------------------------------------------------------------------
# Motore di confronto
# ----------------------------------------------------------------------------
def compare_to_history(cur_long, hist_long, athlete_te=None, pop_sd_map=None):
    """Confronta la sessione attuale con la MEDIA delle sedute precedenti, e
    riporta accanto il miglior valore storico (descrittivo)."""
    athlete_te = athlete_te or {}
    pop_sd_map = pop_sd_map or {}
    if cur_long.empty or hist_long.empty:
        return pd.DataFrame()

    cur = cur_long.drop_duplicates("metric_id").set_index("metric_id")

    rows = []
    for mid, g in hist_long.groupby("metric_id"):
        if mid not in cur.index:
            continue
        storiche = [v for v in g["mean"].tolist()
                    if isinstance(v, (int, float)) and not math.isnan(v)]
        if not storiche:
            continue
        m_cur = float(cur.loc[mid, "mean"])
        if math.isnan(m_cur):
            continue

        etichetta = cur.loc[mid, "metrica"]
        jump_type, var = cur.loc[mid, "jump_type"], cur.loc[mid, "var"]
        if is_symmetry_metric(etichetta) or is_symmetry_metric(var):
            m_cur, storiche = abs(m_cur), [abs(v) for v in storiche]

        k = len(storiche)
        media_rif = sum(storiche) / k
        if media_rif == 0:
            continue

        # Miglior valore storico: direzione nota solo per il catalogo curato,
        # altrove si usa il massimo. È una colonna descrittiva, senza test.
        lower_better, _ = comp_lower_is_better(jump_type, var)
        migliore = min(storiche) if lower_better else max(storiche)
        riga_migliore = g[g["mean"].apply(
            lambda v: isinstance(v, (int, float)) and abs(abs(v) - migliore) < 1e-9
            if (is_symmetry_metric(etichetta) or is_symmetry_metric(var))
            else v == migliore)]
        sess_migliore = riga_migliore["session"].iloc[0] if len(riga_migliore) else "—"

        if mid in athlete_te:
            cv, _ = athlete_te[mid]
            fonte_te = "atleta"
        else:
            cv, fonte_te = te_cv_default(etichetta), "letteratura"

        log_scale = mid in COMP_RATIO_IDS and m_cur > 0 and media_rif > 0
        if log_scale:
            delta = math.log(m_cur / media_rif)
            te_abs = cv / 100.0
            mostra = lambda x: (math.exp(x) - 1) * 100.0
        else:
            delta = m_cur - media_rif
            te_abs = cv / 100.0 * abs(media_rif)
            mostra = lambda x: x

        # SWC a tre livelli: norma di popolazione, poi variabilità storica
        # dell'atleta (per il monitoraggio individuale è anche più pertinente
        # della popolazione), poi niente -> non valutabile.
        pop_sd = pop_sd_map.get(mid)
        sd_storica = _sd_campionaria(storiche) if k >= MIN_SESSIONS_TE_ATLETA else None
        if pop_sd:
            swc = SWC_FACTOR * (pop_sd / abs(media_rif) if log_scale else pop_sd)
            fonte_swc = "norma"
        elif sd_storica:
            swc = SWC_FACTOR * (sd_storica / abs(media_rif) if log_scale else sd_storica)
            fonte_swc = "atleta"
        else:
            swc, fonte_swc = None, "—"

        # La media di k sedute è più precisa della singola: l'incertezza
        # della differenza è TE*sqrt(1 + 1/k), non TE*sqrt(2).
        margine = Z90 * te_abs * math.sqrt(1.0 + 1.0 / k)
        lo, hi = delta - margine, delta + margine

        if swc is None:
            esito, ampiezza, colore_amp = ESITO_ND, "—", None
            rapporto = None
        else:
            rapporto = abs(delta) / swc if swc else 0.0
            etichetta_amp, colore_amp = scala_ampiezza(rapporto)
            freccia = "🔺" if delta > 0 else ("🔻" if delta < 0 else "▪")
            ampiezza = f"{freccia} {etichetta_amp}"
            if lo > swc or hi < -swc:
                esito = ESITO_REALE
            elif lo > -swc and hi < swc:
                esito = ESITO_STABILE
            else:
                esito = ESITO_INCERTO

        rows.append({
            "Test": cur.loc[mid, "test"],
            "Metrica": etichetta,
            "Attendibilità": esito,
            "Cambiamento": ampiezza,
            "Unità": cur.loc[mid, "unit"],
            "Media rif.": media_rif,
            "Attuale": m_cur,
            "Delta": mostra(delta),
            "Delta %": (m_cur / media_rif - 1) * 100.0,
            "Migliore": migliore,
            # Sempre orientata alla prestazione: 100% = pari al record,
            # oltre 100% = meglio del record. Per le metriche dove meno e'
            # meglio il rapporto va invertito, altrimenti un nuovo primato
            # (valore piu' basso) apparirebbe come un calo.
            "% del migliore": ((migliore / m_cur if lower_better else m_cur / migliore) * 100.0
                               if migliore and m_cur else None),
            "Seduta migliore": sess_migliore,
            "N sedute": k,
            "_mid": mid,
            "_display": f"{cur.loc[mid, 'test']} · {etichetta}",
            "_swc": swc,
            "_margine": margine,
            "_rapporto": rapporto,
            "_colore_amp": colore_amp,
            "_log": log_scale,
            "_fonte_te": fonte_te,
            "_fonte_swc": fonte_swc,
        })

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values(
        "_rapporto", ascending=False, na_position="last")


def build_swc_strip(riga, height=134):
    """Striscia a bande per una metrica, nelle sue unità native. Stessa
    geometria del grafico usato per DSI/EUR: fascia sottile, marcatore del
    valore, etichette sopra. Le bande cadono a ±1/3/6/10 SWC attorno alla
    media di riferimento, così l'allenatore legge insieme il valore reale e
    quanto si è spostato."""
    swc, media_rif = riga["_swc"], riga["Media rif."]
    if not swc or media_rif is None:
        return None

    # In scala log il grafico resta in unità native: le soglie sono
    # moltiplicative, quindi si convertono in valori assoluti.
    def confine(mult):
        if riga["_log"]:
            return media_rif * math.exp(mult * swc)
        return media_rif + mult * swc

    valore = riga["Attuale"]
    margine_abs = (abs(confine(riga["_margine"] / swc) - media_rif)
                   if riga["_log"] else riga["_margine"])

    limiti = [1.0, 3.0, 6.0, 10.0]
    estensione = max(limiti[-1], (riga["_rapporto"] or 0) * 1.25 + 1.5)
    lo_ax, hi_ax = confine(-estensione), confine(estensione)
    if lo_ax > hi_ax:
        lo_ax, hi_ax = hi_ax, lo_ax

    fig = go.Figure()
    bordi = [-estensione] + [-m for m in reversed(limiti)] + limiti + [estensione]
    for a, b in zip(bordi, bordi[1:]):
        centro = (abs(a) + abs(b)) / 2
        _, colore = scala_ampiezza(centro if a * b >= 0 else 0.0)
        x0, x1 = confine(a), confine(b)
        fig.add_vrect(x0=min(x0, x1), x1=max(x0, x1), fillcolor=colore,
                      opacity=0.55, line_width=0)

    for mult, etichetta in ((-6.5, "Grande"), (-4.5, "Moderato"), (-2.0, "Piccolo"),
                            (0.0, "Trascurabile"), (2.0, "Piccolo"), (4.5, "Moderato"),
                            (6.5, "Grande")):
        if abs(mult) > estensione:
            continue
        fig.add_annotation(x=confine(mult), y=1.10, xref="x", yref="paper",
                           yanchor="bottom", text=etichetta, showarrow=False,
                           font=dict(size=9, color="#484343"))

    # La media di riferimento e' un rombo VUOTO, non una linea tratteggiata:
    # su una fascia alta ~20px una vline verticale e' praticamente invisibile.
    # Vuoto vs pieno distingue il riferimento dal dato del test a colpo d'occhio.
    fig.add_trace(go.Scatter(
        x=[media_rif], y=[0], mode="markers", name="Media rif.",
        marker=dict(size=15, color=ACCENT, symbol="diamond-open",
                    line=dict(width=3, color=ACCENT)),
        cliponaxis=False,
        hovertemplate=f"Media riferimento: {fmt_valore(media_rif)}<extra></extra>",
    ))
    fig.add_trace(go.Scatter(
        x=[valore], y=[0], mode="markers+text",
        marker=dict(size=13, color=PRIMARY, symbol="diamond",
                    line=dict(width=1.5, color="white")),
        error_x=dict(type="data", array=[margine_abs], color=PRIMARY,
                     thickness=1.5, width=5),
        text=[fmt_valore(valore)], textposition="bottom center",
        textfont=dict(color=TEXT_COLOR, size=13), cliponaxis=False,
        hovertemplate=(f"Attuale: %{{x:.4g}}<br>Media rif.: {fmt_valore(media_rif)}"
                       f"<extra></extra>"),
    ))

    unita = f" ({riga['Unità']})" if riga["Unità"] else ""
    # Il nome della metrica sta SOPRA il grafico come titolo, non sotto come
    # etichetta dell'asse: e' l'informazione che si cerca per prima quando si
    # scorre una sequenza di strisce. Il margine superiore va alzato insieme
    # all'altezza, altrimenti il titolo si sovrappone alle etichette delle
    # fasce (annotazioni in coordinate paper appena sopra l'area di plot).
    fig.update_layout(
        title=dict(text=f"<b>{riga['Metrica']}{unita}</b>", x=0.5, xanchor="center",
                   y=0.90, yanchor="top",
                   font=dict(size=15, color=TEXT_COLOR)),
        xaxis=dict(range=[lo_ax, hi_ax], showgrid=False, automargin=True,
                   tickvals=[media_rif], ticktext=[f"media {fmt_valore(media_rif)}"]),
        yaxis=dict(range=[-1, 1], showgrid=False, showticklabels=False, zeroline=False),
        height=height, margin=dict(t=64, b=50, l=20, r=20), showlegend=False,
        plot_bgcolor=BG_COLOR, paper_bgcolor=BG_COLOR, font=dict(color=TEXT_COLOR),
    )
    return fig


# ----------------------------------------------------------------------------
# Confronto dei profili di forza (T-score) e degli indici DSI / EUR
# ----------------------------------------------------------------------------
# I T-score sono già normalizzati e direzionali (per le metriche dove meno è
# meglio lo z viene invertito in z_t_score), quindi un T più alto è sempre
# migliore: qui, a differenza delle metriche grezze, "migliore" ha senso.
#
# La scala di ampiezza si traduce senza inventare nulla: T = 50 + 10*z e
# SWC = 0,2 SD, quindi 1 SWC = 2 punti di T-score. Piccolo = 2 punti,
# moderato = 6, grande = 12, molto grande = 20.
SWC_IN_PUNTI_T = 10.0 * SWC_FACTOR


def colore_delta_t(delta):
    """Verde se il T-score sale, rosso se scende. Qui il colore e' legittimo
    (a differenza delle metriche grezze): il T-score e' gia' orientato alla
    prestazione, quindi piu' alto e' sempre migliore."""
    if delta is None or (isinstance(delta, float) and math.isnan(delta)):
        return "#8d8d8d"
    return "#2E7D32" if delta > 0 else ("#B00020" if delta < 0 else "#8d8d8d")


def comp_score_specs(pop_dict, sesso_atleta):
    """{metric_id: (categoria, media_pop, sd_pop, lower_is_better)} per le
    sole metriche curate che hanno un T-score."""
    is_female = bool(sesso_atleta) and str(sesso_atleta).upper().startswith("F")
    specs = {}
    for m in METRICS:
        if m.get("kind") != "score" or not m.get("pop_key") or not m.get("jump_type"):
            continue
        pop = pop_dict[m["pop_key"]]
        media = pop["mean_f"] if is_female else pop["mean_m"]
        sd = pop["sd_f"] if is_female else pop["sd_m"]
        if not sd:
            continue
        mid = _metric_id(m["jump_type"], m.get("raw_var") or m["key"])
        specs[mid] = (m["category"], float(media), float(sd), bool(m["lower_is_better"]))
    return specs


def profili_per_sessione(long_df, specs):
    """(profili, t_per_metrica): profili è {sessione: {categoria: T medio}},
    t_per_metrica è {sessione: {metric_id: T}}. Stessa media semplice per
    categoria usata da profilo_forza() nella PARTE 4."""
    profili, t_metrica = {}, {}
    for sessione, g in long_df.groupby("session", sort=False):
        per_cat, per_mid = {}, {}
        for _, riga in g.iterrows():
            spec = specs.get(riga["metric_id"])
            if not spec or riga["mean"] is None:
                continue
            categoria, media_pop, sd_pop, lower = spec
            _z, t = z_t_score(riga["mean"], media_pop, sd_pop, lower)
            if t is None:
                continue
            per_mid[riga["metric_id"]] = t
            per_cat.setdefault(categoria, []).append(t)
        if per_cat:
            profili[sessione] = {c: sum(v) / len(v) for c, v in per_cat.items()}
            t_metrica[sessione] = per_mid
    return profili, t_metrica


def confronta_profili(prof_cur, prof_hist, t_hist, hist_long, specs, athlete_te):
    """Confronta il profilo attuale con la media dei profili precedenti.

    L'incertezza in punti T viene ricavata dal TE delle metriche che compongono
    la categoria: TE_T = (TE assoluto / SD popolazione) * 10. Non viene divisa
    per la radice del numero di metriche, perché le metriche di una stessa
    categoria sono correlate fra loro: la stima resta prudente."""
    if not prof_cur or not prof_hist:
        return pd.DataFrame()

    medie_rif = {}
    for _sess, per_cat in prof_hist.items():
        for categoria, t in per_cat.items():
            medie_rif.setdefault(categoria, []).append(t)

    # TE in punti T per categoria, dalle metriche che la compongono
    medie_metrica = hist_long.groupby("metric_id")["mean"].mean().to_dict()
    te_per_cat = {}
    for mid, (categoria, _media_pop, sd_pop, _lower) in specs.items():
        media = medie_metrica.get(mid)
        if media is None or not sd_pop:
            continue
        cv = athlete_te[mid][0] if mid in athlete_te else te_cv_default(mid)
        te_punti = (cv / 100.0 * abs(media)) / sd_pop * 10.0
        te_per_cat.setdefault(categoria, []).append(te_punti)

    rows = []
    for categoria in CATEGORIES:
        if categoria not in prof_cur or categoria not in medie_rif:
            continue
        storici = medie_rif[categoria]
        k = len(storici)
        media_rif = sum(storici) / k
        t_cur = prof_cur[categoria]
        delta = t_cur - media_rif

        te_punti = te_per_cat.get(categoria)
        if te_punti:
            margine = Z90 * (sum(te_punti) / len(te_punti)) * math.sqrt(1.0 + 1.0 / k)
        else:
            margine = None

        if margine is None:
            esito, ampiezza, colore = ESITO_ND, "—", None
            rapporto = None
        else:
            rapporto = abs(delta) / SWC_IN_PUNTI_T
            etichetta, colore = scala_ampiezza(rapporto)
            freccia = "🔺" if delta > 0 else ("🔻" if delta < 0 else "▪")
            ampiezza = f"{freccia} {etichetta}"
            lo, hi = delta - margine, delta + margine
            if lo > SWC_IN_PUNTI_T or hi < -SWC_IN_PUNTI_T:
                esito = ESITO_REALE
            elif lo > -SWC_IN_PUNTI_T and hi < SWC_IN_PUNTI_T:
                esito = ESITO_STABILE
            else:
                esito = ESITO_INCERTO

        migliore = max(storici)
        sess_migliore = next((s for s, p in prof_hist.items()
                              if p.get(categoria) == migliore), "—")
        banda_cur, _colore_banda = banda_da_tscore(t_cur)

        rows.append({
            "Categoria": categoria,
            "Attendibilità": esito,
            "Cambiamento": ampiezza,
            "T medio rif.": media_rif,
            "T attuale": t_cur,
            "Delta T": delta,
            "Valutazione attuale": banda_cur or "—",
            "T migliore": migliore,
            "Seduta migliore": sess_migliore,
            "N sedute": k,
            "_margine": margine,
            "_rapporto": rapporto,
            "_colore": colore,
        })
    return pd.DataFrame(rows)


def build_profili_radar(cats, serie, nome_atleta):
    """Radar con più serie sovrapposte (attuale, media storico, migliore),
    sulla stessa geometria di build_radar_chart della PARTE 4."""
    if not cats:
        return None
    cats_closed = [_wrap_label(c) for c in cats] + [_wrap_label(cats[0])]
    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=[50] * (len(cats) + 1), theta=cats_closed, mode="lines",
        line=dict(color="rgba(233,74,38,0.6)", dash="dash"),
        name="Media popolazione (T=50)",
    ))
    stili = [
        ("Media sedute precedenti", "#8d8d8d", "dot", None, 2),
        ("Miglior seduta", "#2E7D32", "dash", None, 2),
        (f"{nome_atleta} — attuale", PRIMARY, "solid", "rgba(11,182,255,0.25)", 3),
    ]
    for (etichetta, colore, tratto, riempimento, spessore) in stili:
        valori = serie.get(etichetta)
        if not valori:
            continue
        vals = [valori.get(c) for c in cats]
        if any(v is None for v in vals):
            continue
        fig.add_trace(go.Scatterpolar(
            r=vals + [vals[0]], theta=cats_closed, name=etichetta,
            fill="toself" if riempimento else None, fillcolor=riempimento,
            line=dict(color=colore, width=spessore, dash=tratto),
        ))
    fig.update_layout(
        polar=dict(
            domain=dict(x=[0.12, 0.88], y=[0.08, 0.92]),
            radialaxis=dict(range=[0, 100], showticklabels=True, ticks="",
                            tickfont=dict(color=TEXT_COLOR)),
        ),
        font=dict(color=TEXT_COLOR), paper_bgcolor=BG_COLOR,
        showlegend=True, height=560, margin=dict(t=60, b=60, l=60, r=60),
        legend=dict(orientation="h", yanchor="bottom", y=-0.12, x=0),
    )
    return fig


def build_indice_zone_strip(key, valore_attuale, valore_rif, thr_low, thr_high,
                            decimals=3, height=140):
    """Striscia a zone di profilo per DSI/EUR (stesse bande e stesse etichette
    di ratio_band_strip della PARTE 4), con due marcatori: media delle sedute
    precedenti e test attuale. Serve a vedere se l'atleta ha cambiato ZONA,
    non solo se il numero si è mosso."""
    if valore_attuale is None or not thr_low or not thr_high:
        return None
    valori = [v for v in (valore_attuale, valore_rif) if v is not None]
    lo = min([thr_low - 0.2] + [v - 0.05 for v in valori])
    hi = max([thr_high + 0.2] + [v + 0.05 for v in valori])
    labels = INDEX_ZONE_LABELS[key]

    fig = go.Figure()
    for a, b, i in ((lo, thr_low, 0), (thr_low, thr_high, 1), (thr_high, hi, 2)):
        fig.add_vrect(x0=a, x1=b, fillcolor=INDEX_ZONE_COLORS[i], opacity=0.35, line_width=0)
        fig.add_annotation(x=(a + b) / 2, y=1.10, xref="x", yref="paper", yanchor="bottom",
                           text=_wrap_label(labels[i], 22), showarrow=False,
                           font=dict(size=11, color="#484343"))
    if valore_rif is not None:
        fig.add_trace(go.Scatter(
            x=[valore_rif], y=[0], mode="markers", name="Media precedenti",
            marker=dict(size=12, color="#8d8d8d", symbol="circle-open",
                        line=dict(width=2.5, color="#8d8d8d")),
            cliponaxis=False,
            hovertemplate=f"Media precedenti: %{{x:.{decimals}f}}<extra></extra>",
        ))
    fig.add_trace(go.Scatter(
        x=[valore_attuale], y=[0], mode="markers+text", name="Attuale",
        marker=dict(size=13, color=PRIMARY, symbol="diamond",
                    line=dict(width=1.5, color="white")),
        text=[f"{valore_attuale:.{decimals}f}"], textposition="bottom center",
        textfont=dict(color=TEXT_COLOR, size=13), cliponaxis=False,
        hovertemplate=f"Attuale: %{{x:.{decimals}f}}<extra></extra>",
    ))
    fig.update_layout(
        xaxis=dict(range=[lo, hi], showgrid=False, tickvals=[thr_low, thr_high],
                   ticktext=[f"{thr_low:.2f}", f"{thr_high:.2f}"], automargin=True,
                   title=dict(text=key.upper(), standoff=12)),
        yaxis=dict(range=[-1, 1], showgrid=False, showticklabels=False, zeroline=False),
        height=height, margin=dict(t=64, b=50, l=20, r=20),
        showlegend=True, legend=dict(orientation="h", yanchor="bottom", y=-0.55, x=0),
        plot_bgcolor=BG_COLOR, paper_bgcolor=BG_COLOR, font=dict(color=TEXT_COLOR),
    )
    return fig


# ----------------------------------------------------------------------------
# UI del tab
# ----------------------------------------------------------------------------
with tab_comparazione:
    if not parsed_files:
        st.info("Carica dei file dalla barra laterale per confrontare questo test con i precedenti.")
    else:
        st.markdown(
            "Confronta il test attuale con lo storico dell'atleta. Carica i CSV prodotti da "
            "**🔍 Dettaglio Test → Esporta TUTTE le metriche grezze**."
        )
        st.caption(
            "Il confronto statistico è contro la MEDIA delle sedute precedenti. È riportato anche il miglior "
            "valore storico come riferimento."
        )

        comp_files = st.file_uploader(
            "CSV delle sedute precedenti (export completo)", type=["csv"],
            accept_multiple_files=True, key="comp_files",
            help="Uno o più file 'Scarica export completo (.csv)' dello STESSO atleta. "
                 "Da 4 sedute in su l'app stima il rumore della misura sull'atleta stesso.",
        )

        if not comp_files:
            st.info("Carica almeno un CSV di riferimento per iniziare il confronto.")
        else:
            storico, errori = [], []
            for i, f in enumerate(comp_files):
                try:
                    storico.append(parse_full_export_csv(
                        io.BytesIO(f.getvalue()), f.name.rsplit(".", 1)[0], ordine=i))
                except Exception as e:
                    errori.append(f"**{f.name}**: {e}")
            for msg in errori:
                st.warning(msg)

            if storico:
                storico = ordina_sessioni(storico)
                hist_long = pd.concat(storico, ignore_index=True)
                cur_long = reps_to_long(current_reps_by_type(parsed_files), "Attuale",
                                        _parse_data_test(periodo), ordine=len(storico))
                pop_sd_map = comp_pop_sd_map(st.session_state["pop"], sesso)

                senza_data = [f["session"].iloc[0] for f in storico if f["data"].iloc[0] is None]
                if senza_data:
                    st.caption(
                        "Sedute senza data leggibile, ordinate come caricate: "
                        + ", ".join(senza_data)
                    )

                all_long = pd.concat([hist_long, cur_long], ignore_index=True)
                te_atleta = athlete_te_table(all_long)
                n_sedute = len(storico)

                if te_atleta:
                    st.success(
                        f"✅ {n_sedute} sedute di riferimento: rumore della misura stimato "
                        f"sull'atleta per {len(te_atleta)} metriche. Le altre usano valori "
                        "di letteratura."
                    )
                else:
                    st.info(
                        f"ℹ️ {n_sedute} sedute di riferimento. Da {MIN_SESSIONS_TE_ATLETA} in su "
                        "l'app stima il rumore sull'atleta stesso e ricava una soglia di "
                        "rilevanza anche per le metriche prive di norme; per ora usa valori "
                        "di letteratura, quindi gli intervalli sono più larghi del necessario."
                    )

                res = compare_to_history(cur_long, hist_long, te_atleta, pop_sd_map)

                if res.empty:
                    st.error(
                        "Nessuna metrica in comune tra il test attuale e lo storico. Verifica "
                        "che i CSV appartengano allo stesso atleta e agli stessi tipi di test."
                    )
                else:
                    st.caption(f"Metriche confrontabili trovate: {len(res)}.")
                    primarie_ids = {
                        _metric_id(m["jump_type"], m.get("raw_var") or m["key"])
                        for m in METRICS
                        if m["key"] in COMP_DEFAULT_PRIMARY_KEYS and m.get("jump_type")
                    }
                    opzioni = list(res["_display"])
                    default_sel = [r["_display"] for _, r in res.iterrows()
                                   if r["_mid"] in primarie_ids]
                    scelte = st.multiselect(
                        "Metriche da mostrare", opzioni,
                        default=default_sel or opzioni[: min(5, len(opzioni))],
                        key="comp_scelte",
                        help="Sceglile PRIMA di guardare i dati: è ciò che distingue una "
                             "decisione da un inseguimento del rumore.",
                    )
                    vista = res[res["_display"].isin(scelte)] if scelte else res

                    conteggi = vista["Attendibilità"].value_counts()
                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric(ESITO_REALE, int(conteggi.get(ESITO_REALE, 0)))
                    k2.metric(ESITO_STABILE, int(conteggi.get(ESITO_STABILE, 0)))
                    k3.metric(ESITO_INCERTO, int(conteggi.get(ESITO_INCERTO, 0)))
                    k4.metric(ESITO_ND, int(conteggi.get(ESITO_ND, 0)))

                    tab_df = vista[[
                        "Test", "Metrica", "Unità", "Media rif.", "Attuale",
                        "Delta", "Delta %", "Cambiamento", "Attendibilità",
                        "Migliore", "% del migliore", "Seduta migliore", "N sedute",
                    ]].reset_index(drop=True)

                    def _stile_comp(row):
                        stili = [""] * len(row)
                        col = ESITO_COLORI.get(row["Attendibilità"])
                        if col:
                            stili[tab_df.columns.get_loc("Attendibilità")] = (
                                f"color:{col}; font-weight:600;")
                        return stili

                    st.dataframe(
                        tab_df.style.apply(_stile_comp, axis=1).format({
                            "Media rif.": fmt_valore, "Attuale": fmt_valore,
                            "Delta": lambda v: fmt_valore(v, segno=True),
                            "Delta %": "{:+.1f}%", "Migliore": fmt_valore,
                            "% del migliore": "{:.0f}%",
                        }, na_rep="—"),
                        use_container_width=True, hide_index=True,
                    )

                    nd = [r["_display"] for _, r in vista.iterrows()
                          if r["Attendibilità"] == ESITO_ND]
                    if nd:
                        st.warning(
                            f"⚠️ {len(nd)} metriche non valutabili: manca sia una norma di "
                            f"popolazione sia lo storico necessario ({MIN_SESSIONS_TE_ATLETA} "
                            "sedute) per ricavare una soglia di rilevanza dall'atleta. Il "
                            "valore si vede comunque, ma non si può dire se lo scostamento conti."
                        )

                    # --- Strisce per metrica ---
                    st.markdown("#### Scostamento dalla media, metrica per metrica")
                    st.caption(
                        "Il rombo pieno azzurro è il test attuale con la sua barra di "
                        "incertezza; il rombo vuoto arancione è la media delle sedute "
                        "precedenti."
                    )
                    strip_opzioni = [r["_display"] for _, r in vista.iterrows()
                                     if r["_swc"] is not None]
                    strip_scelte = st.multiselect(
                        "Metriche da graficare", strip_opzioni,
                        default=strip_opzioni[: min(5, len(strip_opzioni))],
                        key="comp_strip",
                        help="Le metriche non valutabili (⚠) non compaiono: senza soglia di "
                             "rilevanza non c'è scala da disegnare.",
                    )
                    if not strip_opzioni:
                        st.info("Nessuna metrica con una soglia di rilevanza disponibile.")
                    for _, riga in vista[vista["_display"].isin(strip_scelte)].iterrows():
                        fig_strip = build_swc_strip(riga)
                        if fig_strip:
                            st.plotly_chart(fig_strip, use_container_width=True)

                    # --- Profili di forza (T-score) e indici -----------------
                    st.markdown("---")
                    st.markdown("#### Profilo di forza: attuale vs storico")
                    specs = comp_score_specs(st.session_state["pop"], sesso)
                    prof_hist, _t_hist = profili_per_sessione(hist_long, specs)
                    prof_cur_all, _t_cur = profili_per_sessione(cur_long, specs)
                    prof_cur = prof_cur_all.get("Attuale", {})

                    prof_res, prof_serie, prof_cats = pd.DataFrame(), None, []
                    if not prof_cur or not prof_hist:
                        st.info(
                            "Profilo non calcolabile: servono metriche con norme di popolazione "
                            "(IMTP, SJ, CMJ o CMJ RE) sia nel test attuale sia nello storico."
                        )
                    else:
                        cats_valide = [c for c in CATEGORIES
                                       if c in prof_cur and any(c in p for p in prof_hist.values())]
                        medie_cat, migliori_cat = {}, {}
                        for c in cats_valide:
                            valori = [p[c] for p in prof_hist.values() if c in p]
                            medie_cat[c] = sum(valori) / len(valori)
                            migliori_cat[c] = max(valori)

                        prof_cats = cats_valide
                        prof_serie = {
                            "Media sedute precedenti": medie_cat,
                            "Miglior seduta": migliori_cat,
                            f"{nome} — attuale": prof_cur,
                        }
                        # confronta_profili() viene chiamata PRIMA del radar
                        # perche' le card sotto il grafico usano il Delta T.
                        prof_res = confronta_profili(prof_cur, prof_hist, _t_hist, hist_long,
                                                     specs, te_atleta)
                        delta_per_cat = ({r["Categoria"]: r["Delta T"]
                                          for _, r in prof_res.iterrows()}
                                         if not prof_res.empty else {})

                        radar = build_profili_radar(cats_valide, prof_serie, nome)
                        if radar:
                            st.plotly_chart(radar, use_container_width=True)

                        # Card come nella scheda Profilo di Forza, piu' il solo
                        # scostamento in punti T: entita' e attendibilita' si
                        # leggono in tabella, qui servirebbero solo a togliere
                        # risalto al T-score.
                        cols_prof = st.columns(len(cats_valide))
                        for col, cat in zip(cols_prof, cats_valide):
                            t = prof_cur[cat]
                            banda, colore = banda_da_tscore(t)
                            d = delta_per_cat.get(cat)
                            html_card = (
                                f"<div style='font-size:11px;letter-spacing:0.5px;"
                                f"opacity:0.65'>{cat}</div>"
                                f"<div style='font-size:17px;font-weight:600;"
                                f"color:{TEXT_COLOR};margin:2px 0 4px 0'>"
                                f"{CATEGORY_QUALITY.get(cat, '')}</div>"
                                f"<div style='font-size:30px;font-weight:700;"
                                f"color:{colore}'>{t:.0f}</div>"
                            )
                            if d is not None:
                                html_card += (
                                    f"<div style='font-size:15px;font-weight:600;"
                                    f"color:{colore_delta_t(d)}'>{d:+.1f}</div>"
                                )
                            html_card += (
                                f"<div style='font-size:13px;opacity:0.75;"
                                f"margin-top:4px'>{banda}</div>"
                            )
                            with col.container(border=True):
                                st.markdown(html_card, unsafe_allow_html=True)

                        if not prof_res.empty:
                            prof_df = prof_res[[
                                "Categoria", "T medio rif.", "T attuale", "Delta T",
                                "Cambiamento", "Attendibilità", "Valutazione attuale",
                                "T migliore", "Seduta migliore", "N sedute",
                            ]].reset_index(drop=True)

                            def _stile_prof(row):
                                stili = [""] * len(row)
                                col = ESITO_COLORI.get(row["Attendibilità"])
                                if col:
                                    stili[prof_df.columns.get_loc("Attendibilità")] = (
                                        f"color:{col}; font-weight:600;")
                                return stili

                            st.dataframe(
                                prof_df.style.apply(_stile_prof, axis=1).format({
                                    "T medio rif.": "{:.1f}", "T attuale": "{:.1f}",
                                    "Delta T": "{:+.1f}", "T migliore": "{:.1f}",
                                }, na_rep="—"),
                                use_container_width=True, hide_index=True,
                            )

                    # --- Indici di profilo DSI / EUR ------------------------
                    idx_export = []
                    idx_disponibili = [
                        k for k in ("dsi", "eur")
                        if _metric_id("indici", k) in set(cur_long["metric_id"])
                    ]
                    if idx_disponibili:
                        st.markdown("#### Indici di profilo")
                        st.caption(
                            "Il rombo è il test attuale, il cerchio vuoto la media delle sedute precedenti."
                        )
                        for key in idx_disponibili:
                            mid = _metric_id("indici", key)
                            att = cur_long[cur_long["metric_id"] == mid]["mean"]
                            rif = hist_long[hist_long["metric_id"] == mid]["mean"]
                            v_att = float(att.iloc[0]) if len(att) else None
                            v_rif = float(rif.mean()) if len(rif) else None
                            lo_thr, hi_thr = st.session_state["idx_thr"][key]
                            fig_idx = build_indice_zone_strip(key, v_att, v_rif, lo_thr, hi_thr)
                            if fig_idx is None:
                                continue
                            zona_att, _c1 = zona_da_indice(key, v_att, lo_thr, hi_thr)
                            zona_rif, _c2 = zona_da_indice(key, v_rif, lo_thr, hi_thr)
                            if zona_att and zona_rif and zona_att != zona_rif:
                                st.markdown(
                                    f"**{key.upper()}** — cambio di zona: "
                                    f"da *{zona_rif}* a *{zona_att}*."
                                )
                            elif zona_att:
                                st.markdown(f"**{key.upper()}** — zona invariata: *{zona_att}*.")
                            st.plotly_chart(fig_idx, use_container_width=True)
                            idx_export.append(dict(
                                key=key, attuale=v_att, riferimento=v_rif,
                                zona_att=zona_att, zona_rif=zona_rif,
                                thr=(lo_thr, hi_thr),
                            ))

                    # Rende i risultati disponibili alla scheda 📄 Report: i
                    # generatori HTML/PDF leggono questa chiave se presente, e
                    # includono la sezione Comparazione. Le figure NON vengono
                    # salvate qui (sono ricostruite al momento dell'export dalle
                    # stesse funzioni), solo i dati.
                    st.session_state["comp_export"] = dict(
                        vista=vista.copy(), n_sedute=n_sedute,
                        prof=prof_res.copy() if not prof_res.empty else None,
                        prof_serie=prof_serie, prof_cats=list(prof_cats),
                        strip=[d for d in strip_scelte],
                        indici=idx_export,
                    )

                    st.download_button(
                        "⬇️ Scarica comparazione (.csv)",
                        data=res.drop(columns=[c for c in res.columns if c.startswith("_")])
                                .to_csv(index=False).encode("utf-8-sig"),
                        file_name=f"comparazione_{nome.replace(' ', '_')}_"
                                  f"{dt.date.today().isoformat()}.csv",
                        mime="text/csv",
                    )

                    with st.expander("ℹ️ Come si leggono questi risultati"):
                        st.markdown(
                            f"""
**Attendibilità — il cambiamento è reale?**

- **{ESITO_REALE}** — l'intervallo di incertezza sta interamente oltre la soglia
  di rilevanza: il cambiamento supera il rumore della misura ed è abbastanza
  grande da contare.
- **{ESITO_STABILE}** — l'intervallo sta interamente dentro la soglia: quasi
  certamente NON c'è stato un cambiamento rilevante. È una conclusione, non un
  dato mancante: sapere che l'atleta è fermo è un'informazione.
- **{ESITO_INCERTO}** — l'intervallo scavalca la soglia: i dati non bastano per
  decidere. Con poche sedute è l'esito più frequente, ed è la risposta onesta.
- **{ESITO_ND}** — manca una soglia di rilevanza: né norma di popolazione né
  {MIN_SESSIONS_TE_ATLETA} sedute da cui ricavarla. Il valore resta leggibile, il
  giudizio no.

**Cambiamento — quanto si è spostato?** Scala di Hopkins in multipli della soglia:
trascurabile (sotto 1), piccolo (1-3), moderato (3-6), grande (6-10), molto grande
(oltre 10). La freccia dice solo la direzione: 🔺 il numero è salito, 🔻 è sceso.
**L'app non dice se sia un bene**: per un tempo di contatto scendere è un
progresso, per un'altezza di salto è un calo. Quel giudizio spetta a te.

**Media rif.** — media delle sedute precedenti, esclusa l'attuale. Escluderla è
necessario: altrimenti confronteresti il test con qualcosa che contiene sé
stesso, e lo scostamento risulterebbe artificialmente più piccolo.

**Migliore** — il miglior valore mai registrato, con la seduta in cui è avvenuto.
È **descrittivo**: non ha né attendibilità né entità. Il massimo di più misure rumorose
è per costruzione una sovrastima, quindi un confronto statistico con il record
segnalerebbe un peggioramento quasi sempre, per un difetto del metodo e non
dell'atleta. Usalo come riferimento pratico ("è al 92% del suo record").

**La soglia di rilevanza (SWC)** vale 0,2 deviazioni standard e viene presa, in
ordine: dalle norme di popolazione della scheda ⚙️ Costanti; se mancano, dalla
variabilità storica dell'atleta stesso (da {MIN_SESSIONS_TE_ATLETA} sedute in su)
— che per il monitoraggio individuale è anche più pertinente della popolazione;
se mancano entrambe, l'esito è {ESITO_ND}.

**Indici di simmetria** — confrontati in valore assoluto: il segno indica il lato
dominante, non la qualità, quindi passare da -2% a +6% è un aumento dello
sbilanciamento anche se il numero "cresce".

**DSI ed EUR** — sono rapporti fra test diversi: variazione e soglia sono in
percentuale, non in unità assolute.
                            """
                        )
# ============================================================================
# PARTE 6 — REPORT SCARICABILE (HTML interattivo)
# ============================================================================
# Il report è un unico file HTML autosufficiente: i grafici Plotly restano
# interattivi (hover, zoom) invece di essere "appiattiti" in immagini
# statiche come richiederebbe un PDF. plotly.js viene caricato una sola
# volta via CDN nell'<head>; ogni grafico è poi un semplice <div> generato
# dalle stesse funzioni build_*_chart usate nella UI live (nessuna
# duplicazione della logica di analisi/plotting). Chi riceve il file può
# aprirlo in qualunque browser e, se serve una copia statica, stampare /
# "Salva come PDF" direttamente da lì.
#
# ORDINE DEL FILE (non modificarlo a caso):
#   PARTE 6    — helper HTML condivisi
#   PARTE 6ter — sezione Comparazione (HTML + PDF)
#   genera_report_html
#   PARTE 6bis — infrastruttura PDF (_pdf_safe, _ReportPDF, ...)
#   genera_report_pdf
#   with tab_report
# La 6ter usa _pdf_safe/_hex_to_rgb/FontFace, definiti più in basso nella
# 6bis: non è un problema, i nomi vengono risolti alla CHIAMATA e non alla
# definizione della funzione.

def _fig_div(fig, div_id):
    """Converte una figura Plotly in un <div> HTML incorporabile nel report,
    senza reincludere plotly.js (già caricato una volta nell'head)."""
    if fig is None:
        return ""
    return fig.to_html(full_html=False, include_plotlyjs=False, div_id=div_id,
                       config={"displaylogo": False, "responsive": True})


def _banda_badge(banda, colore):
    if not banda:
        return "—"
    return f'<span class="badge" style="background:{colore}">{banda}</span>'


def _index_value_html(r):
    """Valore numerico di un indice (DSI/EUR), mostrato accanto al proprio
    grafico invece che in una tabella riepilogativa separata. Include la zona
    di profilo, la media di popolazione e le soglie usate, così il report
    resta leggibile anche a distanza di mesi."""
    if r is None:
        return ""
    extra = []
    if r.get("zona"):
        extra.append(r["zona"])
    if r["pop_mean"] is not None:
        extra.append(f"Media pop. {r['pop_mean']:.3f}")
    if r.get("thr_low") is not None:
        extra.append(f"soglie {r['thr_low']:.2f} / {r['thr_high']:.2f}")
    extra_html = f' <span class="muted">— {" · ".join(extra)}</span>' if extra else ""
    return f'<p class="index-value"><b>{r["label"]}: {r["mean"]:.3f}</b>{extra_html}</p>'


def _metric_descriptions_html(items):
    """Blocco a tendina (nativo, <details>/<summary>, nessun JS) con le
    descrizioni delle metriche — stesso pattern dell'expander "ℹ️ Cosa
    significano queste metriche?" della scheda Dettaglio Test. Ritorna
    stringa vuota se nessuna delle metriche passate ha una descrizione."""
    pairs = [(it["label"], it["desc"]) for it in items if it and it.get("desc")]
    if not pairs:
        return ""
    li_html = "".join(
        f"<li><b>{lbl}</b> — {_html.escape(desc)}</li>" for lbl, desc in pairs
    )
    return f"""<details class="metric-help">
        <summary>ℹ️ Cosa significano queste metriche?</summary>
        <ul>{li_html}</ul>
    </details>"""


def _metric_table_html(cat_results):
    rows_html = []
    for r in cat_results:
        media = f"{r['mean']:.3f}" if r["mean"] is not None else "N/D"
        media_pop = f"{r['pop_mean']:.3f}" if r["pop_mean"] is not None else "—"
        tscore = f"{r['t']:.0f}" if r["t"] is not None else "—"
        rows_html.append(f"""<tr>
            <td>{r['label']}</td><td>{r['unit'] or ''}</td><td>{media}</td><td>{media_pop}</td>
            <td>{r['n']}</td><td>{tscore}</td><td>{_banda_badge(r['banda'], r['colore'])}</td>
        </tr>""")
    return f"""<table class="report-table">
        <thead><tr><th>Metrica</th><th>Unità</th><th>Media</th><th>Media Pop.</th><th>N</th><th>T-score</th><th>Valutazione</th></tr></thead>
        <tbody>{''.join(rows_html)}</tbody>
    </table>"""


# ============================================================================
# PARTE 6ter — SEZIONE COMPARAZIONE NEI REPORT
# ============================================================================
# I risultati della scheda 🔀 Comparazione vengono salvati in
# st.session_state["comp_export"] dalla PARTE 5bis. Qui vengono trasformati
# in HTML e in PDF. Se la chiave non esiste (comparazione non eseguita in
# questa sessione) le funzioni restituiscono contenuto vuoto e il report
# resta identico a prima: la sezione è opzionale, non obbligatoria.
#
# Le FIGURE non vengono salvate nello stato ma ricostruite qui dalle stesse
# funzioni build_swc_strip / build_profili_radar / build_indice_zone_strip
# usate nella UI live: un solo posto dove cambiare la grafica.

# Emoji -> testo per il PDF: i font core di fpdf2 sono Latin-1, quindi ogni
# emoji verrebbe scartata da _pdf_safe lasciando l'etichetta mutila
# ("🟢 Alta" -> " Alta"). Le frecce diventano segni ASCII, i pallini del
# semaforo sparisc­ono perché nel PDF il colore lo dà il riempimento cella.
_PDF_COMP_REPLACEMENTS = {
    "\U0001F7E2": "", "\U0001F7E1": "", "\U0001F534": "",
    "\u26A0\uFE0F": "", "\u26A0": "", "\u2753": "",
    "\U0001F53A": "+", "\U0001F53B": "-", "\u25AA": "=",
}


def _pdf_comp(text):
    """_pdf_safe + traduzione delle emoji di Comparazione."""
    t = str(text if text is not None else "")
    for old, new in _PDF_COMP_REPLACEMENTS.items():
        t = t.replace(old, new)
    return _pdf_safe(t.strip())


def _delta_t_html(delta):
    """Riga compatta col solo scostamento in punti T. Entita' e attendibilita'
    restano in tabella: nella card toglierebbero risalto al T-score."""
    if delta is None or (isinstance(delta, float) and math.isnan(delta)):
        return ""
    return (f'<div class="profile-card-delta" style="color:{colore_delta_t(delta)}">'
            f'{delta:+.1f}</div>')


def _comp_esito_colore(esito):
    """Colore associato a un'etichetta di attendibilità, con fallback grigio."""
    return ESITO_COLORI.get(str(esito), "#8d8d8d")


def _fmt(valore, decimali=3, segno=False, suffisso=""):
    """Wrapper su fmt_valore della PARTE 5bis: stessa resa numerica in UI e
    nei report. `decimali` e' rispettato solo per le percentuali, dove serve
    un numero fisso di cifre."""
    if valore is None or (isinstance(valore, float) and math.isnan(valore)):
        return "—"
    if suffisso == "%" and isinstance(valore, (int, float)):
        return (f"{valore:+.{decimali}f}" if segno else f"{valore:.{decimali}f}") + suffisso
    return fmt_valore(valore, segno=segno) + suffisso


# ----------------------------------------------------------------------------
# HTML
# ----------------------------------------------------------------------------
COMP_HTML_COLS = [
    ("Test", 0, False), ("Metrica", 0, False), ("Unità", 0, False),
    ("Media rif.", 3, False), ("Attuale", 3, False),
    ("Delta", 3, True), ("Delta %", 1, True),
    ("Cambiamento", 0, False), ("Attendibilità", 0, False),
    ("Migliore", 3, False), ("% del migliore", 0, False),
    ("Seduta migliore", 0, False),
]


def _comp_table_html(vista):
    """Tabella della Comparazione: stesse colonne e stesso ordine della UI."""
    if vista is None or vista.empty:
        return ""
    thead = "".join(f"<th>{c}</th>" for c, _d, _s in COMP_HTML_COLS)
    righe = []
    for _, r in vista.iterrows():
        celle = []
        for col, dec, segno in COMP_HTML_COLS:
            v = r.get(col)
            if col == "Attendibilità":
                celle.append(f'<td style="color:{_comp_esito_colore(v)};'
                             f'font-weight:600">{v}</td>')
            elif col == "Delta %":
                celle.append(f"<td>{_fmt(v, 1, True, '%')}</td>")
            elif col == "% del migliore":
                celle.append(f"<td>{_fmt(v, 0, False, '%')}</td>")
            elif dec:
                celle.append(f"<td>{_fmt(v, dec, segno)}</td>")
            else:
                celle.append(f"<td>{v if v not in (None, '') else '—'}</td>")
        righe.append("<tr>" + "".join(celle) + "</tr>")
    return (f'<table class="report-table"><thead><tr>{thead}</tr></thead>'
            f'<tbody>{"".join(righe)}</tbody></table>')


PROF_HTML_COLS = [
    ("Categoria", 0), ("T medio rif.", 1), ("T attuale", 1), ("Delta T", 1),
    ("Cambiamento", 0), ("Attendibilità", 0), ("Valutazione attuale", 0),
    ("T migliore", 1), ("Seduta migliore", 0), ("N sedute", 0),
]


def _prof_table_html(prof):
    if prof is None or prof.empty:
        return ""
    thead = "".join(f"<th>{c}</th>" for c, _d in PROF_HTML_COLS)
    righe = []
    for _, r in prof.iterrows():
        celle = []
        for col, dec in PROF_HTML_COLS:
            v = r.get(col)
            if col == "Attendibilità":
                celle.append(f'<td style="color:{_comp_esito_colore(v)};'
                             f'font-weight:600">{v}</td>')
            elif col == "Delta T":
                celle.append(f"<td>{v:+.1f}</td>" if isinstance(v, (int, float)) else "<td>—</td>")
            elif dec:
                celle.append(f"<td>{v:.{dec}f}</td>" if isinstance(v, (int, float)) else "<td>—</td>")
            else:
                celle.append(f"<td>{v if v not in (None, '') else '—'}</td>")
        righe.append("<tr>" + "".join(celle) + "</tr>")
    return (f'<table class="report-table"><thead><tr>{thead}</tr></thead>'
            f'<tbody>{"".join(righe)}</tbody></table>')


def comparazione_sections_html(comp, nome_atleta, next_id):
    """Sezioni HTML della Comparazione. Lista vuota se comp è None."""
    if not comp:
        return []
    vista = comp.get("vista")
    if vista is None or vista.empty:
        return []

    conteggi = vista["Attendibilità"].value_counts()
    cards = "".join(
        f"""<div class="profile-card">
                <div class="profile-card-cat">{etichetta}</div>
                <div class="profile-card-t" style="color:{_comp_esito_colore(etichetta)}">
                    {int(conteggi.get(etichetta, 0))}</div>
            </div>"""
        for etichetta in (ESITO_REALE, ESITO_INCERTO, ESITO_STABILE, ESITO_ND)
    )

    sezioni = [f"""<section>
        <h2>Comparazione con lo storico</h2>
        <p class="intro-text">Il test è confrontato con la <b>media di
        {comp.get('n_sedute', 0)} sedute precedenti</b> (esclusa l'attuale).
        <b>Cambiamento</b> dice direzione ed entità dello scostamento;
        <b>Attendibilità</b> è un indice statistico per valutare quanto è credibile che un cambiamento ci sia stato davvero</p>
        <div class="profile-cards">{cards}</div>
        {_comp_table_html(vista)}
    </section>"""]

    # Strisce per le metriche scelte nella UI
    scelte = comp.get("strip") or []
    strisce = []
    for _, riga in vista[vista["_display"].isin(scelte)].iterrows():
        fig = build_swc_strip(riga)
        if fig is not None:
            strisce.append(_fig_div(fig, next_id("comp_strip")))
    if strisce:
        sezioni.append(f"""<section>
            <h2>Scostamento dalla media, metrica per metrica</h2>
            <p class="intro-text">Il rombo pieno azzurro è il test attuale con la
            sua barra di incertezza; il rombo vuoto arancione è la media delle
            sedute precedenti.</p>
            {''.join(strisce)}
        </section>""")

    # Profilo di forza: radar sovrapposto + tabella per categoria
    prof, serie, cats = comp.get("prof"), comp.get("prof_serie"), comp.get("prof_cats")
    if serie and cats:
        radar = build_profili_radar(cats, serie, nome_atleta)
        attuale = serie.get(f"{nome_atleta} — attuale", {})
        delta_map = ({r["Categoria"]: r["Delta T"] for _, r in prof.iterrows()}
                     if prof is not None and not prof.empty else {})
        cards = "".join(
            f"""<div class="profile-card">
                    <div class="profile-card-cat">{c}</div>
                    <div class="profile-card-qualita">{CATEGORY_QUALITY.get(c, '')}</div>
                    <div class="profile-card-t" style="color:{banda_da_tscore(attuale[c])[1]}">{attuale[c]:.0f}</div>
                    {_delta_t_html(delta_map.get(c))}
                    <div class="profile-card-banda">{banda_da_tscore(attuale[c])[0]}</div>
                </div>"""
            for c in cats if c in attuale
        )
        sezioni.append(f"""<section>
            <h2>Profilo di forza: attuale vs storico</h2>
            {_fig_div(radar, next_id('comp_radar')) if radar else ''}
            <div class="profile-cards">{cards}</div>
            {_prof_table_html(prof)}
        </section>""")

    # Indici di profilo
    indici = comp.get("indici") or []
    if indici:
        blocchi = []
        for voce in indici:
            fig = build_indice_zone_strip(
                voce["key"], voce["attuale"], voce["riferimento"], *voce["thr"])
            if fig is None:
                continue
            if voce["zona_att"] and voce["zona_rif"] and voce["zona_att"] != voce["zona_rif"]:
                nota = (f"cambio di zona: da <i>{voce['zona_rif']}</i> "
                        f"a <i>{voce['zona_att']}</i>")
            elif voce["zona_att"]:
                nota = f"zona invariata: <i>{voce['zona_att']}</i>"
            else:
                nota = ""
            blocchi.append(
                f"<h3>{voce['key'].upper()}</h3>"
                f'<p class="index-value"><b>{_fmt(voce["attuale"], 3)}</b>'
                f' <span class="muted">— {nota}</span></p>'
                f"{_fig_div(fig, next_id('comp_idx'))}"
            )
        if blocchi:
            sezioni.append(f"""<section>
                <h2>Indici di profilo</h2>
                {''.join(blocchi)}
            </section>""")

    return sezioni


# ----------------------------------------------------------------------------
# PDF
# ----------------------------------------------------------------------------
# Larghezze in mm: somma 180 = larghezza utile con margini 15/15. "Test" e le
# colonne di coda della tabella HTML sono omesse per stare in A4 verticale.
COMP_PDF_COLS = [
    ("Metrica", 40, lambda r: _pdf_comp(r.get("Metrica"))),
    ("UdM", 11, lambda r: _pdf_comp(r.get("Unità"))),
    ("Media rif.", 17, lambda r: _fmt(r.get("Media rif."), 3)),
    ("Attuale", 17, lambda r: _fmt(r.get("Attuale"), 3)),
    ("Delta", 17, lambda r: _fmt(r.get("Delta"), 3, True)),
    ("Cambiam.", 22, lambda r: _pdf_comp(r.get("Cambiamento"))),
    ("Attendib.", 25, lambda r: _pdf_comp(r.get("Attendibilità"))),
    ("Migliore", 17, lambda r: _fmt(r.get("Migliore"), 3)),
    ("% migl.", 14, lambda r: _fmt(r.get("% del migliore"), 0, False, "%")),
]

PROF_PDF_COLS = [
    ("Categoria", 48, lambda r: _pdf_comp(r.get("Categoria"))),
    ("T rif.", 15, lambda r: f"{r.get('T medio rif.'):.1f}"),
    ("T att.", 15, lambda r: f"{r.get('T attuale'):.1f}"),
    ("Delta T", 17, lambda r: f"{r.get('Delta T'):+.1f}"),
    ("Cambiam.", 22, lambda r: _pdf_comp(r.get("Cambiamento"))),
    ("Attendib.", 25, lambda r: _pdf_comp(r.get("Attendibilità"))),
    ("Valutazione", 38, lambda r: _pdf_comp(r.get("Valutazione attuale"))),
]


def _pdf_comp_table(pdf, df, colonne, col_esito="Attendibilità"):
    """Tabella PDF generica per la Comparazione. La cella di attendibilità
    viene riempita col colore del semaforo, perché nel PDF le emoji non
    esistono e il colore è l'unico segnale visivo che resta."""
    if df is None or df.empty:
        return
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_draw_color(200, 200, 200)
    pdf.set_line_width(0.2)
    pdf.set_fill_color(255, 255, 255)
    pdf.set_text_color(30, 30, 30)
    pdf.ensure_space(12)
    with pdf.table(
        col_widths=[w for _h, w, _f in colonne], text_align="LEFT", line_height=4.6,
        headings_style=FontFace(emphasis="BOLD", color=(255, 255, 255),
                                fill_color=_hex_to_rgb(TEXT_COLOR)),
    ) as table:
        riga = table.row()
        for intestazione, _w, _f in colonne:
            riga.cell(intestazione)
        for _, r in df.iterrows():
            riga = table.row()
            for intestazione, _w, estrai in colonne:
                try:
                    testo = estrai(r)
                except (TypeError, ValueError):
                    testo = "-"
                stile = None
                if intestazione.startswith("Attendib"):
                    stile = FontFace(color=(255, 255, 255),
                                     fill_color=_hex_to_rgb(_comp_esito_colore(r.get(col_esito))))
                riga.cell(testo, style=stile)
    pdf.set_x(pdf.l_margin)


def comparazione_sezione_pdf(pdf, comp, nome_atleta):
    """Aggiunge al PDF la sezione Comparazione. No-op se comp è None."""
    if not comp:
        return
    vista = comp.get("vista")
    if vista is None or vista.empty:
        return
    pdf.section_title("Comparazione con lo storico")
    pdf.body_text(
        f"Il test e confrontato con la media di {comp.get('n_sedute', 0)} sedute "
        "precedenti (esclusa l'attuale). La colonna Cambiamento indica direzione "
        "(+ in aumento, - in diminuzione) ed entita dello scostamento; "
        "Attendibilita' e' un indice statistico per valutare quanto e' credibile "
        "che un cambiamento ci sia stato davvero"
    )
    conteggi = vista["Attendibilità"].value_counts()
    pdf.body_text("   ".join(
        f"{_pdf_comp(e)}: {int(conteggi.get(e, 0))}"
        for e in (ESITO_REALE, ESITO_INCERTO, ESITO_STABILE, ESITO_ND)
    ), size=9)
    pdf.ln(1)
    _pdf_comp_table(pdf, vista, COMP_PDF_COLS)

    # Strisce: height_px DEVE coincidere con l'altezza della figura, altrimenti
    # kaleido rende la fascia con uno spessore diverso da quello previsto.
    scelte = comp.get("strip") or []
    strisce = vista[vista["_display"].isin(scelte)]
    if not strisce.empty:
        pdf.ln(2)
        pdf.subsection_title("Scostamento dalla media, metrica per metrica")
        pdf.body_text(
            "Rombo pieno = test attuale con barra di incertezza; rombo vuoto = "
            "media delle sedute precedenti.", size=9)
        for _, riga in strisce.iterrows():
            fig = build_swc_strip(riga)
            if fig is not None:
                pdf.chart_image(fig, width_px=900, height_px=fig.layout.height,
                                content_width_mm=165)

    prof, serie, cats = comp.get("prof"), comp.get("prof_serie"), comp.get("prof_cats")
    if serie and cats:
        pdf.add_page()
        pdf.section_title("Profilo di forza: attuale vs storico")
        radar = build_profili_radar(cats, serie, nome_atleta)
        if radar:
            pdf.chart_image(radar, width_px=1000, height_px=620, content_width_mm=150)
        attuale = serie.get(f"{nome_atleta} — attuale", {})
        if attuale:
            delta_map = ({r["Categoria"]: r["Delta T"] for _, r in prof.iterrows()}
                         if prof is not None and not prof.empty else {})
            pdf.profile_cards([c for c in cats if c in attuale], attuale, delta_map)
        pdf.body_text(
            "Il T-score e gia orientato alla prestazione, quindi qui un valore "
            "piu alto e sempre migliore. 1 soglia di rilevanza = 2 punti di T-score.",
            size=9)
        _pdf_comp_table(pdf, prof, PROF_PDF_COLS)

    indici = comp.get("indici") or []
    if indici:
        pdf.ln(3)
        pdf.subsection_title("Indici di profilo: si e spostato di zona?")
        for voce in indici:
            fig = build_indice_zone_strip(
                voce["key"], voce["attuale"], voce["riferimento"], *voce["thr"])
            if fig is None:
                continue
            if voce["zona_att"] and voce["zona_rif"] and voce["zona_att"] != voce["zona_rif"]:
                nota = f"cambio di zona: da {voce['zona_rif']} a {voce['zona_att']}"
            elif voce["zona_att"]:
                nota = f"zona invariata: {voce['zona_att']}"
            else:
                nota = ""
            pdf.body_text(f"{voce['key'].upper()}: {_fmt(voce['attuale'], 3)}   {nota}",
                          size=9)
            pdf.chart_image(fig, width_px=900, height_px=fig.layout.height,
                            content_width_mm=150)


def genera_report_html(nome, sesso, periodo, results, profilo, commento, thresholds,
                       comp=None):
    div_counter = {"n": 0}

    def next_id(prefix):
        div_counter["n"] += 1
        return f"{prefix}_{div_counter['n']}"

    sections = []

    # --- Profilo di Forza ---
    cats_valide = [c for c in CATEGORIES if profilo.get(c) is not None]
    radar_fig = build_radar_chart(cats_valide, profilo, nome)
    profilo_cards = "".join(
        f"""<div class="profile-card">
                <div class="profile-card-cat">{cat}</div>
                <div class="profile-card-qualita">{CATEGORY_QUALITY.get(cat, '')}</div>
                <div class="profile-card-t" style="color:{banda_da_tscore(profilo[cat])[1]}">{profilo[cat]:.0f}</div>
                <div class="profile-card-banda">{banda_da_tscore(profilo[cat])[0]}</div>
            </div>"""
        for cat in cats_valide
    )
    sections.append(f"""<section>
        <h2>Profilo di Forza</h2>
        {_fig_div(radar_fig, next_id('radar')) if radar_fig else '<p class="muted">Nessuna metrica con confronto di popolazione disponibile.</p>'}
        <div class="profile-cards">{profilo_cards}</div>
    </section>""")

    # --- Categorie di test ---
    # I controlli tecnici a soglia (CMJ Rebound) restano solo nella scheda
    # live "Dettaglio Test": non vengono replicati nel report scaricabile.
    for cat in CATEGORIES:
        cat_results = [r for r in results if r["category"] == cat and r["mean"] is not None]
        if not cat_results:
            continue
        tscore_fig = build_tscore_bar_chart(cat_results)
        rsq_fig = build_rsq_chart(cat, results)

        sections.append(f"""<section>
            <h2>{cat}</h2>
            {_fig_div(tscore_fig, next_id('tscore'))}
            {_metric_table_html(cat_results)}
            {_metric_descriptions_html(cat_results)}
            {f'<h3>RSQ — Reactive Strength Quadrant</h3>{_fig_div(rsq_fig, next_id("rsq"))}' if rsq_fig else ''}
        </section>""")

    # --- Indici (DSI, EUR): barra a bande + grafico a spicchi, con il valore
    # numerico e le soglie usate accanto a ciascun grafico. Nessun T-score:
    # sono indici diagnostici, non metriche "più alto = meglio".
    r_dsi = next((r for r in results if r["key"] == "dsi" and r["mean"] is not None), None)
    r_eur = next((r for r in results if r["key"] == "eur" and r["mean"] is not None), None)
    dsi_fig = build_dsi_chart(results, thresholds)
    eur_fig = build_eur_chart(results, thresholds)
    if r_dsi or r_eur or dsi_fig or eur_fig:
        indici_html = []
        for r_idx, fig_idx, titolo, idx_key in (
            (r_dsi, dsi_fig, "DSI (Dynamic Strength Index)", "dsi"),
            (r_eur, eur_fig, "EUR (Eccentric Utilisation Ratio)", "eur"),
        ):
            if not (r_idx or fig_idx):
                continue
            strip = ratio_band_strip(idx_key, r_idx["mean"], *thresholds[idx_key],
                                     label=idx_key.upper(), decimals=3) if r_idx else None
            indici_html.append(
                f"<h3>{titolo}</h3>{_index_value_html(r_idx)}"
                f"{_fig_div(strip, next_id(idx_key + '_strip'))}"
                f"{_fig_div(fig_idx, next_id(idx_key))}"
            )
        sections.append(f"""<section>
            <h2>Indici</h2>
            {''.join(indici_html)}
            {_metric_descriptions_html([r_dsi, r_eur])}
        </section>""")

    # --- Comparazione con lo storico ---
    # Sezione OPZIONALE: comparazione_sections_html() restituisce una lista
    # vuota se comp è None, cioè se la scheda 🔀 Comparazione non è stata
    # usata in questa sessione. Va prima dell'Analisi, che chiude il report.
    sections.extend(comparazione_sections_html(comp, nome, next_id))

    # --- Analisi: testo scritto dal preparatore nella scheda Report
    # dell'app, incluso qui come testo statico (non modificabile nel report).
    commento_html = _html.escape(commento).replace("\n", "<br>") if commento and commento.strip() else (
        "Nessuna analisi inserita."
    )
    sections.append(f"""<section>
        <h2>Analisi</h2>
        <div class="analysis-box">{commento_html}</div>
    </section>""")

    html = f"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="utf-8">
<title>Force Plate Test Report — {nome}</title>
<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>
<style>
    :root {{ --primary: {PRIMARY}; --accent: {ACCENT}; --text: {TEXT_COLOR}; --bg: {BG_COLOR}; }}
    * {{ box-sizing: border-box; }}
    body {{
        font-family: "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
        background: var(--bg); color: var(--text); margin: 0; padding: 0 0 60px 0;
    }}
    .report {{ max-width: 960px; margin: 0 auto; padding: 0 24px; }}
    header.report-header {{ background: var(--text); color: #fff; padding: 32px 24px; margin-bottom: 24px; }}
    header.report-header h1 {{ margin: 0 0 12px 0; font-size: 26px; letter-spacing: 0.5px; }}
    .meta-row span {{ margin-right: 24px; font-size: 15px; }}
    .meta-row b {{ color: var(--primary); }}
    section {{ margin-bottom: 36px; padding-bottom: 24px; border-bottom: 1px solid #e0e0e0; }}
    h2 {{ color: var(--text); border-left: 5px solid var(--accent); padding-left: 10px; font-size: 20px; }}
    h3 {{ color: var(--text); font-size: 16px; margin-top: 24px; }}
    .intro-text {{ line-height: 1.5; }}
    table.report-table {{ width: 100%; border-collapse: collapse; margin-top: 12px; font-size: 14px; }}
    table.report-table th {{ background: var(--text); color: #fff; text-align: left; padding: 8px 10px; font-weight: 600; }}
    table.report-table td {{ padding: 8px 10px; border-bottom: 1px solid #eee; }}
    table.report-table tr:nth-child(even) td {{ background: #f7fbfd; }}
    .badge {{ color: #fff; padding: 3px 10px; border-radius: 12px; font-size: 12px; font-weight: 600; display: inline-block; }}
    .profile-cards {{ display: flex; flex-wrap: wrap; gap: 16px; margin-top: 16px; }}
    .profile-card {{ flex: 1 1 200px; background: #f7fbfd; border-radius: 8px; padding: 14px 16px; border: 1px solid #e0e0e0; }}
    .profile-card-cat {{ font-size: 12px; text-transform: uppercase; letter-spacing: 0.5px; color: var(--text); opacity: 0.7; }}
    .profile-card-qualita {{ font-size: 17px; font-weight: 600; color: var(--text); margin: 2px 0 4px 0; }}
    .profile-card-t {{ font-size: 30px; font-weight: 700; }}
    .profile-card-banda {{ font-size: 13px; }}
    .profile-card-delta {{ font-size: 15px; font-weight: 600; }}
    .muted {{ color: #667; font-size: 13px; }}
    .index-value {{ font-size: 16px; margin: 6px 0 14px 0; }}
    .metric-help {{ margin: 14px 0 4px 0; font-size: 14px; }}
    .metric-help summary {{ cursor: pointer; color: var(--primary); font-weight: 600; }}
    .metric-help ul {{ margin: 10px 0 0 0; padding-left: 20px; }}
    .metric-help li {{ margin-bottom: 6px; line-height: 1.4; }}
    .analysis-box {{ min-height: 60px; border: 1px solid #e0e0e0; border-left: 4px solid var(--primary); border-radius: 6px; padding: 14px; line-height: 1.6; background: #f7fbfd; }}
    .print-btn {{
        background: var(--accent); color: #fff; border: none; border-radius: 6px;
        padding: 10px 20px; font-size: 15px; font-weight: 600; cursor: pointer;
        margin-top: 18px;
    }}
    .print-btn:hover {{ opacity: 0.9; }}
    @media print {{
        body {{ background: #fff; }}
        .no-print {{ display: none !important; }}
        section {{ break-inside: avoid; page-break-inside: avoid; }}
        .js-plotly-plot .modebar {{ display: none !important; }}
        details.metric-help {{ break-inside: avoid; }}
    }}
</style>
</head>
<body>
    <div class="report">
        <header class="report-header">
            <h1>🏋️ FORCE PLATE TEST REPORT</h1>
            <div class="meta-row">
                <span><b>Atleta:</b> {nome}</span>
                <span><b>Sesso:</b> {sesso}</span>
                <span><b>Data test:</b> {periodo}</span>
            </div>
            <button class="print-btn no-print" onclick="window.print()">🖨️ Stampa / Salva come PDF</button>
        </header>
        <p class="intro-text">
            Per valutare i risultati del test è stato utilizzato il T-Score, un indice standardizzato
            che confronta la prestazione dell'atleta rispetto a un gruppo di riferimento, esprimendo la
            distanza dalla media in deviazioni standard. Punteggi tra 0 e 50 indicano valori inferiori
            alla media, mentre punteggi tra 50 e 100 indicano valori superiori alla media.
        </p>
        {''.join(sections)}
    </div>
</body>
</html>"""
    return html.encode("utf-8")


# ============================================================================
# PARTE 6bis — REPORT SCARICABILE (PDF statico)
# ============================================================================
# Alternativa "un tap e via" all'HTML per chi ha problemi ad aprire report
# interattivi su iOS (Quick Look non esegue JavaScript). I grafici Plotly
# vengono renderizzati come immagini statiche via kaleido (offline, nessuna
# richiesta di rete) e assemblati in un PDF vero con fpdf2 (puro Python,
# nessuna libreria di sistema richiesta — a differenza di weasyprint).

_PDF_CHAR_REPLACEMENTS = {
    "\u2014": " - ",  # em dash
    "\u2013": "-",    # en dash
    "\u2192": "->",   # freccia destra
    "\u2018": "'", "\u2019": "'",
    "\u201c": '"', "\u201d": '"',
    "\u2022": "-",    # bullet
}


def _pdf_safe(text):
    """I font core di fpdf2 (Helvetica) coprono solo Latin-1: sostituisce i
    caratteri tipografici comuni (em dash, frecce, virgolette curve) con
    equivalenti ASCII e scarta silenziosamente qualsiasi altro carattere
    fuori range (es. emoji), invece di far fallire l'export."""
    if text is None:
        return ""
    text = str(text)
    for old, new in _PDF_CHAR_REPLACEMENTS.items():
        text = text.replace(old, new)
    return "".join(ch if ord(ch) <= 255 else "" for ch in text)


def _hex_to_rgb(hex_color):
    if not hex_color:
        return (230, 230, 230)
    h = hex_color.lstrip("#")
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def _fig_png_bytes(fig, width=1000, height=560, scale=2):
    """Renderizza una figura Plotly in PNG (bytes) via kaleido, offline.
    Ritorna (BytesIO, aspect_ratio) o None se fig è None."""
    if fig is None:
        return None
    png_bytes = fig.to_image(format="png", width=width, height=height, scale=scale)
    return io.BytesIO(png_bytes), height / width


class _ReportPDF(FPDF):
    def __init__(self):
        super().__init__(format="A4")
        self.set_auto_page_break(auto=True, margin=16)
        self.set_margins(15, 15, 15)

    def ensure_space(self, height_mm):
        """Forza un salto pagina se l'elemento successivo (altezza nota in
        mm) non entra nello spazio rimasto nella pagina corrente."""
        if self.get_y() + height_mm > self.page_break_trigger:
            self.add_page()

    def section_title(self, text):
        self.ensure_space(14)
        self.set_x(self.l_margin)
        self.set_font("Helvetica", "B", 14)
        self.set_text_color(*_hex_to_rgb(TEXT_COLOR))
        self.cell(0, 10, _pdf_safe(text), new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(*_hex_to_rgb(ACCENT))
        self.set_line_width(0.8)
        self.line(self.l_margin, self.get_y(), self.l_margin + 6, self.get_y())
        self.ln(3)

    def subsection_title(self, text):
        self.ensure_space(10)
        self.set_x(self.l_margin)
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(*_hex_to_rgb(TEXT_COLOR))
        self.cell(0, 8, _pdf_safe(text), new_x="LMARGIN", new_y="NEXT")

    def body_text(self, text, size=10):
        self.set_x(self.l_margin)
        self.set_font("Helvetica", "", size)
        self.set_text_color(60, 60, 70)
        self.multi_cell(0, 5.5, _pdf_safe(text))

    def chart_image(self, fig, width_px=1000, height_px=560, content_width_mm=180):
        result = _fig_png_bytes(fig, width=width_px, height=height_px)
        if result is None:
            return
        img_io, aspect = result
        h_mm = content_width_mm * aspect
        self.ensure_space(h_mm + 4)
        x = self.l_margin + (self.epw - content_width_mm) / 2
        self.image(img_io, x=x, w=content_width_mm)
        self.set_x(self.l_margin)
        self.ln(4)

    def metric_table(self, cat_results):
        headers = ["Metrica", "Unita", "Media", "Media Pop.", "N", "T-score", "Valutazione"]
        widths = [52, 14, 18, 18, 10, 16, 35]  # mm assoluti, sommano ~163mm su 180mm utili
        self.set_x(self.l_margin)
        self.set_font("Helvetica", "", 8.5)
        self.set_draw_color(200, 200, 200)  # i bordi/il fill della tabella non devono
        self.set_line_width(0.2)            # ereditare accent/testo/fill lasciati da altri
        self.set_fill_color(255, 255, 255)  # elementi (es. il rettangolo blu dell'header,
        self.set_text_color(30, 30, 30)     # o il testo blu del blocco Profilo di Forza)
        self.ensure_space(10)
        with self.table(
            col_widths=widths, text_align="LEFT", line_height=5,
            headings_style=FontFace(emphasis="BOLD", color=(255, 255, 255), fill_color=_hex_to_rgb(TEXT_COLOR)),
        ) as table:
            row = table.row()
            for h in headers:
                row.cell(h)
            for r in cat_results:
                row = table.row()
                row.cell(_pdf_safe(r["label"]))
                row.cell(_pdf_safe(r["unit"] or ""))
                row.cell(f"{r['mean']:.3f}" if r["mean"] is not None else "N/D")
                row.cell(f"{r['pop_mean']:.3f}" if r["pop_mean"] is not None else "-")
                row.cell(str(r["n"]))
                row.cell(f"{r['t']:.0f}" if r["t"] is not None else "-")
                style = None
                if r.get("colore"):
                    style = FontFace(fill_color=_hex_to_rgb(r["colore"]))
                row.cell(_pdf_safe(r["banda"] or "-"), style=style)
        # fpdf2 può lasciare il cursore x dopo l'ultima cella dell'ultima riga
        # invece di riportarlo al margine sinistro: senza questo reset, un
        # multi_cell(0, ...) subito dopo la tabella può ricevere una
        # larghezza quasi nulla e sollevare FPDFException.
        self.set_x(self.l_margin)

    def profile_cards(self, cats, profilo, delta_map=None):
        """Riquadri affiancati per il Profilo di Forza, equivalenti alle
        .profile-card del report HTML: nome del protocollo, qualità fisica
        misurata, T-score grande nel colore della banda, lo scostamento in
        punti T (solo nella sezione Comparazione) e infine l'etichetta della
        banda di valutazione.

        delta_map è opzionale: senza, la card è quella del report base e
        l'altezza scende da 36 a 31 mm, altrimenti resterebbe un buco bianco
        sotto la banda. Le coordinate sono assolute, quindi se cambi `h`
        devi spostare anche tutti gli offset y0 + ... qui sotto."""
        if not cats:
            return
        n = len(cats)
        con_delta = bool(delta_map)
        gap = 4
        h = 36 if con_delta else 31
        w = (self.epw - gap * (n - 1)) / n
        self.ensure_space(h + 6)
        y0 = self.get_y()
        for i, cat in enumerate(cats):
            t = profilo[cat]
            banda, colore = banda_da_tscore(t)
            x = self.l_margin + i * (w + gap)

            self.set_fill_color(247, 251, 253)
            self.set_draw_color(224, 224, 224)
            self.set_line_width(0.2)
            self.rect(x, y0, w, h, style="DF")

            # Nome del protocollo: piccolo e grigio, e' il "come"
            self.set_xy(x + 2, y0 + 2.5)
            self.set_font("Helvetica", "B", 6.5)
            self.set_text_color(120, 145, 160)
            self.multi_cell(w - 4, 3.2, _pdf_safe(cat), align="C")

            # Qualita' fisica: e' il "cosa", quindi in evidenza
            self.set_xy(x + 2, y0 + 9.4)
            self.set_font("Helvetica", "B", 9)
            self.set_text_color(*_hex_to_rgb(TEXT_COLOR))
            self.multi_cell(w - 4, 4, _pdf_safe(CATEGORY_QUALITY.get(cat, "")), align="C")

            # T-score nel colore della banda
            self.set_xy(x, y0 + 14.5)
            self.set_font("Helvetica", "B", 20)
            self.set_text_color(*_hex_to_rgb(colore))
            self.cell(w, 9, f"{t:.0f}", align="C")

            # Scostamento in punti T: sta SOPRA la banda, che quindi scende di
            # 5 mm. Senza delta la banda risale al suo posto originale.
            d = (delta_map or {}).get(cat)
            y_banda = y0 + 24
            if d is not None:
                self.set_xy(x + 2, y0 + 23.5)
                self.set_font("Helvetica", "B", 8)
                self.set_text_color(*_hex_to_rgb(colore_delta_t(d)))
                self.multi_cell(w - 4, 3.6, _pdf_safe(f"{d:+.1f}"), align="C")
                y_banda = y0 + 28.5

            self.set_xy(x + 2, y_banda)
            self.set_font("Helvetica", "", 6.5)
            self.set_text_color(90, 90, 100)
            self.multi_cell(w - 4, 3.2, _pdf_safe(banda or "-"), align="C")

        # Reset esplicito: fill/testo/bordo restano impostati sull'ultimo
        # riquadro disegnato e verrebbero ereditati dagli elementi seguenti.
        self.set_xy(self.l_margin, y0 + h)
        self.set_fill_color(255, 255, 255)
        self.set_text_color(30, 30, 30)
        self.set_draw_color(200, 200, 200)
        self.ln(5)

    def descriptions_block(self, items):
        pairs = [(it["label"], it["desc"]) for it in items if it and it.get("desc")]
        if not pairs:
            return
        self.ensure_space(8)
        self.set_x(self.l_margin)
        self.set_font("Helvetica", "I", 8.5)
        self.set_text_color(90, 90, 100)
        for label, desc in pairs:
            self.ensure_space(6)
            self.set_x(self.l_margin)
            self.multi_cell(0, 4.5, _pdf_safe(f"- {label}: {desc}"))
        self.ln(1)


def genera_report_pdf(nome, sesso, periodo, results, profilo, commento, thresholds,
                      comp=None):
    pdf = _ReportPDF()
    pdf.add_page()

    # --- Header ---
    pdf.set_fill_color(*_hex_to_rgb(TEXT_COLOR))
    pdf.rect(0, 0, 210, 26, style="F")
    pdf.set_xy(15, 7)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 17)
    pdf.cell(0, 8, "FORCE PLATE TEST REPORT", new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(15)
    pdf.set_font("Helvetica", "", 10.5)
    pdf.cell(0, 6, _pdf_safe(f"Atleta: {nome}   Sesso: {sesso}   Data test: {periodo}"),
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_y(32)

    pdf.body_text(
        "Per valutare i risultati del test e stato utilizzato il T-Score, un indice standardizzato "
        "che confronta la prestazione dell'atleta rispetto a un gruppo di riferimento, esprimendo la "
        "distanza dalla media in deviazioni standard. Punteggi tra 0 e 50 indicano valori inferiori "
        "alla media, mentre punteggi tra 50 e 100 indicano valori superiori alla media. DSI ed EUR "
        "fanno eccezione: essendo rapporti tra due test, non vengono letti come 'piu alto = meglio' "
        "ma per zona di profilo, rispetto alle soglie indicate accanto a ciascun grafico."
    )
    pdf.ln(3)

    # --- Profilo di Forza ---
    cats_valide = [c for c in CATEGORIES if profilo.get(c) is not None]
    pdf.section_title("Profilo di Forza")
    radar_fig = build_radar_chart(cats_valide, profilo, nome)
    if radar_fig:
        pdf.chart_image(radar_fig, width_px=1000, height_px=620, content_width_mm=150)
    else:
        pdf.body_text("Nessuna metrica con confronto di popolazione disponibile.")
    pdf.profile_cards(cats_valide, profilo)

    pdf.add_page()

    # --- Categorie di test ---
    for cat in CATEGORIES:
        cat_results = [r for r in results if r["category"] == cat and r["mean"] is not None]
        if not cat_results:
            continue
        pdf.section_title(cat)
        tscore_fig = build_tscore_bar_chart(cat_results)
        if tscore_fig:
            n_metrics = len([r for r in cat_results if r["t"] is not None])
            h_px = max(360, 70 * n_metrics + 150)
            pdf.chart_image(tscore_fig, width_px=1000, height_px=h_px, content_width_mm=180)
        pdf.metric_table(cat_results)
        pdf.ln(1)
        rsq_fig = build_rsq_chart(cat, results)
        if rsq_fig:
            pdf.subsection_title("RSQ - Reactive Strength Quadrant")
            pdf.chart_image(rsq_fig, width_px=900, height_px=430, content_width_mm=150)
        pdf.descriptions_block(cat_results)
        pdf.add_page()

    # --- Indici (DSI, EUR) ---
    r_dsi = next((r for r in results if r["key"] == "dsi" and r["mean"] is not None), None)
    r_eur = next((r for r in results if r["key"] == "eur" and r["mean"] is not None), None)
    dsi_fig = build_dsi_chart(results, thresholds)
    eur_fig = build_eur_chart(results, thresholds)
    if r_dsi or r_eur or dsi_fig or eur_fig:
        pdf.section_title("Indici")
        for r_idx, fig_idx, titolo, idx_key in (
            (r_dsi, dsi_fig, "DSI (Dynamic Strength Index)", "dsi"),
            (r_eur, eur_fig, "EUR (Eccentric Utilisation Ratio)", "eur"),
        ):
            if not (r_idx or fig_idx):
                continue
            pdf.subsection_title(titolo)
            if r_idx:
                lo, hi = thresholds[idx_key]
                zona = f"  ({r_idx['zona']})" if r_idx.get("zona") else ""
                pop = f"   Media pop. {r_idx['pop_mean']:.3f}" if r_idx["pop_mean"] is not None else ""
                pdf.body_text(f"{idx_key.upper()}: {r_idx['mean']:.3f}{zona}{pop}   "
                              f"soglie {lo:.2f} / {hi:.2f}")
                # height della figura e height_px di kaleido DEVONO coincidere:
                # lo spessore della fascia e' (height - margini) dell'altezza
                # effettivamente renderizzata, non di quella dichiarata nella
                # figura. Se divergono, in PDF la fascia diventa un blocco.
                strip = ratio_band_strip(idx_key, r_idx["mean"], lo, hi,
                                         label=idx_key.upper(), decimals=3)
                if strip:
                    pdf.chart_image(strip, width_px=900, height_px=strip.layout.height,
                                    content_width_mm=150)
            if fig_idx:
                pdf.chart_image(fig_idx, width_px=900, height_px=470, content_width_mm=150)
            pdf.descriptions_block([r_idx])
            pdf.add_page()

    # --- Comparazione con lo storico (sezione opzionale) ---
    comparazione_sezione_pdf(pdf, comp, nome)

    # --- Analisi ---
    # add_page() perche' la sezione Comparazione lascia la pagina a metà.
    pdf.add_page()
    pdf.section_title("Analisi")
    testo = commento.strip() if commento and commento.strip() else "Nessuna analisi inserita."
    pdf.body_text(testo)

    return bytes(pdf.output())


with tab_report:
    if not parsed_files:
        st.info("Carica dei file dalla barra laterale per generare il report.")
    else:
        st.markdown(
            "Scegli in che formato generare il report: **HTML interattivo** (grafici zoomabili, pensato per "
            "desktop/laptop) oppure **PDF statico** (più facilmente visualizzabile da qualsiasi dispositivo)."
        )
        st.caption(
            "Se hai usato la scheda 🔀 Comparazione in questa sessione, il report include anche "
            "il confronto con lo storico: tabella, strisce delle metriche selezionate e profilo "
            "di forza attuale vs storico."
        )
        st.markdown("**Analisi del preparatore**")
        st.caption(
            "Scrivi qui il commento tecnico da includere nel report: punti di forza, aree di "
            "miglioramento e indicazioni di lavoro."
        )
        commento = st.text_area(
            "Analisi del preparatore", key="coach_comment", height=160, label_visibility="collapsed",
            placeholder="Es. Buoni valori di forza isometrica, mRSI sopra media. Da lavorare sulla "
                        "reattività nel CMJ Rebound...",
        )
        col_html, col_pdf = st.columns(2)
        with col_html:
            if st.button("📄 Genera report HTML", type="primary", use_container_width=True):
                html_bytes = genera_report_html(nome, sesso, periodo, results, profilo, commento,
                                                st.session_state["idx_thr"],
                                                comp=st.session_state.get("comp_export"))
                st.download_button(
                    "⬇️ Scarica report (.html)", data=html_bytes,
                    file_name=f"Report_{nome.replace(' ', '_')}_{dt.date.today().isoformat()}.html",
                    mime="text/html", use_container_width=True,
                )
        with col_pdf:
            if st.button("📕 Genera report PDF", use_container_width=True):
                with st.spinner("Rendering grafici e impaginazione PDF..."):
                    pdf_bytes = genera_report_pdf(nome, sesso, periodo, results, profilo, commento,
                                                  st.session_state["idx_thr"],
                                                  comp=st.session_state.get("comp_export"))
                st.download_button(
                    "⬇️ Scarica report (.pdf)", data=pdf_bytes,
                    file_name=f"Report_{nome.replace(' ', '_')}_{dt.date.today().isoformat()}.pdf",
                    mime="application/pdf", use_container_width=True,
                )